"""
Test Stabilization Tracker - Local Web App
Backend: Flask + SQLite
Run: python app.py
Then open http://localhost:5000 (or http://<your-lan-ip>:5000 for teammates)
"""
import sqlite3
import io
import csv
import re
import functools
import json
import os
import shutil
import threading
import time
import urllib.request
import urllib.error
import urllib.parse
from datetime import datetime, date, timedelta
from flask import Flask, request, jsonify, render_template, send_file, g, session, redirect
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, BarChart, PieChart, Reference

DB_PATH = "tracker.db"
USERS_DB_PATH = "users.db"  # DB rieng cho tai khoan - tach khoi tracker.db (du lieu song)
DEFAULT_MODELS = ["S721", "A175", "F966", "X526", "F741"]
PASS_STATES = {"pass", "check", "na", "manual check"}  # MANUAL CHECK tương tự CHECK = Pass
SKIP_STATES = {"running"}  # test chua chay xong, khong dua vao thong ke Pass/Fail
# TIMEOUT = Fail (không có trong PASS_STATES → quy vào Fail)
# Mat khau mac dinh khi admin reset tai khoan (hoac tao sap dat khi them owner moi)
DEFAULT_RESET_PASSWORD = "abc123"

# ------------------------------------------------------------------
# Phan quyen theo tung tab chuc nang (key = data-tab trong index.html)
# ------------------------------------------------------------------
ALL_TABS = [
    "dashboard", "new-scripts", "input-results", "input-fix",
    "priority", "cycle-compare", "fix-tracking", "reports", "integrations", "settings",
]
# Quyen mac dinh cho user thuong: moi tab TRU Nhap ket qua, Dong bo & Cai dat.
USER_DEFAULT_TABS = [
    "dashboard", "new-scripts", "input-fix", "priority", "cycle-compare", "fix-tracking", "reports",
]
NS_EXTRA_PERMS = ["ns-assign", "ns-edit"]  # quyen rieng cho tab new-scripts, khong phai tab
ALL_PERMS = ALL_TABS + NS_EXTRA_PERMS
ROLE_DEFAULT_PERMS = {
    "admin": list(ALL_TABS) + list(NS_EXTRA_PERMS),
    "moderator": list(ALL_TABS) + list(NS_EXTRA_PERMS),  # moi tab ke ca Cai dat; khac admin o cho khong quan tri tai khoan
    "user": list(USER_DEFAULT_TABS),
}
BOOTSTRAP_ADMIN = "anh.hh"  # tai khoan admin khoi tao mac dinh

# Nhom nguyen nhan goc CHUAN cho form Ghi nhan Fix (dropdown) - theo chien luoc quan ly.
# "Infra/Device" = loi moi truong/farm, KHONG phai loi script -> tach rieng trong bao cao.
ROOT_CAUSE_GROUPS = [
    "Locator/UI change", "Timing/Sync", "Test data",
    "Infra/Device", "App bug", "Script logic", "Khác",
]

# Cac setting nhay cam (token/API key): GET /api/settings tra ve mask, POST bo qua gia tri mask.
SENSITIVE_SETTINGS = {"farm_api_token", "github_token", "import_token", "company_cookie"}
SETTINGS_MASK = "********"

# Trang thai ben TC Hub duoc coi la "da hoan thanh script" / "loai tru" (so sanh lowercase).
# "excluded" = vocab TC Hub that; "skip" = giu tuong thich voi vocab paste tay cu.
COMPANY_PERFORMED_STATES = {"performed"}
COMPANY_SKIP_STATES = {"skip", "excluded"}
# "target" (automationTarget) = TC CAN hoan thanh script nhung CHUA xong (khac performed/excluded).
COMPANY_TARGET_STATES = {"target"}

BACKUP_DIR = "backups"


# Keyword heuristic de suy NHOM nguyen nhan tu root_cause free text (fix cu / admin bulk
# import khong co dropdown). Uu tien prefix "<Group> - " (convention team da dung), roi
# keyword; khong khop gi -> "Khác".
_ROOT_CAUSE_GROUP_KEYWORDS = [
    ("Locator/UI change", r"locator|xpath|element|resource[- ]?id|\bui\b|id đổi|doi id|selector"),
    ("Timing/Sync", r"wait|timing|sleep|time\s*out|timeout|sync|race|delay|cho\b|chậm"),
    ("Test data", r"test\s*data|du lieu|dữ liệu|account|tai khoan|tài khoản|het han|hết hạn|data\b"),
    ("Infra/Device", r"infra|device|farm|wifi|\bstp\b|mang\b|mạng|network|treo|disconnect|not free|adb"),
    ("App bug", r"app\s*bug|loi app|lỗi app|issue app|bug app|app issue"),
    ("Script logic", r"script|logic|import|exception|traceback|code|hàm|ham\b"),
]


def classify_root_cause_group(text):
    """Suy nhom nguyen nhan chuan (ROOT_CAUSE_GROUPS) tu root_cause free text."""
    s = str(text or "").strip()
    if not s:
        return "Khác"
    low = s.lower()
    # 1) Prefix "<Group> - chi tiet" (hoac "<Group>:") theo convention docs
    for grp in ROOT_CAUSE_GROUPS:
        g = grp.lower()
        if low == g or low.startswith(g + " -") or low.startswith(g + "-") or low.startswith(g + ":"):
            return grp
    # 2) Keyword heuristic
    for grp, pat in _ROOT_CAUSE_GROUP_KEYWORDS:
        if re.search(pat, low, re.I):
            return grp
    return "Khác"


def get_root_cause_groups(db):
    """Danh sach nhom nguyen nhan HIEN HANH: doc tu settings key 'root_cause_groups'
    (JSON list) neu admin da doi ten; else tra ve hang so mac dinh ROOT_CAUSE_GROUPS.
    Cho phep doi ten nhom ma khong sua code (persist trong DB)."""
    raw = get_setting(db, "root_cause_groups", "").strip()
    if raw:
        try:
            groups = json.loads(raw)
            if isinstance(groups, list) and all(isinstance(g, str) and g.strip() for g in groups) and groups:
                return [g.strip() for g in groups]
        except (ValueError, TypeError):
            pass
    return list(ROOT_CAUSE_GROUPS)


app = Flask(__name__)
# Session cookie (dang nhap). Hardcode cung mo hinh secret hien co (LAN noi bo).
app.secret_key = "smartlab-tracker-session-2026-haianh6729"


# ------------------------------------------------------------------
# DB helpers
# ------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL;")
        g.db.execute("PRAGMA foreign_keys=ON;")
        # Tang toc doc, van an toan voi WAL: synchronous=NORMAL (khong fsync moi commit,
        # WAL checkpoint van dam bao ben vung); cache_size=-8000 = ~8MB page cache/ket noi.
        g.db.execute("PRAGMA synchronous=NORMAL;")
        g.db.execute("PRAGMA cache_size=-8000;")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()
    udb = g.pop("users_db", None)
    if udb is not None:
        udb.close()


def get_users_db():
    """Ket noi toi users.db (DB rieng cho tai khoan). Tach khoi tracker.db."""
    if "users_db" not in g:
        g.users_db = sqlite3.connect(USERS_DB_PATH)
        g.users_db.row_factory = sqlite3.Row
    return g.users_db


ADMIN_SECRET_KEY = "haianh6729"

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row  # cho phep truy cap cot bang ten trong migration
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cycle INTEGER NOT NULL,
            cycle_date TEXT,
            test_id TEXT,
            model TEXT NOT NULL,
            test_suite TEXT NOT NULL,
            test_case TEXT NOT NULL,
            state TEXT NOT NULL,
            description TEXT,
            result TEXT NOT NULL,
            created_by TEXT,
            author TEXT,
            team TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS fixes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fix_date TEXT NOT NULL,
            owner TEXT NOT NULL,
            test_suite TEXT NOT NULL,
            test_case TEXT NOT NULL,
            model_fixed TEXT NOT NULL,
            fixed_after_cycle INTEGER NOT NULL,
            note TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );

        CREATE TABLE IF NOT EXISTS test_suites (
            name TEXT PRIMARY KEY
        );

        CREATE TABLE IF NOT EXISTS models (
            name TEXT PRIMARY KEY,
            sort_order INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS owners (
            name TEXT PRIMARY KEY,
            active INTEGER DEFAULT 1,
            team TEXT
        );

        CREATE TABLE IF NOT EXISTS assignments (
            test_suite TEXT NOT NULL,
            test_case TEXT NOT NULL,
            owner TEXT,
            assigned_date TEXT,
            PRIMARY KEY (test_suite, test_case)
        );

        -- Ghi nhan script duoc VIET MOI cho tung Test Case (tab "Script viet moi").
        -- Item = test_suite chuan (suy tu tien to tc_id), tc_id = test_case -> join duoc
        -- voi bang results/fixes. tc_id DUY NHAT: moi Test Case chi co 1 ban ghi viet moi.
        CREATE TABLE IF NOT EXISTS new_scripts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item TEXT NOT NULL,
            tc_id TEXT NOT NULL UNIQUE,
            member TEXT,
            team TEXT,
            assign_week INTEGER,
            completed_date TEXT,
            status TEXT NOT NULL,          -- DONE / SKIP / ASSIGNED
            models_written TEXT,           -- CSV ten model tu bang models
            sdf_id TEXT,
            remark TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        -- Cache danh sach test case tu HE THONG CONG TY (nguon tong so TC can script,
        -- status SKIP khong tinh vao tong). source='api' (tu sync) hoac 'manual' (paste tay).
        CREATE TABLE IF NOT EXISTS company_testcases (
            tc_id TEXT PRIMARY KEY,
            item TEXT,
            status TEXT,
            raw TEXT,
            source TEXT,
            synced_at TEXT DEFAULT (datetime('now'))
        );

        -- Cache danh sach file script tren nhanh main cua repo GitHub (doi chieu 3 chieu).
        CREATE TABLE IF NOT EXISTS repo_files (
            path TEXT PRIMARY KEY,
            source TEXT,
            synced_at TEXT DEFAULT (datetime('now'))
        );

        -- Audit log: ai lam gi luc nao (mutation o admin page, danh muc, sync, backup).
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT DEFAULT (datetime('now')),
            username TEXT,
            action TEXT,
            target TEXT,
            detail TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_results_lookup ON results(test_suite, test_case, model, cycle);
        CREATE INDEX IF NOT EXISTS idx_results_cycle ON results(cycle);
        CREATE INDEX IF NOT EXISTS idx_fixes_owner ON fixes(owner);
        CREATE INDEX IF NOT EXISTS idx_assignments_owner ON assignments(owner);
        CREATE INDEX IF NOT EXISTS idx_audit_ts ON audit_log(ts);
        """
    )
    defaults = {
        "target_pass_rate": "0.88",
        "deadline_date": "",
        "project_start_date": date.today().isoformat(),
        # --- Tich hop (dien URL/token khi co tai lieu API; de trong = chua cau hinh) ---
        "farm_api_url": "",
        "farm_api_token": "",
        "import_token": "",
        "company_api_url": "",
        "company_cookie": "",
        "company_items": "",
        "github_api_base": "https://api.github.com",
        "github_repo": "",       # dang owner/repo
        "github_branch": "main",
        "github_token": "",
        # --- Tieu chi & KPI ---
        "flaky_window": "5",             # so cycle gan nhat de xet flaky
        "flaky_min_flips": "2",          # so lan doi pass<->fail toi thieu de coi la flaky
        "exit_criteria_cycles": "2",     # Done = pass N cycle lien tiep tren moi model
        "persistent_fail_cycles": "3",   # TC fail lien tiep >= N cycle gan nhat -> "dai dang"
        "exclude_new_scripts_cycles": "0",  # >0: them pass rate phu loai script moi N cycle dau
        # Whitelist loi nhieu (moi dong 1 cum, match substring khong phan biet hoa/thuong):
        # ket qua co mo ta khop -> luu voi result='Excluded' (KHONG tinh vao thong ke pass/fail).
        "error_whitelist": "",
        # --- Backup ---
        "backup_enabled": "1",
        "backup_retention": "30",
        "last_backup_date": "",
    }
    for k, v in defaults.items():
        conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", (k, v))

    # First-time seeding of models list
    n_models = conn.execute("SELECT COUNT(*) c FROM models").fetchone()[0]
    if n_models == 0:
        for i, m in enumerate(DEFAULT_MODELS):
            conn.execute("INSERT OR IGNORE INTO models (name, sort_order) VALUES (?, ?)", (m, i))

    # Backfill test_suites / models / owners from existing data (upgrade-safe for older tracker.db files)
    for row in conn.execute("SELECT DISTINCT test_suite FROM results"):
        conn.execute("INSERT OR IGNORE INTO test_suites (name) VALUES (?)", (row[0],))
    for row in conn.execute("SELECT DISTINCT model FROM results"):
        conn.execute("INSERT OR IGNORE INTO models (name) VALUES (?)", (row[0],))
    for row in conn.execute("SELECT DISTINCT owner FROM fixes"):
        conn.execute("INSERT OR IGNORE INTO owners (name) VALUES (?)", (row[0],))

    # Migration: add cac cot moi neu chua co (upgrade-safe cho tracker.db cu)
    for stmt in (
        "ALTER TABLE results ADD COLUMN created_by TEXT",
        "ALTER TABLE results ADD COLUMN author TEXT",
        "ALTER TABLE results ADD COLUMN team TEXT",
        "ALTER TABLE owners ADD COLUMN team TEXT",
        "ALTER TABLE fixes ADD COLUMN root_cause TEXT",
        "ALTER TABLE fixes ADD COLUMN root_cause_group TEXT",
        "ALTER TABLE fixes ADD COLUMN sdf_id TEXT",
        "ALTER TABLE results ADD COLUMN serial TEXT",
        "ALTER TABLE test_suites ADD COLUMN script_path TEXT DEFAULT ''",
    ):
        try:
            conn.execute(stmt)
        except Exception:
            pass  # Column already exists

    # Migration: chuan hoa ten model bi trung do hau to khac nhau (VD 'SM-X526B' vs
    # 'SM-X526' la cung 1 model) ve 1 dang chuan duy nhat qua normalize_model_name().
    # Idempotent - chi UPDATE khi gia tri khac ban chuan.
    for row in conn.execute("SELECT DISTINCT model FROM results").fetchall():
        canon = normalize_model_name(row["model"])
        if canon != row["model"]:
            conn.execute("UPDATE results SET model=? WHERE model=?", (canon, row["model"]))

    for row in conn.execute("SELECT DISTINCT model_fixed FROM fixes").fetchall():
        canon = normalize_model_name(row["model_fixed"])
        if canon != row["model_fixed"]:
            conn.execute("UPDATE fixes SET model_fixed=? WHERE model_fixed=?", (canon, row["model_fixed"]))

    for r in conn.execute(
        "SELECT id, models_written FROM new_scripts WHERE models_written IS NOT NULL AND models_written != ''"
    ).fetchall():
        parts = [p.strip() for p in r["models_written"].split(",") if p.strip()]
        canon_csv = ", ".join(normalize_model_name(p) for p in parts)
        if canon_csv != r["models_written"]:
            conn.execute("UPDATE new_scripts SET models_written=? WHERE id=?", (canon_csv, r["id"]))

    # De-duplicate bang models: gop cac ten bi coi la trung (VD 'SM-X526B' -> 'SM-X526') ve
    # 1 hang duy nhat, giu sort_order NHO NHAT (uu tien thu tu xuat hien som hon) cua nhom trung.
    dupe_groups = {}
    for row in conn.execute("SELECT name, sort_order FROM models").fetchall():
        dupe_groups.setdefault(normalize_model_name(row["name"]), []).append(row)
    for canon, members in dupe_groups.items():
        if len(members) == 1 and members[0]["name"] == canon:
            continue  # da chuan, khong co gi de gop
        min_order = min(m["sort_order"] for m in members)
        for m in members:
            conn.execute("DELETE FROM models WHERE name=?", (m["name"],))
        conn.execute("INSERT OR IGNORE INTO models (name, sort_order) VALUES (?, ?)", (canon, min_order))

    # Migration: sua du lieu cu bi loi do derive_test_suite() truoc day hardcode dau "_"
    # lam ky tu phan tach (khong nhan dien duoc tien to "sm-" dung dau "-") -> cac dong
    # SamsungMember bi luu nham thanh 'SamsungMember_ui90' (ten file .ts trong cot Test
    # Suite tho, do fallback parse khi khong khop tien to nao). Da fix root cause trong
    # derive_test_suite(); backfill nay chi sua lai du lieu cu.
    conn.execute("UPDATE results SET test_suite='SamsungMember' WHERE test_suite='SamsungMember_ui90'")

    # Ket noi du lieu co san: tu dong dien Team cho owner tu cot Team trong ket qua da nhap
    # (lay ban ghi Author=owner moi nhat co Team). Chi dien khi owner CHUA co team -> khong
    # ghi de team da chinh tay trong Cai dat. Import moi van cap nhat team truc tiep (POST /results).
    for row in conn.execute(
        """
        SELECT r.author AS author, r.team AS team
        FROM results r
        WHERE r.author IS NOT NULL AND r.author != '' AND r.team IS NOT NULL AND r.team != ''
          AND r.id = (
              SELECT MAX(r2.id) FROM results r2
              WHERE r2.author = r.author AND r2.team IS NOT NULL AND r2.team != ''
          )
        """
    ).fetchall():
        conn.execute(
            "UPDATE owners SET team=? WHERE name=? AND (team IS NULL OR team='')",
            (row["team"], row["author"]),
        )

    # Migration: backfill cycle_date tu Test ID (cho cac dong co Test ID ma hoa ngay
    # dang SDF 'YYMMDD-...'), roi danh so lai cycle theo ngay. Dong nao Test ID khong
    # ma hoa ngay (VD data demo cu) giu nguyen cycle_date da co.
    for r in conn.execute("SELECT id, test_id FROM results").fetchall():
        d = extract_date_from_test_id(r["test_id"])
        if d:
            conn.execute("UPDATE results SET cycle_date=? WHERE id=?", (d, r["id"]))
    recompute_cycles(conn)

    # Backfill assignment "dang phu trach" tu Author cua ket qua o cycle gan nhat cho
    # cac script da co du lieu tu truoc khi tinh nang nay ton tai. Dung INSERT OR IGNORE
    # de khong ghi de assignment da duoc gan thu cong tu truoc.
    conn.execute(
        """
        INSERT OR IGNORE INTO assignments (test_suite, test_case, owner, assigned_date)
        SELECT r.test_suite, r.test_case, r.author, r.cycle_date
        FROM results r
        WHERE r.author IS NOT NULL AND r.author != ''
          AND r.id = (
              SELECT r2.id FROM results r2
              WHERE r2.test_suite = r.test_suite AND r2.test_case = r.test_case
                AND r2.author IS NOT NULL AND r2.author != ''
              ORDER BY r2.cycle DESC, r2.id DESC LIMIT 1
          )
        """
    )

    # Don dep cac dong fix bi TRUNG da lo phat sinh tu truoc: cung
    # (owner, test_suite, test_case, model_fixed, fixed_after_cycle) -> chi giu dong id nho nhat.
    conn.execute(
        """
        DELETE FROM fixes WHERE id NOT IN (
            SELECT MIN(id) FROM fixes
            GROUP BY owner, test_suite, test_case, model_fixed, fixed_after_cycle
        )
        """
    )

    # Migration: backfill root_cause_group cho cac fix cu (free text) bang heuristic
    # classify_root_cause_group(). Chi chay 1 lan cho moi dong (WHERE group IS NULL),
    # KHONG sua text goc. Dong bi phan loai sai co the sua lai qua admin Fix Log.
    for r in conn.execute(
        "SELECT id, root_cause FROM fixes WHERE root_cause_group IS NULL "
        "AND root_cause IS NOT NULL AND root_cause != ''"
    ).fetchall():
        conn.execute("UPDATE fixes SET root_cause_group=? WHERE id=?",
                     (classify_root_cause_group(r["root_cause"]), r["id"]))

    # Performance indexes (idempotent). Dat o day - SAU cac ALTER TABLE ADD COLUMN o tren -
    # de index tren cot moi (VD root_cause_group) khong loi tren DB cu chua co cot.
    # KHONG doi logic truy van, chi tang toc (planner tu chon dung index).
    perf_indexes = (
        # results (bang lon nhat, tac dong cao nhat)
        "CREATE INDEX IF NOT EXISTS idx_results_result           ON results(result)",
        "CREATE INDEX IF NOT EXISTS idx_results_suite_case        ON results(test_suite, test_case)",
        "CREATE INDEX IF NOT EXISTS idx_results_suite_case_cycle  ON results(test_suite, test_case, cycle)",
        "CREATE INDEX IF NOT EXISTS idx_results_model_cycle       ON results(model, cycle)",
        "CREATE INDEX IF NOT EXISTS idx_results_suite_model_cycle ON results(test_suite, model, cycle)",
        # fixes
        "CREATE INDEX IF NOT EXISTS idx_fixes_fix_date            ON fixes(fix_date)",
        "CREATE INDEX IF NOT EXISTS idx_fixes_root_cause_group    ON fixes(root_cause_group)",
        # new_scripts
        "CREATE INDEX IF NOT EXISTS idx_new_scripts_status        ON new_scripts(status)",
        "CREATE INDEX IF NOT EXISTS idx_new_scripts_status_date   ON new_scripts(status, completed_date)",
        "CREATE INDEX IF NOT EXISTS idx_new_scripts_member_status ON new_scripts(member, status)",
        # owners
        "CREATE INDEX IF NOT EXISTS idx_owners_active             ON owners(active)",
    )
    for stmt in perf_indexes:
        conn.execute(stmt)
    conn.execute("ANALYZE")  # cap nhat thong ke de planner dung index moi hieu qua

    conn.commit()
    conn.close()


def ensure_account_for_owner(udb, name):
    """Neu owner chua co tai khoan trong users.db, tao sap dat: role=user,
    quyen = mac dinh cua user, mat khau = DEFAULT_RESET_PASSWORD. Neu co roi -> bo qua
    (khong ghi de tai khoan da ton tai)."""
    if udb.execute("SELECT 1 FROM users WHERE username=?", (name,)).fetchone():
        return  # Co roi
    udb.execute(
        "INSERT INTO users (username, password_hash, role, permissions, active) VALUES (?, ?, ?, ?, 1)",
        (name, generate_password_hash(DEFAULT_RESET_PASSWORD), "user",
         ",".join(ROLE_DEFAULT_PERMS["user"]))
    )


def init_users_db():
    """Tao users.db (DB rieng) + seed tai khoan admin khoi tao (anh.hh). Idempotent."""
    conn = sqlite3.connect(USERS_DB_PATH)
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            permissions TEXT,               -- CSV cac tab duoc phep
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );
        """
    )
    # Migration idempotent (upgrade-safe cho users.db cu neu them cot ve sau)
    for stmt in (
        "ALTER TABLE users ADD COLUMN permissions TEXT",
        "ALTER TABLE users ADD COLUMN active INTEGER DEFAULT 1",
    ):
        try:
            conn.execute(stmt)
        except Exception:
            pass
    # Seed/ep anh.hh la admin (mat khau mac dinh neu chua ton tai).
    row = conn.execute("SELECT username FROM users WHERE username=?", (BOOTSTRAP_ADMIN,)).fetchone()
    if not row:
        conn.execute(
            "INSERT INTO users (username, password_hash, role, permissions, active) VALUES (?,?,?,?,1)",
            (BOOTSTRAP_ADMIN, generate_password_hash(DEFAULT_RESET_PASSWORD),
             "admin", ",".join(ROLE_DEFAULT_PERMS["admin"])),
        )
    else:
        # Luon dam bao anh.hh giu role admin + du quyen.
        conn.execute(
            "UPDATE users SET role='admin', permissions=?, active=1 WHERE username=?",
            (",".join(ROLE_DEFAULT_PERMS["admin"]), BOOTSTRAP_ADMIN),
        )
    conn.commit()
    # Backfill: dam bao tai khoan cu co du cac quyen MOI theo role (permissions luu CSV
    # snapshot tai thoi diem tao nen quyen them ve sau khong tu co). CHI APPEND quyen
    # thieu, khong ghi de/xoa permission tuy chinh ho dang co.
    #   - admin/moderator: ns-assign/ns-edit + tab moi reports/integrations
    #   - user: tab moi reports
    _role_backfill = {
        "admin": set(NS_EXTRA_PERMS) | {"reports", "integrations"},
        "moderator": set(NS_EXTRA_PERMS) | {"reports", "integrations"},
        "user": {"reports"},
    }
    for username, role, permissions in conn.execute(
        "SELECT username, role, permissions FROM users"
    ).fetchall():
        ensure = _role_backfill.get(role)
        if not ensure:
            continue
        current = set(perms_to_list(permissions))
        missing = ensure - current
        if missing:
            conn.execute("UPDATE users SET permissions=? WHERE username=?",
                         (",".join(sorted(current | missing)), username))
    conn.commit()
    # Backfill: tao tai khoan cho tat ca owner active chua co
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    for owner in db.execute("SELECT name FROM owners WHERE active=1").fetchall():
        ensure_account_for_owner(conn, owner["name"])
    db.close()
    conn.commit()
    conn.close()


# ------------------------------------------------------------------
# Auth helpers + decorators
# ------------------------------------------------------------------
def perms_to_list(csv_str):
    return [p for p in (csv_str or "").split(",") if p]


def owner_is_active(name):
    """True neu owner ton tai va con active trong tracker.db (owner nghi -> khoa dang nhap)."""
    db = get_db()
    row = db.execute("SELECT active FROM owners WHERE name=?", (name,)).fetchone()
    return bool(row) and (row["active"] == 1)


def current_user():
    """Tra ve dict tai khoan dang dang nhap (con hop le) hoac None."""
    name = session.get("user")
    if not name:
        return None
    udb = get_users_db()
    row = udb.execute(
        "SELECT username, role, permissions, active FROM users WHERE username=?", (name,)
    ).fetchone()
    if not row or row["active"] != 1:
        return None
    if not owner_is_active(name):
        return None
    return {
        "username": row["username"],
        "role": row["role"],
        "permissions": perms_to_list(row["permissions"]),
    }


def require_login(fn):
    @functools.wraps(fn)
    def wrapper(*args, **kwargs):
        if current_user() is None:
            return jsonify({"error": "Chưa đăng nhập"}), 401
        return fn(*args, **kwargs)
    return wrapper


def require_perm(tab):
    """Chan API ghi: phai dang nhap VA co quyen dung tab tuong ung."""
    def deco(fn):
        @functools.wraps(fn)
        def wrapper(*args, **kwargs):
            err = perm_error(tab)
            if err:
                return err
            return fn(*args, **kwargs)
        return wrapper
    return deco


def perm_error(tab):
    """Dung cho handler GET+POST chung: chi chan nhanh ghi. Tra ve response loi hoac None."""
    u = current_user()
    if u is None:
        return jsonify({"error": "Chưa đăng nhập"}), 401
    if tab not in u["permissions"]:
        return jsonify({"error": f"Bạn không có quyền dùng chức năng này ({tab})."}), 403
    return None


def log_audit(db, action, target="", detail=""):
    """Ghi 1 dong audit_log (ai / lam gi / len cai gi / chi tiet). KHONG commit -
    di cung transaction cua mutation goi no (caller commit). An toan goi trong
    request context; ngoai request (backup thread) truyen username qua detail."""
    username = "(anonymous)"
    try:
        u = current_user()
        if u:
            username = u["username"]
        elif request.headers.get("X-Admin-Key"):
            username = "admin-key"
        elif request.headers.get("X-Import-Token"):
            username = "import-token"
    except RuntimeError:
        username = "(system)"  # ngoai request context (backup daemon...)
    db.execute(
        "INSERT INTO audit_log (username, action, target, detail) VALUES (?,?,?,?)",
        (username, action, str(target or ""), str(detail or "")),
    )


def _http_json(url, token="", method="GET", payload=None, timeout=30, cookie=""):
    """Goi HTTP JSON ra ngoai (farm API / TC Hub / GitHub) bang urllib stdlib.
    Tra ve object da parse JSON; nem RuntimeError voi message tieng Viet de hien thang
    len UI khi loi (khong ket noi duoc / HTTP status loi / body khong phai JSON).
    token -> header Authorization: Bearer (farm dung JWT). cookie -> header Cookie
    (TC Hub dung cookie phien dang nhap san, khong dung Bearer)."""
    headers = {"User-Agent": "test-stabilization-tracker", "Accept": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if cookie:
        headers["Cookie"] = cookie
    body = None
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"API trả về HTTP {e.code} ({e.reason}) — kiểm tra URL/token trong Cài đặt.")
    except urllib.error.URLError as e:
        raise RuntimeError(f"Không kết nối được tới API ({e.reason}) — kiểm tra URL/mạng.")
    except TimeoutError:
        raise RuntimeError(f"API không phản hồi sau {timeout}s (timeout).")
    try:
        return json.loads(raw)
    except ValueError:
        raise RuntimeError("API trả về dữ liệu không phải JSON — kiểm tra lại URL endpoint.")


def _settings_cache(db):
    """Cache toan bo bang settings (nho, ~30 dong) vao request context (g) o lan doc dau,
    tranh round-trip SQLite lap lai nhieu lan/1 request. Tu dong huy khi request ket thuc.
    Phai goi _invalidate_settings_cache() sau moi lan GHI settings de tranh doc gia tri cu."""
    cache = g.get("_settings_cache")
    if cache is None:
        cache = {r["key"]: r["value"]
                 for r in db.execute("SELECT key, value FROM settings").fetchall()}
        g._settings_cache = cache
    return cache


def _invalidate_settings_cache():
    """Xoa cache settings (goi sau khi INSERT/UPDATE settings trong cung request)."""
    if g.get("_settings_cache") is not None:
        g._settings_cache = None


def get_setting(db, key, default=""):
    val = _settings_cache(db).get(key)
    return val if val is not None else default


def get_setting_int(db, key, default):
    try:
        return int(float(get_setting(db, key, str(default)) or default))
    except (ValueError, TypeError):
        return default


def classify_result(state):
    return "Pass" if state.strip().lower() in PASS_STATES else "Fail"


def is_skip_state(state):
    return state.strip().lower() in SKIP_STATES


def get_error_whitelist(db):
    """List cac cum tu whitelist loi nhieu (setting 'error_whitelist', moi dong 1 cum).
    Ket qua co Description khop 1 trong cac cum nay se duoc luu voi result='Excluded'
    (khong tinh vao thong ke pass/fail) de khong gay nhieu viec fix."""
    raw = get_setting(db, "error_whitelist", "") or ""
    return [line.strip() for line in raw.splitlines() if line.strip()]


def matches_whitelist(description, patterns):
    """True neu description chua bat ky cum whitelist nao (substring, khong phan biet hoa/thuong)."""
    if not patterns:
        return False
    low = str(description or "").lower()
    return any(p.lower() in low for p in patterns)


# ------------------------------------------------------------------
# Root-cause classifier: gom mo ta loi (Description) - dac biet cac state
# Exception/Error/Traceback - ve 1 so NHOM NGUYEN NHAN co nghia, thay vi de nguyen
# text tho (moi dong khac nhau vi chua device SN / test id / duong dan).
# Moi rule: (regex quet toan bo description, khong phan biet hoa thuong) -> nhan nhom.
# Xep tu CU THE nhat -> chung nhat; rule dau tien khop se duoc chon.
# ------------------------------------------------------------------
ROOT_CAUSE_RULES = [
    (r"some device is not Free", "Hạ tầng: Device đang bận (không xin được máy để chạy)"),
    (r"Device number\s*:\s*\d+\s*/\s*\d+|DeviceNotFoundException|device.*not.*found",
     "Hạ tầng: Mất kết nối / không đủ thiết bị (PC ↔ điện thoại)"),
    (r"FAIL TO CONNECT STP|connect\w*\s+stp|stp\s+server", "Hạ tầng: Lỗi kết nối tới STP server"),
    (r"WIFI_FAIL|WIFI-\d+|not found AP|WiFi connection failed|k\w* AP\b",
     "Hạ tầng: Lỗi kết nối WiFi (không tìm thấy/không kết nối được AP)"),
    (r"ModuleNotFoundError|No module named", "Script: Thiếu module / lỗi import"),
    (r"scriptLog is None|MTCN_Lib\s*\|\s*ERROR", "Hạ tầng: Lỗi script log (MTCN)"),
    (r"NoneType'?\s+object\s+has\s+no\s+attribute|NoneType'?\s+object\s+is\s+not\s+(subscriptable|iterable)",
     "Script: Không tìm thấy phần tử UI (đối tượng None)"),
    (r"TimeoutExpired|\bTIMEOUT\b|timed?\s*out", "Timeout: Quá thời gian chờ"),
    (r"ParseError|not well-formed|invalid token", "Script: Lỗi phân tích XML"),
    (r"pre[\s\-]?condition", "Tiền đề: Lỗi thiết lập điều kiện tiền đề (pre-condition)"),
    (r"Cannot run Samsung Browser", "Ứng dụng: Không mở được Samsung Browser"),
    (r"\b(IndexError|KeyError|ValueError|TypeError|AttributeError|UnboundLocalError|RuntimeError|ZeroDivisionError|StopIteration|AssertionError)\b",
     "Script: Lỗi logic (Python exception)"),
]

# Prefix marker o dau dong mo ta functional-fail (VD "# 8.1. FAIL, ...", "# Step 2. FAILED, ...")
_FAIL_PREFIX_RE = re.compile(r"^\s*#?\s*(?:step\s*)?[\d.\s]*\.?\s*(?:fail(?:ed)?|warning|error)\s*[:,\-]?\s*", re.I)


def summarize_root_cause(description, state=""):
    """Tra ve 1 nhan NHOM NGUYEN NHAN goi cho description. Voi Exception/Traceback ->
    quy ve nhom ha tang/script tuong ung. Voi functional-fail thuong -> lam sach dong
    dau (bo tien to '# N. FAIL,', bo device SN trong [...], gop khoang trang) de cac loi
    giong nhau gom chung 1 nhom."""
    text = (description or "").strip()
    if not text or text.lower() == "none":
        st = (state or "").strip().upper()
        return f"Không có mô tả ({st})" if st else "Không có mô tả lỗi"
    for pat, label in ROOT_CAUSE_RULES:
        if re.search(pat, text, re.I):
            return label
    # Fallback: lay dong dau co nghia, lam sach de gom nhom
    first = next((l.strip() for l in text.splitlines() if l.strip()), text)
    first = _FAIL_PREFIX_RE.sub("", first)
    first = re.sub(r"\[[^\]]*\]", "[...]", first)      # bo noi dung trong ngoac vuong (device SN...)
    first = re.sub(r"['\"].*?['\"]", "'...'", first)   # bo chuoi trong nhay
    first = re.sub(r"\d+", "N", first)                 # bo so cu the
    first = re.sub(r"\s+", " ", first).strip()
    return ("Khác: " + first[:70]) if first else "Khác: (không rõ)"


# Ban dich sang tieng Anh cho cac nhan CO DINH cua ROOT_CAUSE_RULES - chi dung cho
# Dashboard + bao cao sinh ra (2 noi bat buoc tieng Anh). Export Excel toan bo du lieu
# (tab Uu tien) va cac consumer khac VAN dung nhan tieng Viet goc (summarize_root_cause).
_ROOT_CAUSE_LABEL_EN = {label: label for _, label in ROOT_CAUSE_RULES}
_ROOT_CAUSE_LABEL_EN.update({
    "Hạ tầng: Device đang bận (không xin được máy để chạy)": "Infra: Device busy (could not acquire a device to run)",
    "Hạ tầng: Mất kết nối / không đủ thiết bị (PC ↔ điện thoại)": "Infra: Connection lost / insufficient devices (PC ↔ phone)",
    "Hạ tầng: Lỗi kết nối tới STP server": "Infra: STP server connection error",
    "Hạ tầng: Lỗi kết nối WiFi (không tìm thấy/không kết nối được AP)": "Infra: WiFi connection error (AP not found / could not connect)",
    "Script: Thiếu module / lỗi import": "Script: Missing module / import error",
    "Hạ tầng: Lỗi script log (MTCN)": "Infra: Script log error (MTCN)",
    "Script: Không tìm thấy phần tử UI (đối tượng None)": "Script: UI element not found (None object)",
    "Timeout: Quá thời gian chờ": "Timeout: Exceeded wait time",
    "Script: Lỗi phân tích XML": "Script: XML parse error",
    "Tiền đề: Lỗi thiết lập điều kiện tiền đề (pre-condition)": "Precondition: Failed to set up precondition",
    "Ứng dụng: Không mở được Samsung Browser": "App: Could not open Samsung Browser",
    "Script: Lỗi logic (Python exception)": "Script: Logic error (Python exception)",
})


def _translate_root_cause_label(label):
    """Dich 1 nhan root-cause sang tieng Anh cho Dashboard/bao cao. Nhan dong
    ('Khác: ...', 'Không có mô tả...') giu nguyen tien to dich, phan con lai
    (trich tu description raw) la DU LIEU nguoi dung nhap - khong dich."""
    if label in _ROOT_CAUSE_LABEL_EN:
        return _ROOT_CAUSE_LABEL_EN[label]
    if label.startswith("Khác: "):
        return "Other: " + label[len("Khác: "):]
    if label == "Không có mô tả lỗi":
        return "No failure description"
    if label.startswith("Không có mô tả ("):
        return "No description (" + label[len("Không có mô tả ("):]
    return label


def compute_root_cause_pareto(db, limit=15, english=False):
    """Pareto nguyen nhan loi da GOM NHOM (thay vi group text tho). Tra ve list dict
    {description(=ten nhom), count, pct, cum_pct} da sap giam dan + tinh % cong don.
    english=True: dich nhan sang tieng Anh (chi dung cho Dashboard/bao cao)."""
    groups = {}
    for r in db.execute("SELECT description, state FROM results WHERE result='Fail'"):
        label = summarize_root_cause(r["description"], r["state"])
        groups[label] = groups.get(label, 0) + 1
    total_fail = sum(groups.values())
    ordered = sorted(groups.items(), key=lambda kv: (-kv[1], kv[0]))[:limit]
    out, cum = [], 0
    for label, cnt in ordered:
        pct = (cnt / total_fail) if total_fail else 0
        cum += pct
        out.append({"key": label,
                    "description": _translate_root_cause_label(label) if english else label,
                    "count": cnt, "pct": pct, "cum_pct": cum})
    return out


def compute_cycle_trend(db):
    """Pass Rate theo cycle, theo cong thuc rieng (khac voi cot 'result' dung cho fail_count/
    priority engine): NA duoc coi la trung lap - KHONG tinh vao ca tu so lan mau so.

    Pass rate = (so PASS + so CHECK + so MANUAL CHECK) / (Tong so - Skip - NA)
    Vi RUNNING (Skip) da bi loai bo tu luc import (khong bao gio duoc luu vao DB), nen
    o day chi can tru NA khoi mau so: denom = total - na_count.
    """
    rows = db.execute(
        """
        SELECT cycle, MIN(cycle_date) as cycle_date, COUNT(*) as total,
               SUM(CASE WHEN LOWER(TRIM(state)) IN ('pass','check','manual check') THEN 1 ELSE 0 END) as pass_like_count,
               SUM(CASE WHEN LOWER(TRIM(state))='na' THEN 1 ELSE 0 END) as na_count
        FROM results WHERE result <> 'Excluded' GROUP BY cycle ORDER BY cycle
        """
    ).fetchall()
    trend = []
    prev_fail, prev_rate = None, None
    for c in rows:
        denom = c["total"] - c["na_count"]
        fail_count = denom - c["pass_like_count"]  # tat ca state khong phai Pass/Check/ManualCheck/NA
        rate = (c["pass_like_count"] / denom) if denom > 0 else None
        delta_fail = (fail_count - prev_fail) if prev_fail is not None else None
        delta_rate = (rate - prev_rate) if (rate is not None and prev_rate is not None) else None
        trend.append({
            "cycle": c["cycle"], "cycle_date": c["cycle_date"], "total": c["total"],
            "pass_count": c["pass_like_count"], "na_count": c["na_count"], "fail_count": fail_count,
            "pass_rate": rate, "delta_fail": delta_fail, "delta_rate": delta_rate,
        })
        prev_fail, prev_rate = fail_count, rate
    return trend


def compute_fix_root_cause_pareto(db, date_from=None, date_to=None):
    """Pareto theo NHOM nguyen nhan DA XAC NHAN khi fix (fixes.root_cause_group) - khac
    voi compute_root_cause_pareto (bucket description cua ket qua Fail = 'dang fail vi gi').
    Loc theo fix_date [date_from, date_to] neu truyen. Tra ve list nhu pareto kia."""
    q = "SELECT root_cause_group, COUNT(*) c FROM fixes WHERE root_cause_group IS NOT NULL AND root_cause_group != ''"
    args = []
    if date_from:
        q += " AND fix_date >= ?"
        args.append(date_from)
    if date_to:
        q += " AND fix_date <= ?"
        args.append(date_to)
    q += " GROUP BY root_cause_group ORDER BY c DESC"
    rows = db.execute(q, args).fetchall()
    total = sum(r["c"] for r in rows)
    out, cum = [], 0
    for r in rows:
        pct = (r["c"] / total) if total else 0
        cum += pct
        out.append({"group": r["root_cause_group"], "count": r["c"], "pct": pct, "cum_pct": cum})
    return out


def compute_adjusted_trend(db, exclude_n):
    """Pass rate theo cycle nhung LOAI cac dong cua script dang trong exclude_n cycle DAU
    TIEN cua chinh no (script moi chua on dinh khong keo tut pass rate chung - B2).
    Tra ve {cycle: pass_rate_or_None}."""
    rows = db.execute(
        """
        SELECT test_suite, test_case, cycle, COUNT(*) as total,
               SUM(CASE WHEN LOWER(TRIM(state)) IN ('pass','check','manual check') THEN 1 ELSE 0 END) as pass_like,
               SUM(CASE WHEN LOWER(TRIM(state))='na' THEN 1 ELSE 0 END) as na
        FROM results WHERE result <> 'Excluded' GROUP BY test_suite, test_case, cycle
        """
    ).fetchall()
    first_cycle = {}
    for r in rows:
        k = (r["test_suite"], r["test_case"])
        first_cycle[k] = min(first_cycle.get(k, r["cycle"]), r["cycle"])
    agg = {}
    for r in rows:
        k = (r["test_suite"], r["test_case"])
        if r["cycle"] < first_cycle[k] + exclude_n:
            continue  # script con "moi" tai cycle nay -> loai khoi pass rate dieu chinh
        a = agg.setdefault(r["cycle"], {"total": 0, "pass": 0, "na": 0})
        a["total"] += r["total"]
        a["pass"] += r["pass_like"]
        a["na"] += r["na"]
    out = {}
    for cyc, a in agg.items():
        denom = a["total"] - a["na"]
        out[cyc] = (a["pass"] / denom) if denom > 0 else None
    return out


def compute_coverage(db):
    """Tien do phu script so voi TONG DONG tu he thong cong ty (cache company_testcases).
    total_needed = so TC status != SKIP. done = so TC DONE trong new_scripts VA nam trong
    danh sach can lam (khong SKIP ben cong ty). DONE ngoai danh sach -> out_of_plan.
    Cache rong -> {"configured": False} (KHONG fallback ve con so cung nao)."""
    comp_rows = db.execute("SELECT tc_id, item, status, synced_at FROM company_testcases").fetchall()
    done_rows = db.execute("SELECT tc_id, item FROM new_scripts WHERE status='DONE'").fetchall()
    if not comp_rows:
        return {"configured": False, "done_recorded": len(done_rows)}

    def norm(tc):
        return str(tc or "").strip().lower()

    needed = {}   # tc_id_norm -> item
    skip_ids = set()
    target_ids = set()   # TC status 'target' = can hoan thanh nhung chua xong
    for r in comp_rows:
        item = (r["item"] or "").strip() or item_from_tc_id(r["tc_id"]) or "Unknown"
        status_l = (r["status"] or "").strip().lower()
        if status_l in COMPANY_SKIP_STATES:
            skip_ids.add(norm(r["tc_id"]))
        else:
            needed[norm(r["tc_id"])] = item
            if status_l in COMPANY_TARGET_STATES:
                target_ids.add(norm(r["tc_id"]))

    done_in_plan = set()
    out_of_plan = []
    done_by_item = {}
    for r in done_rows:
        key = norm(r["tc_id"])
        if key in needed:
            done_in_plan.add(key)
            it = needed[key]
            done_by_item[it] = done_by_item.get(it, 0) + 1
        else:
            out_of_plan.append(r["tc_id"])  # bi SKIP ben cong ty hoac khong co trong danh sach

    needed_by_item = {}
    for it in needed.values():
        needed_by_item[it] = needed_by_item.get(it, 0) + 1

    by_item = [
        {
            "item": it, "needed": n, "done": done_by_item.get(it, 0),
            "pct": (done_by_item.get(it, 0) / n) if n else 0,
        }
        for it, n in sorted(needed_by_item.items())
    ]
    total_needed = len(needed)
    done_n = len(done_in_plan)
    # TC status 'target' con lai (chua ghi nhan DONE ben new_scripts) = viec con phai lam.
    target_pending = len(target_ids - done_in_plan)
    synced_at = max((r["synced_at"] or "" for r in comp_rows), default="")
    return {
        "configured": True,
        "total_needed": total_needed,
        "skip": len(skip_ids),
        "done": done_n,
        "target": len(target_ids),
        "target_pending": target_pending,
        "pct": (done_n / total_needed) if total_needed else 0,
        "out_of_plan": len(out_of_plan),
        "by_item": by_item,
        "synced_at": synced_at,
    }


def compute_script_cycle_matrix(db, group_by_model=False):
    """Pass rate + fail count của TỪNG SCRIPT (test_suite/test_case), tách riêng theo TỪNG
    CYCLE (dùng đúng công thức Pass Rate như compute_cycle_trend), để so sánh 1 script
    qua các lần chạy: cycle này so cycle trước tốt lên/xấu đi/không đổi ra sao.

    Nếu group_by_model=True, mỗi dòng tách riêng theo TỪNG MODEL (test_suite/test_case/model)
    thay vì gộp chung mọi model — dùng khi cần soi pass rate của 1 item trên 1 model cụ thể.

    Trả về {"cycles": [{cycle, cycle_date}, ...], "scripts": [{
        test_suite, test_case, model (chỉ có khi group_by_model=True),
        current_owner, team, priority_tier,
        by_cycle: {cycle_num: {total, pass_count, na_count, fail_count, pass_rate, verdict}},
        overall_trend: "improved"|"regressed"|"unchanged"|"insufficient_data",
        first_cycle_with_data, last_cycle_with_data,
    }, ...]}
    verdict của 1 cell = so với cycle GẦN NHẤT TRƯỚC ĐÓ mà script này có chạy (không nhất
    thiết là cycle liền kề, vì có thể script không chạy ở 1 vài cycle).
    """
    cycles = db.execute("SELECT DISTINCT cycle, cycle_date FROM results ORDER BY cycle").fetchall()
    cycle_list = [{"cycle": r["cycle"], "cycle_date": r["cycle_date"]} for r in cycles]

    group_cols = "test_suite, test_case, model, cycle" if group_by_model else "test_suite, test_case, cycle"
    rows = db.execute(
        f"""
        SELECT {group_cols}, COUNT(*) as total,
               SUM(CASE WHEN LOWER(TRIM(state)) IN ('pass','check','manual check') THEN 1 ELSE 0 END) as pass_like_count,
               SUM(CASE WHEN LOWER(TRIM(state))='na' THEN 1 ELSE 0 END) as na_count
        FROM results WHERE result <> 'Excluded' GROUP BY {group_cols}
        """
    ).fetchall()

    by_script = {}
    for r in rows:
        key = (r["test_suite"], r["test_case"], r["model"]) if group_by_model else (r["test_suite"], r["test_case"])
        by_script.setdefault(key, {})[r["cycle"]] = {
            "total": r["total"], "pass_like_count": r["pass_like_count"], "na_count": r["na_count"],
        }

    priority_map = {(p["test_suite"], p["test_case"]): p for p in get_script_priority(db)}

    scripts_out = []
    for key, cell_by_cycle in by_script.items():
        ordered_cycles = sorted(cell_by_cycle.keys())
        by_cycle = {}
        prev_rate = None
        rates_with_data = []
        for cn in ordered_cycles:
            c = cell_by_cycle[cn]
            denom = c["total"] - c["na_count"]
            rate = (c["pass_like_count"] / denom) if denom > 0 else None
            fail_count = denom - c["pass_like_count"] if denom > 0 else 0
            verdict = None
            if prev_rate is not None and rate is not None:
                if rate > prev_rate:
                    verdict = "improved"
                elif rate < prev_rate:
                    verdict = "regressed"
                else:
                    verdict = "unchanged"
            by_cycle[cn] = {
                "total": c["total"], "pass_count": c["pass_like_count"], "na_count": c["na_count"],
                "fail_count": fail_count, "pass_rate": rate, "verdict": verdict,
            }
            if rate is not None:
                prev_rate = rate
                rates_with_data.append(rate)

        if len(rates_with_data) < 2:
            overall = "insufficient_data"
        elif rates_with_data[-1] > rates_with_data[-2]:
            overall = "improved"
        elif rates_with_data[-1] < rates_with_data[-2]:
            overall = "regressed"
        else:
            overall = "unchanged"

        if group_by_model:
            suite, case, model = key
            p = priority_map.get((suite, case), {})
        else:
            suite, case = key
            model = None
            p = priority_map.get(key, {})

        entry = {
            "test_suite": suite, "test_case": case,
            "current_owner": p.get("current_owner", ""),
            "team": p.get("team", ""),
            "priority_tier": p.get("priority_tier", ""),
            "by_cycle": by_cycle,
            "overall_trend": overall,
            "first_cycle_with_data": ordered_cycles[0] if ordered_cycles else None,
            "last_cycle_with_data": ordered_cycles[-1] if ordered_cycles else None,
        }
        if group_by_model:
            entry["model"] = model
        scripts_out.append(entry)

    return {"cycles": cycle_list, "scripts": scripts_out}


def compute_suite_model_matrix(db):
    """Pass rate cua TUNG ITEM (test_suite: Reminder, Wallpaper, Weather...) tren TUNG MODEL,
    tach rieng theo TUNG CYCLE. Dung cho Dashboard de xem 1 item chay tren 1 model qua cac cycle
    tot len/xau di ra sao, va tinh overall pass rate cho cac cycle nguoi dung chon.

    Tra ve {
      "cycles": [{cycle, cycle_date}, ...],
      "rows": [{test_suite, model, by_cycle: {cycle: {total, pass_count, na_count, fail_count, pass_rate}}}],
      "overall_by_cycle": {cycle: {total, pass_count, na_count, fail_count, pass_rate}},  # tat ca script
    }
    Frontend co the cong don cac cycle duoc chon: overall = sum(pass_count) / (sum(total) - sum(na)).
    """
    cycles = db.execute("SELECT DISTINCT cycle, cycle_date FROM results ORDER BY cycle").fetchall()
    cycle_list = [{"cycle": r["cycle"], "cycle_date": r["cycle_date"]} for r in cycles]

    def cell(total, pass_like, na):
        denom = total - na
        return {
            "total": total, "pass_count": pass_like, "na_count": na,
            "fail_count": (denom - pass_like) if denom > 0 else 0,
            "pass_rate": (pass_like / denom) if denom > 0 else None,
        }

    rows_raw = db.execute(
        """
        SELECT test_suite, model, cycle, COUNT(*) as total,
               SUM(CASE WHEN LOWER(TRIM(state)) IN ('pass','check','manual check') THEN 1 ELSE 0 END) as pass_like,
               SUM(CASE WHEN LOWER(TRIM(state))='na' THEN 1 ELSE 0 END) as na
        FROM results WHERE result <> 'Excluded' GROUP BY test_suite, model, cycle
        """
    ).fetchall()
    by_key = {}
    for r in rows_raw:
        by_key.setdefault((r["test_suite"], r["model"]), {})[r["cycle"]] = cell(r["total"], r["pass_like"], r["na"])

    rows = []
    for (suite, model), by_cycle in sorted(by_key.items()):
        rows.append({"test_suite": suite, "model": model, "by_cycle": by_cycle})

    overall_raw = db.execute(
        """
        SELECT cycle, COUNT(*) as total,
               SUM(CASE WHEN LOWER(TRIM(state)) IN ('pass','check','manual check') THEN 1 ELSE 0 END) as pass_like,
               SUM(CASE WHEN LOWER(TRIM(state))='na' THEN 1 ELSE 0 END) as na
        FROM results WHERE result <> 'Excluded' GROUP BY cycle
        """
    ).fetchall()
    overall_by_cycle = {r["cycle"]: cell(r["total"], r["pass_like"], r["na"]) for r in overall_raw}

    return {"cycles": cycle_list, "rows": rows, "overall_by_cycle": overall_by_cycle}


def extract_date_from_test_id(test_id):
    """Test ID/Request ID dang SDF ma hoa ngay tao o 6 chu so dau: 'YYMMDD-NNNNN'.
    VD '260706-01949' -> '2026-07-06'. Tra ve ISO date string, hoac None neu khong hop le.
    """
    s = str(test_id or "").strip()
    m = re.match(r"^(\d{2})(\d{2})(\d{2})\b", s)
    if not m:
        return None
    yy, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(2000 + yy, mm, dd).isoformat()
    except ValueError:
        return None  # thang/ngay khong hop le


def recompute_cycles(db):
    """Cycle duoc suy hoan toan tu ngay chay (cycle_date): moi ngay khac nhau = 1 cycle,
    danh so tang dan theo thu tu thoi gian (ngay som nhat = cycle 1). Goi sau moi lan
    import hoac migration. Dam bao cycle luon nhat quan & khoa hoc theo ngay tao Test ID.
    """
    dates = [
        r["cycle_date"]
        for r in db.execute(
            "SELECT DISTINCT cycle_date FROM results WHERE cycle_date IS NOT NULL AND cycle_date != '' ORDER BY cycle_date"
        ).fetchall()
    ]
    for idx, d in enumerate(dates, start=1):
        db.execute("UPDATE results SET cycle=? WHERE cycle_date=?", (idx, d))


# TC name prefix -> ten Test Suite/chuc nang chuan hoa (khong phan biet hoa/thuong,
# nhung phai la tien to dung nghia, VD "Weather_" chu khong khop "Weathermania_")
TC_SUITE_RULES = [
    (("browser", "internet"), "Internet"),
    (("keyboard", "skbd"), "Keyboard"),
    (("weather",), "Weather"),
    (("wallpaper",), "Wallpaper"),
    (("sm",), "SamsungMember"),
    (("reminder",), "Reminder"),
    (("nowbrief",), "Now-brief"),
]

# TC ID cua SamsungMember dung dau gach ngang (VD "SM-00-002") thay vi gach duoi nhu
# cac Item khac -> rieng tien to "sm" CHI chap nhan dau "-". Dung chung boi ca
# derive_test_suite() va item_from_tc_id() (mot nguon su that, tranh lech logic).
_TC_ID_PREFIX_SEPARATORS = {"sm": ("-",)}


def extract_test_case_name(raw):
    """'Internet/.../Browser_000118.py' hoac 'Browser_000118.py' -> 'Browser_000118'"""
    name = str(raw or "").strip().replace("\\", "/")
    name = name.split("/")[-1]
    if name.lower().endswith(".py"):
        name = name[:-3]
    return name


def derive_test_suite(test_case_name, raw_test_suite):
    """Uu tien suy ra ten chuc nang chuan tu tien to cua Test Case (dang tin cay hon
    vi cot Test Suite tho co the la duong dan day du, ten file .ts, hoac bi dien nham
    bang chinh ten Test Case). Neu khong khop tien to nao, fallback ve parse cot tho."""
    lower = test_case_name.lower()
    for prefixes, suite in TC_SUITE_RULES:
        for p in prefixes:
            seps = _TC_ID_PREFIX_SEPARATORS.get(p, ("_",))
            if any(lower.startswith(p + sep) for sep in seps):
                return suite

    raw = str(raw_test_suite or "").strip().replace("\\", "/")
    if not raw:
        return "Unknown"
    base = raw.split("/")[-1]
    if base.lower().endswith(".ts"):
        base = base[:-3]
    # Neu cot Test Suite bi dien nham bang chinh ten Test Case (.py) thi khong dung duoc
    if base.lower() == test_case_name.lower() or base.lower().endswith(".py"):
        return "Unknown"
    return base


# Item mau + vi du dinh dang TC ID (chi de HIEN THI goi y; validate thuc te dua vao tien to).
ITEM_TC_EXAMPLES = {
    "Internet": "Browser_000001",
    "Keyboard": "SKBD_000001",
    "Now-brief": "NowBrief_000001",
    "Reminder": "Reminder_000001",
    "SamsungMember": "SM-00-002",
    "Wallpaper": "Wallpaper_000001",
    "Weather": "Weather_000001",
}


def item_from_tc_id(tc_id):
    """Suy Item (ten test_suite chuan) tu TC ID theo dung tien to trong TC_SUITE_RULES.
    Tra ve None neu khong khop tien to nao (dung de validate 'sai dinh dang')."""
    lower = str(tc_id or "").strip().lower()
    for prefixes, suite in TC_SUITE_RULES:
        for p in prefixes:
            seps = _TC_ID_PREFIX_SEPARATORS.get(p, ("_",))
            if any(lower.startswith(p + sep) for sep in seps):
                return suite
    return None


def get_models_list(db):
    rows = db.execute("SELECT name FROM models ORDER BY sort_order, name").fetchall()
    return [r["name"] for r in rows]


_MODEL_NAME_RE = re.compile(r"^(SM-[A-Za-z]\d+)")


def normalize_model_name(raw):
    """Chuan hoa ten model ve dang goc 'SM-' + 1 chu cai + chuoi so lien tiep, bo moi hau to
    theo sau (VD 'SM-X526B' -> 'SM-X526', 'SM-A175F' -> 'SM-A175'). No-op (giu nguyen) neu
    khong khop dinh dang SM-<letter><digits> (VD 'All Models', chuoi rong, ten khac he SM-)."""
    s = str(raw or "").strip()
    if not s:
        return s
    m = _MODEL_NAME_RE.match(s.upper())
    return m.group(1) if m else s


# ------------------------------------------------------------------
# Pages
# ------------------------------------------------------------------
@app.route("/")
def index():
    if current_user() is None:
        return redirect("/login")
    return render_template("index.html")


@app.route("/login")
def login_page():
    if current_user() is not None:
        return redirect("/")
    return render_template("login.html")


@app.route("/register")
def register_page():
    return render_template("register.html")


# ------------------------------------------------------------------
# Auth API (dang ky / dang nhap / dang xuat / doi mat khau / thong tin tai khoan)
# ------------------------------------------------------------------
@app.route("/api/auth/register", methods=["POST"])
def api_register():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""   # chap nhan moi ky tu
    if not username or not password:
        return jsonify({"error": "Vui lòng nhập tên tài khoản và mật khẩu."}), 400
    # Ten dang ky phai trung danh sach owner (con active) trong tracker.db.
    if not owner_is_active(username):
        return jsonify({"error": "Tên này không có trong danh sách Owner (hoặc đã ngừng hoạt động) — không thể đăng ký."}), 400
    udb = get_users_db()
    if udb.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
        return jsonify({"error": "Tài khoản này đã được đăng ký."}), 400
    udb.execute(
        "INSERT INTO users (username, password_hash, role, permissions, active) VALUES (?,?,?,?,1)",
        (username, generate_password_hash(password), "user", ",".join(ROLE_DEFAULT_PERMS["user"])),
    )
    udb.commit()
    return jsonify({"status": "ok", "username": username})


@app.route("/api/auth/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    udb = get_users_db()
    row = udb.execute(
        "SELECT username, password_hash, active FROM users WHERE username=?", (username,)
    ).fetchone()
    if not row or not check_password_hash(row["password_hash"], password):
        return jsonify({"error": "Sai tên tài khoản hoặc mật khẩu."}), 401
    if row["active"] != 1:
        return jsonify({"error": "Tài khoản đã bị vô hiệu hoá."}), 403
    if not owner_is_active(username):
        return jsonify({"error": "Owner tương ứng đã ngừng hoạt động — không thể đăng nhập."}), 403
    session["user"] = username
    session.permanent = True
    return jsonify({"status": "ok"})


@app.route("/api/auth/logout", methods=["POST"])
def api_logout():
    session.pop("user", None)
    return jsonify({"status": "ok"})


@app.route("/api/me")
def api_me():
    u = current_user()
    if u is None:
        return jsonify({"error": "Chưa đăng nhập"}), 401
    return jsonify(u)


@app.route("/api/auth/change-password", methods=["POST"])
@require_login
def api_change_password():
    data = request.get_json(force=True)
    current_pw = data.get("current") or ""
    new_pw = data.get("new") or ""     # chap nhan moi ky tu
    if not new_pw:
        return jsonify({"error": "Mật khẩu mới không được rỗng."}), 400
    name = session.get("user")
    udb = get_users_db()
    row = udb.execute("SELECT password_hash FROM users WHERE username=?", (name,)).fetchone()
    if not row or not check_password_hash(row["password_hash"], current_pw):
        return jsonify({"error": "Mật khẩu hiện tại không đúng."}), 400
    udb.execute("UPDATE users SET password_hash=? WHERE username=?",
                (generate_password_hash(new_pw), name))
    udb.commit()
    return jsonify({"status": "ok"})


# ------------------------------------------------------------------
# Reference data
# ------------------------------------------------------------------
@app.route("/api/models")
def api_models():
    db = get_db()
    return jsonify(get_models_list(db))


@app.route("/api/owners")
def api_owners():
    db = get_db()
    rows = db.execute(
        "SELECT name, active, team FROM owners WHERE active=1 ORDER BY name"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/scripts")
def api_scripts():
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT test_suite, test_case FROM results ORDER BY test_suite, test_case"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/cycles/latest")
def api_latest_cycle():
    db = get_db()
    row = db.execute("SELECT MAX(cycle) as c FROM results").fetchone()
    return jsonify({"latest_cycle": row["c"] or 0})


# ------------------------------------------------------------------
# Master lists: Test Suites / Models / Owners  (Cài đặt)
# Không có xác thực đăng nhập — bất kỳ ai mở trang cũng chỉnh được các danh mục này.
# ------------------------------------------------------------------
@app.route("/api/lists", methods=["GET"])
def api_lists():
    db = get_db()
    suite_rows = db.execute("SELECT name, script_path FROM test_suites ORDER BY name").fetchall()
    test_suites = [r["name"] for r in suite_rows]
    # Kem script_path (duong dan thu muc script tren repo GitHub) cho bang Cai dat + doi chieu.
    test_suites_detail = [{"name": r["name"], "script_path": r["script_path"] or ""} for r in suite_rows]
    models = get_models_list(db)
    owners = [dict(r) for r in db.execute("SELECT name, active, team FROM owners ORDER BY name")]
    return jsonify({
        "test_suites": test_suites,
        "test_suites_detail": test_suites_detail,
        "models": models,
        "owners": owners,
        "root_cause_groups": get_root_cause_groups(db),
    })


@app.route("/api/lists/test_suites", methods=["POST"])
@require_perm("settings")
def add_test_suite():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Ten khong duoc rong"}), 400
    db = get_db()
    db.execute("INSERT OR IGNORE INTO test_suites (name) VALUES (?)", (name,))
    log_audit(db, "lists.suite_add", target=name)
    db.commit()
    return jsonify({"status": "ok"})


@app.route("/api/lists/test_suites/<path:old_name>", methods=["PUT", "DELETE"])
@require_perm("settings")
def edit_test_suite(old_name):
    db = get_db()
    if request.method == "DELETE":
        used = db.execute("SELECT COUNT(*) c FROM results WHERE test_suite=?", (old_name,)).fetchone()["c"]
        if used:
            return jsonify({"error": f"Khong the xoa: da co {used} ket qua dung Test suite nay. Doi ten thay vi xoa."}), 400
        db.execute("DELETE FROM test_suites WHERE name=?", (old_name,))
        log_audit(db, "lists.suite_delete", target=old_name)
        db.commit()
        return jsonify({"status": "deleted"})
    data = request.get_json(force=True)
    new_name = (data.get("new_name") or "").strip()
    if not new_name:
        return jsonify({"error": "Ten moi khong duoc rong"}), 400
    db.execute("UPDATE OR IGNORE test_suites SET name=? WHERE name=?", (new_name, old_name))
    db.execute("DELETE FROM test_suites WHERE name=?", (old_name,))
    db.execute("INSERT OR IGNORE INTO test_suites (name) VALUES (?)", (new_name,))
    db.execute("UPDATE results SET test_suite=? WHERE test_suite=?", (new_name, old_name))
    db.execute("UPDATE fixes SET test_suite=? WHERE test_suite=?", (new_name, old_name))
    log_audit(db, "lists.suite_rename", target=old_name, detail=f"-> {new_name}")
    db.commit()
    return jsonify({"status": "renamed", "affected_results": db.total_changes})


@app.route("/api/lists/test_suites/<path:name>/path", methods=["PUT"])
@require_perm("settings")
def set_suite_script_path(name):
    """Gan/cap nhat duong dan thu muc script cua 1 Item tren repo GitHub (nhanh main).
    Dung khi doi chieu 3 chieu: file script cua Item phai nam trong thu muc nay."""
    data = request.get_json(force=True)
    script_path = (data.get("script_path") or "").strip().strip("/")
    db = get_db()
    db.execute("INSERT OR IGNORE INTO test_suites (name) VALUES (?)", (name,))
    db.execute("UPDATE test_suites SET script_path=? WHERE name=?", (script_path, name))
    log_audit(db, "lists.suite_path", target=name, detail=script_path)
    db.commit()
    return jsonify({"status": "ok", "script_path": script_path})


def _persist_root_cause_groups(db, groups):
    """Ghi danh sach nhom nguyen nhan hien hanh vao settings key 'root_cause_groups' (JSON)."""
    db.execute(
        "INSERT INTO settings (key, value) VALUES ('root_cause_groups', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (json.dumps(groups, ensure_ascii=False),),
    )
    _invalidate_settings_cache()


@app.route("/api/lists/root_cause_groups", methods=["POST"])
@require_perm("settings")
def add_root_cause_group():
    """Them 1 nhom nguyen nhan goc moi (mirror add_model)."""
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Tên nhóm không được rỗng"}), 400
    db = get_db()
    groups = get_root_cause_groups(db)
    if name in groups:
        return jsonify({"error": f"Nhóm '{name}' đã tồn tại."}), 400
    new_groups = groups + [name]
    _persist_root_cause_groups(db, new_groups)
    log_audit(db, "lists.root_cause_group_add", target=name)
    db.commit()
    return jsonify({"status": "ok", "groups": new_groups})


@app.route("/api/lists/root_cause_groups/<path:old_name>", methods=["PUT", "DELETE"])
@require_perm("settings")
def edit_root_cause_group(old_name):
    """Doi ten hoac xoa 1 nhom nguyen nhan goc. Persist danh sach vao settings
    (key 'root_cause_groups', JSON). Rename cascade cap nhat cac dong fixes cu
    (cot root_cause_group VA prefix text trong root_cause) de Pareto/consumer khong lech.
    Delete chi cho phep khi KHONG co fix nao dang dung nhom nay va con >=1 nhom sau khi xoa."""
    db = get_db()
    groups = get_root_cause_groups(db)
    if old_name not in groups:
        return jsonify({"error": f"Không tìm thấy nhóm '{old_name}'."}), 404

    if request.method == "DELETE":
        used = db.execute("SELECT COUNT(*) c FROM fixes WHERE root_cause_group=?", (old_name,)).fetchone()["c"]
        if used:
            return jsonify({"error": f"Không thể xoá: đã có {used} fix dùng nhóm này. Đổi tên thay vì xoá."}), 400
        new_groups = [g for g in groups if g != old_name]
        if not new_groups:
            return jsonify({"error": "Phải giữ ít nhất 1 nhóm nguyên nhân."}), 400
        _persist_root_cause_groups(db, new_groups)
        log_audit(db, "lists.root_cause_group_delete", target=old_name)
        db.commit()
        return jsonify({"status": "deleted", "groups": new_groups})

    data = request.get_json(force=True)
    new_name = (data.get("new_name") or "").strip()
    if not new_name:
        return jsonify({"error": "Ten moi khong duoc rong"}), 400
    if new_name != old_name and new_name in groups:
        return jsonify({"error": f"Nhom '{new_name}' da ton tai — chon ten khac."}), 400
    if new_name == old_name:
        return jsonify({"status": "unchanged"})
    new_groups = [new_name if g == old_name else g for g in groups]
    _persist_root_cause_groups(db, new_groups)
    # Cascade: cot root_cause_group + prefix "<old> - " trong text root_cause.
    db.execute("UPDATE fixes SET root_cause_group=? WHERE root_cause_group=?", (new_name, old_name))
    db.execute(
        "UPDATE fixes SET root_cause = ? || SUBSTR(root_cause, ?) WHERE root_cause LIKE ?",
        (new_name + " - ", len(old_name + " - ") + 1, old_name + " - %"),
    )
    log_audit(db, "lists.root_cause_group_rename", target=old_name, detail=f"-> {new_name}")
    db.commit()
    return jsonify({"status": "renamed", "groups": new_groups})


@app.route("/api/lists/models", methods=["POST"])
@require_perm("settings")
def add_model():
    data = request.get_json(force=True)
    name = normalize_model_name((data.get("name") or "").strip())
    if not name:
        return jsonify({"error": "Ten khong duoc rong"}), 400
    db = get_db()
    max_order = db.execute("SELECT MAX(sort_order) m FROM models").fetchone()["m"] or 0
    db.execute("INSERT OR IGNORE INTO models (name, sort_order) VALUES (?, ?)", (name, max_order + 1))
    log_audit(db, "lists.model_add", target=name)
    db.commit()
    return jsonify({"status": "ok"})


@app.route("/api/lists/models/<path:old_name>", methods=["PUT", "DELETE"])
@require_perm("settings")
def edit_model(old_name):
    db = get_db()
    if request.method == "DELETE":
        used = db.execute("SELECT COUNT(*) c FROM results WHERE model=?", (old_name,)).fetchone()["c"]
        if used:
            return jsonify({"error": f"Khong the xoa: da co {used} ket qua dung Model nay. Doi ten (thay the) thay vi xoa."}), 400
        db.execute("DELETE FROM models WHERE name=?", (old_name,))
        log_audit(db, "lists.model_delete", target=old_name)
        db.commit()
        return jsonify({"status": "deleted"})
    data = request.get_json(force=True)
    new_name = normalize_model_name((data.get("new_name") or "").strip())
    if not new_name:
        return jsonify({"error": "Ten moi khong duoc rong"}), 400
    row = db.execute("SELECT sort_order FROM models WHERE name=?", (old_name,)).fetchone()
    order = row["sort_order"] if row else 0
    db.execute("DELETE FROM models WHERE name=?", (old_name,))
    db.execute("INSERT OR IGNORE INTO models (name, sort_order) VALUES (?, ?)", (new_name, order))
    db.execute("UPDATE results SET model=? WHERE model=?", (new_name, old_name))
    db.execute("UPDATE fixes SET model_fixed=? WHERE model_fixed=?", (new_name, old_name))
    log_audit(db, "lists.model_rename", target=old_name, detail=f"-> {new_name}")
    db.commit()
    return jsonify({"status": "renamed (thay the model cu bang model moi, giu nguyen lich su ket qua)"})


# ------------------------------------------------------------------
# Owner operations - helper dung chung cho ca route Cai dat (session-gated)
# va route quan tri (ADMIN_KEY-gated). Bao gom dong bo owner <-> tai khoan (users.db).
# ------------------------------------------------------------------
def owner_op_add(db, name, team):
    if not name:
        return {"error": "Ten khong duoc rong"}, 400
    db.execute("INSERT OR IGNORE INTO owners (name, active, team) VALUES (?, 1, ?)", (name, team))
    if team:
        db.execute("UPDATE owners SET team=? WHERE name=?", (team, name))
    log_audit(db, "owner.add", target=name, detail=f"team={team}")
    db.commit()
    # Tao tai khoan mat dinh cho owner moi (neu chua co)
    udb = get_users_db()
    ensure_account_for_owner(udb, name)
    udb.commit()
    return {"status": "ok"}, 200


def owner_op_set_team(db, name, team):
    db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (name,))
    db.execute("UPDATE owners SET team=? WHERE name=?", (team, name))
    log_audit(db, "owner.set_team", target=name, detail=f"team={team}")
    db.commit()
    return {"status": "ok", "team": team}, 200


def owner_op_rename(db, old_name, new_name):
    if not new_name:
        return {"error": "Ten moi khong duoc rong"}, 400
    # "Rename" = SAME nguoi, chi doi ten hien thi -> cascade toan bo, ke ca assignment dang mo.
    old_row = db.execute("SELECT team, active FROM owners WHERE name=?", (old_name,)).fetchone()
    old_team = old_row["team"] if old_row else None
    old_active = old_row["active"] if old_row else 1
    db.execute("DELETE FROM owners WHERE name=?", (old_name,))
    db.execute("INSERT OR IGNORE INTO owners (name, active, team) VALUES (?, ?, ?)", (new_name, old_active, old_team))
    db.execute("UPDATE fixes SET owner=? WHERE owner=?", (new_name, old_name))
    db.execute("UPDATE assignments SET owner=? WHERE owner=?", (new_name, old_name))
    log_audit(db, "owner.rename", target=old_name, detail=f"-> {new_name}")
    db.commit()
    # Dong bo: doi ten tai khoan tuong ung (neu co) de username khong lech voi owner.
    udb = get_users_db()
    if udb.execute("SELECT 1 FROM users WHERE username=?", (old_name,)).fetchone():
        if not udb.execute("SELECT 1 FROM users WHERE username=?", (new_name,)).fetchone():
            udb.execute("UPDATE users SET username=? WHERE username=?", (new_name, old_name))
            udb.commit()
    return {"status": "renamed", "note": "Lich su fix + script dang phu trach + tai khoan da chuyen sang ten moi."}, 200


def owner_op_deactivate(db, name):
    # Soft-delete (deactivate) de giu lich su fix; dong bo vo hieu hoa tai khoan.
    db.execute("UPDATE owners SET active=0 WHERE name=?", (name,))
    log_audit(db, "owner.deactivate", target=name)
    db.commit()
    udb = get_users_db()
    udb.execute("UPDATE users SET active=0 WHERE username=?", (name,))
    udb.commit()
    return {"status": "deactivated"}, 200


def owner_op_hard_delete(db, name):
    # Xoa han - CHI khi owner khong con tham chieu o results/fixes/assignments.
    refs = 0
    for tbl, col in (("results", "author"), ("fixes", "owner"), ("assignments", "owner")):
        refs += db.execute(f"SELECT COUNT(*) c FROM {tbl} WHERE {col}=?", (name,)).fetchone()["c"]
    if refs:
        return {"error": f"Khong the xoa han: owner con {refs} tham chieu (ket qua/fix/assignment). Dung 'Ngung hoat dong' thay the."}, 400
    db.execute("DELETE FROM owners WHERE name=?", (name,))
    log_audit(db, "owner.hard_delete", target=name)
    db.commit()
    udb = get_users_db()
    udb.execute("DELETE FROM users WHERE username=?", (name,))
    udb.commit()
    return {"status": "deleted"}, 200


@app.route("/api/lists/owners", methods=["POST"])
@require_perm("settings")
def add_owner():
    data = request.get_json(force=True)
    body, code = owner_op_add(get_db(), (data.get("name") or "").strip(), (data.get("team") or "").strip())
    return jsonify(body), code


@app.route("/api/lists/owners/<path:name>/team", methods=["PUT"])
@require_perm("settings")
def set_owner_team(name):
    """Gan/cap nhat Team cho 1 owner trong Cai dat. Team la nhom nho (thuong dat theo
    ten team-lead). Import ket qua moi (co cot Team) cung tu dong cap nhat truong nay."""
    data = request.get_json(force=True)
    body, code = owner_op_set_team(get_db(), name, (data.get("team") or "").strip())
    return jsonify(body), code


@app.route("/api/lists/owners/<path:old_name>", methods=["PUT", "DELETE"])
@require_perm("settings")
def edit_owner(old_name):
    db = get_db()
    if request.method == "DELETE":
        body, code = owner_op_deactivate(db, old_name)
        return jsonify(body), code
    data = request.get_json(force=True)
    body, code = owner_op_rename(db, old_name, (data.get("new_name") or "").strip())
    return jsonify(body), code


# ------------------------------------------------------------------
# Settings
# ------------------------------------------------------------------
@app.route("/api/settings", methods=["GET", "POST"])
def api_settings():
    db = get_db()
    if request.method == "POST":
        err = perm_error("settings")
        if err:
            return err
        data = request.get_json(force=True)
        changed = []
        for k, v in data.items():
            if k == "_owner":
                continue
            # Token nhay cam: client gui lai gia tri mask nghia la "khong doi" -> bo qua.
            if k in SENSITIVE_SETTINGS and str(v) == SETTINGS_MASK:
                continue
            db.execute(
                "INSERT INTO settings (key, value) VALUES (?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                (k, str(v)),
            )
            changed.append(k)
        _invalidate_settings_cache()
        if changed:
            log_audit(db, "settings.update", detail="keys: " + ", ".join(sorted(changed)))
        db.commit()
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    out = {}
    for r in rows:
        # Khong tra token plaintext ve moi client dang nhap - mask neu da co gia tri.
        if r["key"] in SENSITIVE_SETTINGS and r["value"]:
            out[r["key"]] = SETTINGS_MASK
        else:
            out[r["key"]] = r["value"]
    return jsonify(out)


@app.route("/api/settings/apply-whitelist", methods=["POST"])
@require_perm("settings")
def api_apply_whitelist():
    """Ap dung whitelist loi nhieu cho DU LIEU DA CO: quet cac ket qua Pass/Fail co mo ta
    khop whitelist -> reclass sang result='Excluded' (loai khoi thong ke). Dung khi nguoi
    dung vua tinh chinh whitelist va muon loai ca noise cu. Chi dong theo 1 chieu (khong tu
    dua Excluded ve lai Pass/Fail)."""
    db = get_db()
    whitelist = get_error_whitelist(db)
    if not whitelist:
        return jsonify({"error": "Whitelist đang trống — nhập ít nhất 1 cụm từ trong ô 'Whitelist lỗi nhiễu' rồi lưu trước khi áp dụng."}), 400
    updated = 0
    for r in db.execute(
        "SELECT id, description FROM results WHERE result IN ('Pass','Fail')"
    ).fetchall():
        if matches_whitelist(r["description"], whitelist):
            db.execute("UPDATE results SET result='Excluded' WHERE id=?", (r["id"],))
            updated += 1
    log_audit(db, "results.apply_whitelist", detail=f"reclassified {updated} rows to Excluded")
    db.commit()
    return jsonify({"status": "ok", "updated": updated})


# ------------------------------------------------------------------
# New scripts (Script viet moi) - ghi nhan script duoc viet moi cho tung Test Case
# ------------------------------------------------------------------
@app.route("/api/new-scripts/items")
def api_new_script_items():
    """Danh sach Item co dinh + tien to TC ID + dau noi + vi du dinh dang, dung tu
    TC_SUITE_RULES / _TC_ID_PREFIX_SEPARATORS. Frontend dung de tu suy Item, validate
    dinh dang va hien goi y (mot nguon su that, JS khong tu hardcode lai)."""
    out = []
    for prefixes, suite in TC_SUITE_RULES:
        out.append({
            "item": suite,
            "prefixes": list(prefixes),
            "separators": {p: list(_TC_ID_PREFIX_SEPARATORS.get(p, ("_",))) for p in prefixes},
            "example": ITEM_TC_EXAMPLES.get(suite, prefixes[0] + "_000001"),
        })
    return jsonify(out)


def validate_new_script_row(db, data, exclude_id=None):
    """Validate + chuan hoa 1 dong new_scripts tu dict tho. Tra ve (row_dict, None) neu
    hop le, hoac (None, error_message) neu khong. Dung chung cho POST don le, bulk import,
    va sua dong da co (exclude_id = id cua chinh dong do, de bo qua check trung tc_id voi
    chinh no) - mot nguon su that, tranh lech logic giua cac duong nhap/sua."""
    raw_tc = str(data.get("tc_id") or "").strip()
    if not raw_tc:
        return None, "Vui lòng nhập TC ID."
    tc_id = extract_test_case_name(raw_tc)
    item = item_from_tc_id(tc_id)
    if not item:
        examples = ", ".join(ITEM_TC_EXAMPLES.values())
        return None, f"TC ID '{tc_id}' không đúng định dạng — không nhận diện được Item. Ví dụ hợp lệ: {examples}."

    dup_q = "SELECT 1 FROM new_scripts WHERE tc_id=?" + (" AND id!=?" if exclude_id else "")
    dup_args = (tc_id, exclude_id) if exclude_id else (tc_id,)
    if db.execute(dup_q, dup_args).fetchone():
        return None, f"TC ID '{tc_id}' đã được nhập trước đó — mỗi Test Case chỉ ghi nhận 1 lần."

    status = str(data.get("status") or "").strip().upper()
    if status not in ("DONE", "SKIP", "ASSIGNED"):
        return None, "Status phải là DONE, SKIP hoặc ASSIGNED."

    remark = str(data.get("remark") or "").strip()
    if status == "SKIP" and not remark:
        return None, "Bắt buộc nhập Remark (lý do) khi Status = SKIP."

    models_written = data.get("models_written") or []
    if isinstance(models_written, str):
        models_written = re.split(r"[,/]", models_written)
    models_written = [normalize_model_name(str(m).strip()) for m in models_written if str(m).strip()]
    if status == "DONE" and not models_written:
        return None, "Chọn ít nhất 1 model đã viết script khi Status = DONE."
    models_csv = ", ".join(models_written)

    member = str(data.get("member") or "").strip()
    team = str(data.get("team") or "").strip()
    if member and not team:
        row = db.execute("SELECT team FROM owners WHERE name=?", (member,)).fetchone()
        team = row["team"] if row and row["team"] else ""

    completed_date = str(data.get("completed_date") or "").strip()
    assign_week = data.get("assign_week")
    try:
        assign_week = int(assign_week) if assign_week not in (None, "") else None
    except (ValueError, TypeError):
        assign_week = None
    if assign_week is None and completed_date:
        try:
            assign_week = date.fromisoformat(completed_date[:10]).isocalendar()[1]
        except ValueError:
            assign_week = None

    sdf_id = str(data.get("sdf_id") or "").strip()
    created_by = str(data.get("created_by") or "").strip() or member

    return {
        "item": item, "tc_id": tc_id, "member": member, "team": team,
        "assign_week": assign_week, "completed_date": completed_date, "status": status,
        "models_written": models_csv, "sdf_id": sdf_id, "remark": remark, "created_by": created_by,
    }, None


def insert_new_script_row(db, row):
    db.execute(
        "INSERT INTO new_scripts (item, tc_id, member, team, assign_week, completed_date, status, models_written, sdf_id, remark, created_by) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (row["item"], row["tc_id"], row["member"], row["team"], row["assign_week"], row["completed_date"],
         row["status"], row["models_written"], row["sdf_id"], row["remark"], row["created_by"]),
    )
    if row["member"]:
        db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (row["member"],))
        if row["team"]:
            db.execute("UPDATE owners SET team=? WHERE name=?", (row["team"], row["member"]))


@app.route("/api/new-scripts", methods=["GET", "POST"])
def api_new_scripts():
    db = get_db()
    if request.method == "GET":
        rows = db.execute(
            "SELECT * FROM new_scripts ORDER BY completed_date DESC, id DESC"
        ).fetchall()
        return jsonify([dict(r) for r in rows])

    err = perm_error("new-scripts")
    if err:
        return err
    data = request.get_json(force=True)
    if str(data.get("status", "")).strip().upper() == "ASSIGNED":
        err = perm_error("ns-assign")
        if err:
            return err
    row, error = validate_new_script_row(db, data)
    if error:
        return jsonify({"error": error}), 400
    insert_new_script_row(db, row)
    db.commit()
    return jsonify({"status": "ok", "item": row["item"], "tc_id": row["tc_id"], "assign_week": row["assign_week"], "team": row["team"]})


@app.route("/api/new-scripts/<int:sid>", methods=["PUT"])
@require_login
def api_update_new_script(sid):
    """Sua lai 1 dong new_scripts da ghi nhan. 3 cap quyen: (1) admin/moderator hoac tai
    khoan co quyen 'ns-edit' -> sua duoc moi field ke ca 'member' (vd de 'assign lai' cho
    nguoi khac); (2) tai khoan co username == member cua chinh dong do (self-edit) -> chi
    duoc cap nhat ket qua/trang thai cua chinh minh (status/models_written/completed_date/
    assign_week/sdf_id/remark), KHONG duoc doi member (khong tu reassign di nguoi khac).
    tc_id/item luon bat bien qua route nay (rieng biet voi route admin PUT X-Admin-Key,
    cho sua moi field ke ca tc_id)."""
    u = current_user()
    db = get_db()
    row = db.execute("SELECT * FROM new_scripts WHERE id=?", (sid,)).fetchone()
    if not row:
        return jsonify({"error": "Không tìm thấy dòng này."}), 404

    is_admin_or_mod = u["role"] in ("admin", "moderator")
    is_ns_edit = "ns-edit" in u["permissions"]
    is_self = bool(row["member"]) and row["member"] == u["username"]
    can_reassign = is_admin_or_mod or is_ns_edit
    if not (can_reassign or is_self):
        return jsonify({"error": "Bạn không có quyền sửa dòng này."}), 403

    data = request.get_json(force=True)
    if "member" in data and not can_reassign:
        new_member = str(data["member"] or "").strip()
        if new_member != (row["member"] or ""):
            return jsonify({"error": "Bạn không có quyền đổi Member (reassign) của dòng này."}), 403

    if str(data.get("status", row["status"])).strip().upper() == "ASSIGNED" and row["status"] != "ASSIGNED":
        err = perm_error("ns-assign")
        if err:
            return err

    merged = dict(row)
    for field in ("member", "status", "models_written", "completed_date", "assign_week", "sdf_id", "remark"):
        if field in data:
            merged[field] = data[field]
    merged["tc_id"] = row["tc_id"]
    if "member" in data and str(merged["member"] or "").strip() != (row["member"] or ""):
        merged["team"] = ""  # buoc validate_new_script_row tu suy lai team theo member moi

    row_out, error = validate_new_script_row(db, merged, exclude_id=sid)
    if error:
        return jsonify({"error": error}), 400

    db.execute(
        "UPDATE new_scripts SET item=?, member=?, team=?, assign_week=?, completed_date=?, "
        "status=?, models_written=?, sdf_id=?, remark=? WHERE id=?",
        (row_out["item"], row_out["member"], row_out["team"], row_out["assign_week"],
         row_out["completed_date"], row_out["status"], row_out["models_written"],
         row_out["sdf_id"], row_out["remark"], sid),
    )
    if row_out["member"] and row_out["member"] != (row["member"] or ""):
        db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (row_out["member"],))
        if row_out["team"]:
            db.execute("UPDATE owners SET team=? WHERE name=?", (row_out["team"], row_out["member"]))
    db.commit()
    return jsonify({"status": "updated"})


@app.route("/api/new-scripts/bulk-assign", methods=["POST"])
@require_perm("ns-assign")
def api_bulk_assign_new_scripts():
    """Assign hang loat: nhan { pairs: [ {tc_id, member}, ... ] }, tao dong new_scripts moi
    voi status=ASSIGNED cho tung cap. Dung chung validate_new_script_row/insert_new_script_row
    nen tu dong co check trung TC ID (ke ca trung trong cung batch)."""
    payload = request.get_json(force=True)
    pairs = payload.get("pairs", [])
    if not pairs:
        return jsonify({"error": "Không có dòng nào."}), 400
    db = get_db()
    inserted = 0
    errors = []
    for i, p in enumerate(pairs):
        data = {"tc_id": p.get("tc_id", ""), "member": p.get("member", ""), "status": "ASSIGNED"}
        row, error = validate_new_script_row(db, data)
        if error:
            errors.append({"row_index": i, "tc_id": p.get("tc_id", ""), "error": error})
            continue
        insert_new_script_row(db, row)
        inserted += 1
    db.commit()
    return jsonify({"inserted": inserted, "errors": errors})


# ------------------------------------------------------------------
# Results (History_Log equivalent)
# ------------------------------------------------------------------
def insert_result_rows(db, rows, created_by):
    """Chen danh sach row ket qua vao bang results (logic dung chung cho paste UI,
    /api/results/import va farm fetch). Moi row: {cycle_date?, test_id, model, test_suite,
    test_case, state, description?, author?, team?}.

    KHONG commit va KHONG recompute_cycles - caller tu lam sau khi chen xong.
    Tra ve dict {inserted, skipped_running, skipped_duplicate, duplicates, errors, warnings}.
    """
    min_row = db.execute(
        "SELECT MIN(cycle_date) d FROM results WHERE cycle_date IS NOT NULL AND cycle_date != ''"
    ).fetchone()
    existing_min_date = min_row["d"] if min_row else None

    whitelist = get_error_whitelist(db)

    inserted = 0
    skipped_running = 0
    skipped_duplicate = 0
    excluded = 0
    errors = []
    duplicates = []
    warnings = []
    backdated = False
    for i, row in enumerate(rows):
        try:
            model = normalize_model_name(str(row["model"]).strip())
            raw_test_suite = str(row["test_suite"]).strip()
            raw_test_case = str(row["test_case"]).strip()
            state = str(row["state"]).strip()
            if not (model and raw_test_suite and raw_test_case and state):
                raise ValueError("Thieu du lieu bat buoc (model/test_suite/test_case/state)")

            if is_skip_state(state):
                skipped_running += 1
                continue

            test_case = extract_test_case_name(raw_test_case)
            test_suite = derive_test_suite(test_case, raw_test_suite)
            test_id = str(row.get("test_id") or "")
            description = str(row.get("description") or "")

            # Kiem tra trung: cung Test ID + Test Case + Model + State + Description
            # thi coi la ban ghi da co san, bo qua khong chen lai (tranh nhan doi Fail_Count).
            existing = db.execute(
                "SELECT 1 FROM results WHERE test_id=? AND test_case=? AND model=? AND state=? AND description=? LIMIT 1",
                (test_id, test_case, model, state, description),
            ).fetchone()
            if existing:
                skipped_duplicate += 1
                duplicates.append({
                    "row_index": i, "test_id": test_id, "model": model,
                    "test_suite": test_suite, "test_case": test_case, "state": state,
                })
                continue

            # Cycle_date uu tien suy tu ngay ma hoa trong Test ID (VD 260706-... -> 2026-07-06);
            # neu Test ID khong ma hoa ngay thi dung ngay chay nguoi dung nhap (fallback).
            cycle_date = extract_date_from_test_id(test_id) or str(row.get("cycle_date") or date.today().isoformat())
            if existing_min_date and cycle_date < existing_min_date:
                backdated = True
            # Cycle (so thu tu) se duoc danh so lai theo ngay o cuoi ham -> dat tam 0.
            cycle = 0
            author = str(row.get("author") or "").strip()
            team = str(row.get("team") or "").strip()
            serial = str(row.get("serial") or "").strip()
            result = classify_result(state)
            # Whitelist loi nhieu: mo ta khop -> luu nhung KHONG tinh vao thong ke (result='Excluded').
            if matches_whitelist(description, whitelist):
                result = "Excluded"
                excluded += 1
            db.execute(
                "INSERT INTO results (cycle, cycle_date, test_id, model, test_suite, test_case, state, description, result, created_by, author, team, serial) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (cycle, cycle_date, test_id, model, test_suite, test_case, state, description, result, created_by, author, team, serial),
            )
            db.execute("INSERT OR IGNORE INTO test_suites (name) VALUES (?)", (test_suite,))
            if db.execute("SELECT 1 FROM models WHERE name=?", (model,)).fetchone() is None:
                max_order = db.execute("SELECT MAX(sort_order) m FROM models").fetchone()["m"] or 0
                db.execute("INSERT OR IGNORE INTO models (name, sort_order) VALUES (?, ?)", (model, max_order + 1))
            if author:
                # Author chinh la Owner cua script - tu dong dang ky vao danh sach owners,
                # kem ten Team hien tai cua nguoi do.
                db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (author,))
                if team:
                    db.execute("UPDATE owners SET team=? WHERE name=?", (team, author))
                # Tu dong gan "dang phu trach" trong Bang uu tien = Author cua lan nhap gan nhat.
                db.execute(
                    "INSERT INTO assignments (test_suite, test_case, owner, assigned_date) VALUES (?,?,?,?) "
                    "ON CONFLICT(test_suite, test_case) DO UPDATE SET owner=excluded.owner, assigned_date=excluded.assigned_date",
                    (test_suite, test_case, author, cycle_date),
                )
            inserted += 1
        except Exception as e:
            errors.append({"row_index": i, "error": str(e), "row": row})

    if backdated and inserted:
        warnings.append(
            "Có kết quả với ngày SỚM HƠN dữ liệu hiện có — toàn bộ số cycle sẽ được đánh lại "
            "theo ngày, các fix đã ghi 'fixed_after_cycle' có thể lệch số cycle. Kiểm tra lại tab Theo dõi Fix."
        )
    return {
        "inserted": inserted,
        "skipped_running": skipped_running,
        "skipped_duplicate": skipped_duplicate,
        "excluded": excluded,
        "duplicates": duplicates,
        "errors": errors,
        "warnings": warnings,
    }


@app.route("/api/results", methods=["POST"])
@require_perm("input-results")
def api_add_results():
    """
    Accepts:
    { "rows": [ {cycle, cycle_date, test_id, model, test_suite, test_case, state, description, author, team}, ... ], "created_by": "name" }

    test_suite/test_case duoc gui len o dang tho (co the la duong dan day du, ten file
    .ts/.py, hoac bi dien nham) - server tu trich xuat lai ten chuan:
      - test_case: lay basename, bo duoi ".py"
      - test_suite: uu tien suy tu tien to ten test_case (Browser/Internet->Internet,
        Keyboard/SKBD->Keyboard, Weather->Weather, Wallpaper->Wallpaper, SM->SamsungMember,
        Reminder->Reminder, NowBrief->Now-brief), fallback parse cot tho neu khong khop.
    State "RUNNING" (chua chay xong) se bi bo qua, khong luu vao database.
    Neu co "author": tu dong dang ky owner + gan author do lam nguoi "dang phu trach"
    (assignment) cho script trong Bang uu tien - ghi de moi lan nhap moi (phan anh dung
    ai vua cham vao script nay gan nhat), khong anh huong lich su fix (Daily_Fix_Log).
    """
    payload = request.get_json(force=True)
    rows = payload.get("rows", [])
    u = current_user()
    created_by = str(payload.get("created_by") or "").strip() or (u["username"] if u else "(unknown)")
    if not rows:
        return jsonify({"error": "No rows provided"}), 400

    db = get_db()
    summary = insert_result_rows(db, rows, created_by)
    # Danh so lai cycle theo ngay sau khi da chen xong (moi ngay = 1 cycle, tang dan).
    recompute_cycles(db)
    db.commit()
    return jsonify(summary)


@app.route("/api/results", methods=["GET"])
def api_get_results():
    db = get_db()
    limit = int(request.args.get("limit", 500))
    rows = db.execute(
        "SELECT * FROM results ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


# ------------------------------------------------------------------
# Fixes (Daily_Fix_Log equivalent)
# ------------------------------------------------------------------
@app.route("/api/fixes", methods=["POST"])
@require_perm("input-fix")
def api_add_fix():
    row = request.get_json(force=True)
    required = ["fix_date", "owner", "test_suite", "test_case", "model_fixed", "fixed_after_cycle"]
    for f in required:
        if not row.get(f) and row.get(f) != 0:
            return jsonify({"error": "Thieu truong bat buoc: " + f}), 400

    # Root cause (BAT BUOC): uu tien dang chuan hoa {root_cause_group + root_cause_detail}
    # tu dropdown (A3); van chap nhan root_cause free text cu (bulk import/script) -> tu suy
    # group bang heuristic. Luon luu CA root_cause_group (cot moi) VA root_cause (text cu,
    # compose "Group - detail") de moi consumer cu khong doi.
    db = get_db()
    valid_groups = get_root_cause_groups(db)
    group = str(row.get("root_cause_group") or "").strip()
    detail = str(row.get("root_cause_detail") or "").strip()
    root_cause = str(row.get("root_cause") or "").strip()
    sdf_id = str(row.get("sdf_id") or "").strip()
    if group:
        if group not in valid_groups:
            return jsonify({"error": "Nhóm nguyên nhân không hợp lệ. Chọn 1 trong: " + ", ".join(valid_groups)}), 400
        # Chi tiet nguyen nhan BAT BUOC (form UI). Bulk import free-text di nhanh 'elif' ben duoi.
        if not detail:
            return jsonify({"error": "Bắt buộc nhập Chi tiết nguyên nhân (mô tả cụ thể) khi ghi nhận fix."}), 400
        # SDF ID BAT BUOC khi ghi fix qua form (bang chung da chay lai de dam bao fix hieu qua).
        # Bulk import admin di nhanh 'elif' ben duoi -> khong bat buoc (tuong thich du lieu cu).
        if not sdf_id:
            return jsonify({"error": "Bắt buộc nhập SDF ID (bằng chứng đã chạy lại để đảm bảo fix)."}), 400
        root_cause = f"{group} - {detail}"
    elif root_cause:
        group = classify_root_cause_group(root_cause)
    else:
        return jsonify({"error": "Root cause (nguyên nhân lỗi) là bắt buộc — chọn nhóm nguyên nhân và mô tả chi tiết."}), 400
    owner_name = row["owner"].strip()
    test_suite = row["test_suite"].strip()
    test_case = row["test_case"].strip()
    model_fixed = normalize_model_name(str(row["model_fixed"] or "").strip())
    fixed_after_cycle = int(row["fixed_after_cycle"])
    note = row.get("note", "")
    db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (owner_name,))

    # Chong trung: 1 lan fix duoc dinh danh boi (owner, test_suite, test_case, model_fixed,
    # fixed_after_cycle). Neu da ton tai -> KHONG tao dong moi giong het, ma CAP NHAT dong cu
    # (root_cause/note/fix_date moi nhat) de tranh nhan doi trong Theo doi Fix.
    existing = db.execute(
        "SELECT id FROM fixes WHERE owner=? AND test_suite=? AND test_case=? AND model_fixed=? AND fixed_after_cycle=?",
        (owner_name, test_suite, test_case, model_fixed, fixed_after_cycle),
    ).fetchone()
    if existing:
        db.execute(
            "UPDATE fixes SET fix_date=?, note=?, root_cause=?, root_cause_group=?, sdf_id=? WHERE id=?",
            (row["fix_date"], note, root_cause, group, sdf_id, existing["id"]),
        )
        db.commit()
        return jsonify({
            "status": "updated",
            "message": "Fix này đã được ghi nhận trước đó (cùng owner/script/model/cycle) — đã cập nhật thông tin thay vì tạo dòng trùng.",
        })

    db.execute(
        "INSERT INTO fixes (fix_date, owner, test_suite, test_case, model_fixed, fixed_after_cycle, note, root_cause, root_cause_group, sdf_id) "
        "VALUES (?,?,?,?,?,?,?,?,?,?)",
        (row["fix_date"], owner_name, test_suite, test_case, model_fixed, fixed_after_cycle, note, root_cause, group, sdf_id),
    )
    db.commit()
    return jsonify({"status": "ok"})


@app.route("/api/fixes", methods=["GET"])
def api_get_fixes():
    db = get_db()
    limit = int(request.args.get("limit", 500))
    rows = db.execute(
        "SELECT * FROM fixes ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


# ------------------------------------------------------------------
# Fix tracking: mỗi lần fix đã báo -> hết lỗi chưa? hết rồi có fail lại không?
# ------------------------------------------------------------------
# 4 trạng thái rõ ràng, khoa học:
#   pending        (⏳): chưa có lần chạy nào SAU cycle fix để đối chiếu
#   still_failing  (❌): lần chạy ngay sau fix VẪN còn Fail (chưa hết lỗi)
#   verified       (✅): sau fix chạy lại Pass và GIỮ NGUYÊN Pass tới nay (hết lỗi ổn định)
#   regressed      (⚠️): đã hết lỗi ngay sau fix, nhưng cycle sau đó lại Fail (tái lỗi)
FIX_STATUS_LABEL = {
    "pending": "Chờ dữ liệu",
    "still_failing": "Chưa hết lỗi",
    "verified": "Đã hết lỗi",
    "regressed": "Hết rồi fail lại",
}


def _status_from_after_rows(after_rows):
    """after_rows: list dict {cycle, result} của các lần chạy SAU cycle fix, cùng 1 script
    (đã lọc theo model nếu fix theo model cụ thể). Trả về 1 trong 4 trạng thái."""
    if not after_rows:
        return "pending"
    by_cycle = {}
    for r in after_rows:
        by_cycle.setdefault(r["cycle"], []).append(r["result"])
    ordered = sorted(by_cycle.items())  # theo cycle tăng dần
    first_all_pass = all(res == "Pass" for res in ordered[0][1])
    if not first_all_pass:
        return "still_failing"
    for _cycle, results in ordered[1:]:
        if any(res == "Fail" for res in results):
            return "regressed"
    return "verified"


def compute_fix_tracking(db):
    """Làm giàu mỗi dòng fix với trạng thái resolution + thông tin lần chạy sau fix."""
    fixes = db.execute("SELECT * FROM fixes ORDER BY id DESC").fetchall()
    owner_team = {r["name"]: r["team"] for r in db.execute("SELECT name, team FROM owners").fetchall()}
    if not fixes:
        return []
    # Prefetch cac lan chay "sau fix" trong 1 truy van thay vi N truy van/fix (chong N+1).
    # Chi lay results thuoc suite co fix va cycle > nguong nho nhat (min fixed_after_cycle);
    # loc chi tiet theo (case, model, cycle>after_cycle) trong Python - cho ket qua Y HET
    # vong lap cu (status/count/min deu doc lap thu tu nen khong can ORDER BY).
    suites = sorted({f["test_suite"] for f in fixes})
    min_after = min(f["fixed_after_cycle"] for f in fixes)
    ph = ",".join("?" * len(suites))
    by_pair = {}
    for r in db.execute(
        f"SELECT test_suite, test_case, model, cycle, result FROM results "
        f"WHERE test_suite IN ({ph}) AND cycle>? AND result <> 'Excluded'",
        (*suites, min_after),
    ).fetchall():
        by_pair.setdefault((r["test_suite"], r["test_case"]), []).append(r)
    out = []
    for f in fixes:
        suite, case = f["test_suite"], f["test_case"]
        after_cycle = f["fixed_after_cycle"]
        model_fixed = f["model_fixed"]
        pair_rows = by_pair.get((suite, case), ())
        if model_fixed and model_fixed != "All Models":
            after_rows = [{"cycle": r["cycle"], "result": r["result"]} for r in pair_rows
                          if r["model"] == model_fixed and r["cycle"] > after_cycle]
        else:
            after_rows = [{"cycle": r["cycle"], "result": r["result"]} for r in pair_rows
                          if r["cycle"] > after_cycle]
        status = _status_from_after_rows(after_rows)
        next_cycle = min((r["cycle"] for r in after_rows), default=None)
        n_fail_after = sum(1 for r in after_rows if r["result"] == "Fail")
        out.append({
            "id": f["id"], "fix_date": f["fix_date"], "owner": f["owner"],
            "team": owner_team.get(f["owner"]) or "",
            "test_suite": suite, "test_case": case, "model_fixed": model_fixed,
            "fixed_after_cycle": after_cycle, "note": f["note"],
            "root_cause": (f["root_cause"] if "root_cause" in f.keys() else "") or "",
            "root_cause_group": (f["root_cause_group"] if "root_cause_group" in f.keys() else "") or "",
            "sdf_id": (f["sdf_id"] if "sdf_id" in f.keys() else "") or "",
            "status": status, "status_label": FIX_STATUS_LABEL[status],
            "next_cycle_after_fix": next_cycle,
            "runs_after": len(after_rows), "fails_after": n_fail_after,
        })
    return out


@app.route("/api/fix-tracking")
def api_fix_tracking():
    db = get_db()
    return jsonify(compute_fix_tracking(db))


@app.route("/api/failing-scripts")
def api_failing_scripts():
    """Danh sách script HIỆN còn lỗi (tier != Done) để form Ghi nhận Fix chọn trực tiếp,
    kèm các model đang fail và cycle gần nhất (điền sẵn Fixed_after_cycle)."""
    db = get_db()
    priority = get_script_priority(db)
    failing = [
        {
            "test_suite": p["test_suite"], "test_case": p["test_case"],
            "failing_models": p["failing_models"],
            "fail_count": p["fail_count"], "priority_tier": p["priority_tier"],
            "last_updated_cycle": p["last_updated_cycle"],
            "current_owner": p["current_owner"],
        }
        for p in priority if p["priority_tier"] != "Done"
    ]
    failing.sort(key=lambda x: (x["test_suite"], x["test_case"]))
    return jsonify(failing)


# ------------------------------------------------------------------
# Core computed views (mirrors the Excel logic, computed live via SQL)
# ------------------------------------------------------------------
def get_latest_status(db):
    """One row per (test_suite, test_case, model): state at the most recent cycle it ran."""
    return db.execute(
        """
        SELECT test_suite, test_case, model, state, result, cycle
        FROM (
            SELECT *, ROW_NUMBER() OVER (
                PARTITION BY test_suite, test_case, model
                ORDER BY cycle DESC, id DESC
            ) as rn
            FROM results WHERE result <> 'Excluded'
        )
        WHERE rn = 1
        """
    ).fetchall()


def get_script_priority(db):
    """One row per script: total fail_count (sum of all Fail across all cycles/models),
    priority tier, latest per-model status, and current responsible owner (assignment).
    Tier thresholds scale with the current number of models (breadth-of-failure logic).

    fail_count = tổng số lần Fail của script này trên tất cả các model qua tất cả các cycle
    (không chỉ cycle gần nhất) — để owner biết bao nhiêu lần đã chạy xong Fail.

    Exit criteria (B3): script chỉ "Done" khi MỌI model nó từng chạy đều Pass ở
    N cycle gần nhất liên tiếp (N = setting exit_criteria_cycles, mặc định 2; model chạy
    chưa đủ N lần -> chưa Done). Cycle mới nhất không fail nhưng chưa đạt N -> tier "Verify".

    Flaky (A2): script đổi trạng thái pass<->fail >= flaky_min_flips lần trong
    flaky_window cycle gần nhất mà nó CÓ CHẠY -> is_flaky. reopen_count = số lần
    Pass->Fail trên toàn lịch sử (script "tái lỗi" sau khi đã xanh).
    """
    total_models = len(get_models_list(db)) or 5
    exit_n = max(1, get_setting_int(db, "exit_criteria_cycles", 2))
    flaky_window = max(2, get_setting_int(db, "flaky_window", 5))
    flaky_min_flips = max(1, get_setting_int(db, "flaky_min_flips", 2))
    persist_n = max(2, get_setting_int(db, "persistent_fail_cycles", 3))

    # Chuoi verdict theo cycle cua tung (script, model) va tung script (gop moi model).
    # Verdict cua 1 (model, cycle) = ket qua cua LAN CHAY CUOI CUNG trong cycle do
    # (id lon nhat) - NHAT QUAN voi get_latest_status/tier cu: fail roi re-run pass
    # trong cung cycle thi cycle do tinh Pass. (N=1 => Done trung khop hanh vi cu.)
    # Dung chung cho ca exit criteria (per-model) lan flaky/reopen (per-script).
    model_seq = {}   # (suite, case, model) -> [(cycle, fail_bool), ...] theo cycle tang dan
    script_seq = {}  # (suite, case) -> {cycle: fail_bool}
    for row in db.execute(
        """
        SELECT test_suite, test_case, model, cycle, result FROM (
            SELECT *, ROW_NUMBER() OVER (
                PARTITION BY test_suite, test_case, model, cycle
                ORDER BY id DESC
            ) as rn FROM results WHERE result <> 'Excluded'
        ) WHERE rn = 1 ORDER BY cycle
        """
    ).fetchall():
        failed = row["result"] == "Fail"
        model_seq.setdefault((row["test_suite"], row["test_case"], row["model"]), []).append(
            (row["cycle"], failed)
        )
        sc = script_seq.setdefault((row["test_suite"], row["test_case"]), {})
        sc[row["cycle"]] = sc.get(row["cycle"], False) or failed

    # Tính tổng fail_count (tất cả rows Fail) VÀ số model KHÁC NHAU từng fail (độ rộng lỗi)
    # cho mỗi script — cả hai đều tính trên TẤT CẢ model & TẤT CẢ cycle đã chạy.
    fail_counts = {}
    fail_model_breadth = {}
    for row in db.execute(
        "SELECT test_suite, test_case, COUNT(*) as fail_cnt, "
        "COUNT(DISTINCT model) as fail_models FROM results "
        "WHERE result='Fail' GROUP BY test_suite, test_case"
    ).fetchall():
        fail_counts[(row["test_suite"], row["test_case"])] = row["fail_cnt"]
        fail_model_breadth[(row["test_suite"], row["test_case"])] = row["fail_models"]

    # Lấy status gần nhất trên mỗi model (cho hiển thị model_detail + last_cycle)
    latest = get_latest_status(db)
    agg = {}
    for r in latest:
        key = (r["test_suite"], r["test_case"])
        a = agg.setdefault(key, {"pass": 0, "models_seen": 0, "last_cycle": 0, "models": {}})
        a["models_seen"] += 1
        a["last_cycle"] = max(a["last_cycle"], r["cycle"])
        a["models"][r["model"]] = r["result"]  # "Pass" or "Fail" at latest cycle for this model
        if r["result"] == "Pass":
            a["pass"] += 1

    assignments = {
        (row["test_suite"], row["test_case"]): row["owner"]
        for row in db.execute("SELECT test_suite, test_case, owner FROM assignments")
    }
    # Map owner -> team (nhom nho) de gan cot Team cho tung script theo nguoi dang phu trach.
    owner_team = {
        row["name"]: (row["team"] or "")
        for row in db.execute("SELECT name, team FROM owners")
    }

    out = []
    for (suite, case), a in agg.items():
        fail = fail_counts.get((suite, case), 0)  # Tổng lần Fail, không chỉ latest
        not_run = total_models - a["models_seen"]

        # Tier vẫn dựa vào "bao nhiêu model hiện tại còn Fail" (latest status)
        # để phản ánh ưu tiên fix ngay bây giờ
        fail_latest = sum(1 for res in a["models"].values() if res == "Fail")
        if fail_latest == 0:
            # Exit criteria: moi model tung chay phai Pass >= exit_n cycle GAN NHAT lien tiep.
            # Chua du du lieu / chua du chuoi pass -> "Verify" (dang xac minh, chua tinh Done).
            meets_exit = True
            for m in a["models"]:
                seq = model_seq.get((suite, case, m), [])
                if len(seq) < exit_n or any(failed for _cyc, failed in seq[-exit_n:]):
                    meets_exit = False
                    break
            tier = "Done" if meets_exit else "Verify"
        elif fail_latest >= max(total_models - 1, 2):
            tier = "P0"
        elif fail_latest == max(total_models - 2, 1):
            tier = "P1"
        elif fail_latest == max(total_models - 3, 1):
            tier = "P2"
        else:
            tier = "P3"

        failing_models = sorted([m for m, res in a["models"].items() if res == "Fail"])
        breadth = fail_model_breadth.get((suite, case), 0)  # số model KHÁC NHAU từng fail
        # Điểm ưu tiên (kết hợp có trọng số): TỔNG lần fail × ĐỘ RỘNG model từng fail.
        # Case vừa fail nhiều lần vừa fail trên nhiều model -> điểm cao nhất -> ưu tiên trước.
        priority_score = fail * breadth
        current_owner = assignments.get((suite, case)) or ""

        # Flaky + reopen tu chuoi verdict theo cycle cua script (gop moi model)
        verdicts = [failed for _cyc, failed in sorted(script_seq.get((suite, case), {}).items())]
        flips_total = sum(1 for j in range(1, len(verdicts)) if verdicts[j] != verdicts[j - 1])
        window = verdicts[-flaky_window:]
        flip_count = sum(1 for j in range(1, len(window)) if window[j] != window[j - 1])
        reopen_count = sum(
            1 for j in range(1, len(verdicts)) if verdicts[j] and not verdicts[j - 1]
        )  # số lần Pass -> Fail (tái lỗi)
        is_flaky = flip_count >= flaky_min_flips
        # Persistent (dai dang): fail lien tuc >= persist_n cycle GAN NHAT (khong pass xen ke),
        # va cycle moi nhat cung fail -> loi ben vung, uu tien xu ly dut diem. Khac flaky.
        is_persistent = (
            fail_latest > 0 and len(verdicts) >= persist_n and all(verdicts[-persist_n:])
        )

        out.append({
            "test_suite": suite, "test_case": case,
            "fail_count": fail,  # Tổng lần Fail qua tất cả cycle & model
            "fail_model_breadth": breadth,  # Số model khác nhau từng fail (độ rộng)
            "priority_score": priority_score,  # fail_count × breadth (điểm ưu tiên chính)
            "fail_latest_count": fail_latest,  # Số model Fail ở latest status (dùng xếp tier)
            "pass_count": a["pass"], "not_run_count": not_run,
            "priority_tier": tier, "last_updated_cycle": a["last_cycle"],
            "model_detail": a["models"],
            "failing_models": failing_models,
            "current_owner": current_owner,
            "team": owner_team.get(current_owner, ""),  # Team (nhóm nhỏ) của người phụ trách
            "is_flaky": is_flaky,
            "is_persistent": is_persistent, # fail liên tục >= persistent_fail_cycles cycle gần nhất
            "flip_count": flip_count,       # số lần đổi pass<->fail trong flaky_window cycle gần nhất
            "flips_total": flips_total,     # tổng số lần đổi trên toàn lịch sử
            "reopen_count": reopen_count,   # số lần Pass->Fail toàn lịch sử (tái lỗi)
        })
    return out


def get_owner_stats(db, priority=None):
    """Per-owner KPI leaderboard: distinct scripts touched, distinct fully resolved,
    fix-verification rate (per attempt), resolution rate (per distinct script), and
    current open workload (from assignments). Used by both /api/dashboard and Excel export."""
    if priority is None:
        priority = get_script_priority(db)
    priority_map = {(p["test_suite"], p["test_case"]): p for p in priority}

    owner_rows = db.execute("SELECT name FROM owners ORDER BY name").fetchall()
    assignment_rows = db.execute("SELECT test_suite, test_case, owner FROM assignments").fetchall()
    # So script viet moi da hoan thanh (DONE) theo tung nguoi - all-time (cot Scripts Written).
    written_map = {
        r["member"]: r["c"]
        for r in db.execute(
            "SELECT member, COUNT(*) c FROM new_scripts "
            "WHERE status='DONE' AND member IS NOT NULL AND TRIM(member)!='' GROUP BY member"
        ).fetchall()
    }
    open_workload = {}
    for r in assignment_rows:
        if not r["owner"]:
            continue
        p = priority_map.get((r["test_suite"], r["test_case"]))
        # Open workload = script DANG con loi that su (P0-P3); Verify khong tinh
        # (khong can fix, chi cho du chuoi pass de len Done).
        if p and p["priority_tier"] in ("P0", "P1", "P2", "P3"):
            open_workload[r["owner"]] = open_workload.get(r["owner"], 0) + 1

    # Prefetch (chong N+1): gom tat ca fixes theo owner trong 1 truy van, va tinh san
    # so lieu verify (total / fail) cho moi (suite, case, model, cycle) + ban gop-moi-model
    # (suite, case, cycle). Thay cho 1 SELECT fixes/owner + 2 COUNT/fix. Ket qua Y HET.
    fixes_by_owner = {}
    for f in db.execute("SELECT * FROM fixes ORDER BY owner, id").fetchall():
        fixes_by_owner.setdefault(f["owner"], []).append(f)
    agg_model = {}   # (suite, case, model, cycle) -> [total, fail]
    agg_allmod = {}  # (suite, case, cycle)        -> [total, fail]  (gop moi model)
    for r in db.execute(
        "SELECT test_suite, test_case, model, cycle, COUNT(*) total, "
        "SUM(CASE WHEN result='Fail' THEN 1 ELSE 0 END) fail "
        "FROM results WHERE result <> 'Excluded' GROUP BY test_suite, test_case, model, cycle"
    ).fetchall():
        agg_model[(r["test_suite"], r["test_case"], r["model"], r["cycle"])] = (r["total"], r["fail"])
        k2 = (r["test_suite"], r["test_case"], r["cycle"])
        t, fl = agg_allmod.get(k2, (0, 0))
        agg_allmod[k2] = (t + r["total"], fl + (r["fail"] or 0))

    owner_stats = []
    for orow in owner_rows:
        owner = orow["name"]
        fx = fixes_by_owner.get(owner, [])
        if not fx and owner not in open_workload and not written_map.get(owner):
            continue
        seen = set()
        distinct_scripts = []
        fully_resolved = []
        verified = reopened = pending = 0
        for f in fx:
            key = (f["test_suite"], f["test_case"])
            if key not in seen:
                seen.add(key)
                distinct_scripts.append({"test_suite": key[0], "test_case": key[1]})
                p = priority_map.get(key)
                if p and p["fail_count"] == 0:
                    fully_resolved.append({"test_suite": key[0], "test_case": key[1]})

            # per-fix-attempt verification status (uses cycle-scoped before/after counts)
            verify_cycle = f["fixed_after_cycle"] + 1
            model_fixed = f["model_fixed"]
            if model_fixed == "All Models":
                after_total, after_fail = agg_allmod.get(
                    (f["test_suite"], f["test_case"], verify_cycle), (0, 0))
            else:
                after_total, after_fail = agg_model.get(
                    (f["test_suite"], f["test_case"], model_fixed, verify_cycle), (0, 0))
            if after_total == 0:
                pending += 1
            elif after_fail == 0:
                verified += 1
            else:
                reopened += 1

        distinct_fixed_n = len(distinct_scripts)
        fully_resolved_n = len(fully_resolved)
        owner_stats.append({
            "owner": owner,
            "fixes_logged": len(fx),
            "scripts_written": written_map.get(owner, 0),  # so script viet moi DONE (all-time)
            "distinct_scripts_fixed": distinct_fixed_n,
            "distinct_scripts_fully_resolved": fully_resolved_n,
            "verified": verified, "reopened": reopened, "pending": pending,
            # % cac lan fix duoc xac nhan dung ngay lan dau (khong bi reopen), tren so lan da co ket qua doi chieu
            "verification_rate": (verified / (verified + reopened)) if (verified + reopened) else None,
            # % SO SCRIPT KHAC NHAU tung dong gop ma NAY DA HET LOI HOAN TOAN tren tat ca model - KPI chinh
            "resolution_rate": (fully_resolved_n / distinct_fixed_n) if distinct_fixed_n else None,
            "open_workload": open_workload.get(owner, 0),
        })

    owner_stats.sort(key=lambda x: (
        -(x["resolution_rate"] if x["resolution_rate"] is not None else -1),
        -x["distinct_scripts_fully_resolved"],
    ))
    for i, o in enumerate(owner_stats, start=1):
        o["rank"] = i
    return owner_stats


def _activity_days(db, date_from=None, date_to=None):
    """So NGAY co hoat dong (co fix hoac co script DONE) trong khoang [date_from, date_to].
    Dung lam mau so cho nang suat TB/nguoi/ngay (tranh chia cho ngay khong ai lam)."""
    rng = ""
    params = []
    if date_from and date_to:
        rng = " WHERE d >= ? AND d <= ?"
        params = [date_from, date_to]
    q = (
        "SELECT COUNT(DISTINCT d) c FROM ("
        "  SELECT fix_date d FROM fixes WHERE fix_date IS NOT NULL AND TRIM(fix_date)!=''"
        "  UNION"
        "  SELECT completed_date d FROM new_scripts WHERE status='DONE' AND completed_date IS NOT NULL AND TRIM(completed_date)!=''"
        ")" + rng
    )
    return db.execute(q, params).fetchone()["c"] or 0


def _week_range(week_str):
    """Parse 'YYYY-Wnn' -> (monday_iso, sunday_iso). Raise ValueError neu sai."""
    m = re.match(r"^(\d{4})-W(\d{1,2})$", (week_str or "").strip())
    if not m:
        raise ValueError("Định dạng tuần không hợp lệ — dùng YYYY-Wnn (VD 2026-W28).")
    monday = date.fromisocalendar(int(m.group(1)), int(m.group(2)), 1)
    return monday.isoformat(), (monday + timedelta(days=6)).isoformat()


@app.route("/api/leaderboard")
@require_login
def api_leaderboard():
    """Owner leaderboard co the chon theo NGAY / TUAN / CONG DON.
    - cumulative: KPI all-time day du (resolution/verification rate, workload) + scripts_written.
    - day/week: dem fix (fix_date) & script viet moi DONE (completed_date) TRONG khoang;
      rate all-time van kem de tham chieu. Kem totals + nang suat TB/nguoi/ngay."""
    db = get_db()
    scope = (request.args.get("scope") or "cumulative").strip().lower()
    base = get_owner_stats(db)
    base_map = {o["owner"]: o for o in base}

    date_from = date_to = None
    if scope == "day":
        date_from = date_to = (request.args.get("date") or "").strip() or date.today().isoformat()
    elif scope == "week":
        try:
            date_from, date_to = _week_range(request.args.get("week") or date.today().strftime("%G-W%V"))
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

    if scope == "cumulative":
        rows = base
    else:
        fx_map = {
            r["owner"]: r["c"] for r in db.execute(
                "SELECT owner, COUNT(*) c FROM fixes WHERE fix_date>=? AND fix_date<=? "
                "AND owner IS NOT NULL AND TRIM(owner)!='' GROUP BY owner", (date_from, date_to)
            ).fetchall()
        }
        wr_map = {
            r["member"]: r["c"] for r in db.execute(
                "SELECT member, COUNT(*) c FROM new_scripts WHERE status='DONE' AND completed_date>=? AND completed_date<=? "
                "AND member IS NOT NULL AND TRIM(member)!='' GROUP BY member", (date_from, date_to)
            ).fetchall()
        }
        rows = []
        for name in sorted(set(fx_map) | set(wr_map)):
            b = base_map.get(name, {})
            rows.append({
                "owner": name,
                "fixes_logged": fx_map.get(name, 0),
                "scripts_written": wr_map.get(name, 0),
                "distinct_scripts_fixed": b.get("distinct_scripts_fixed", 0),
                "distinct_scripts_fully_resolved": b.get("distinct_scripts_fully_resolved", 0),
                "verified": b.get("verified", 0), "reopened": b.get("reopened", 0), "pending": b.get("pending", 0),
                "verification_rate": b.get("verification_rate"),
                "resolution_rate": b.get("resolution_rate"),
                "open_workload": b.get("open_workload", 0),
            })
        rows.sort(key=lambda x: (-(x["scripts_written"] + x["fixes_logged"]), x["owner"]))
        for i, o in enumerate(rows, start=1):
            o["rank"] = i

    total_written = sum(o.get("scripts_written", 0) for o in rows)
    total_fixes = sum(o.get("fixes_logged", 0) for o in rows)
    people = sum(1 for o in rows if (o.get("scripts_written", 0) or o.get("fixes_logged", 0)))
    days = _activity_days(db, date_from, date_to)
    return jsonify({
        "scope": scope,
        "date_from": date_from, "date_to": date_to,
        "rows": rows,
        "totals": {
            "scripts_written": total_written,
            "fixes_logged": total_fixes,
            "people": people,
            "days": days,
            "avg_write_per_person_day": (total_written / people / days) if (people and days) else None,
            "avg_fix_per_person_day": (total_fixes / people / days) if (people and days) else None,
        },
    })


@app.route("/api/priority")
def api_priority():
    db = get_db()
    data = get_script_priority(db)
    # Sắp xếp CHÍNH theo điểm ưu tiên (fail_count × breadth) giảm dần — case lỗi nhiều
    # nhất & rộng nhất lên đầu. Done (hết lỗi) luôn xuống cuối, Verify (đang xác minh)
    # ngay trên Done. Rồi tie-break theo tổng fail.
    data.sort(key=lambda x: (
        {"Done": 2, "Verify": 1}.get(x["priority_tier"], 0),
        -x["priority_score"], -x["fail_count"],
        x["test_suite"], x["test_case"],
    ))
    for i, p in enumerate(data, start=1):
        p["rank"] = i  # thứ hạng ưu tiên (#1 = cần fix trước nhất)
    return jsonify(data)


@app.route("/api/script-cycle-matrix")
def api_script_cycle_matrix():
    """Pass rate/fail count của từng script tách theo từng cycle, để so sánh script đó
    thay đổi ra sao qua các lần chạy (cải thiện / giảm sút / không đổi).
    Query param ?by_model=1 -> tách riêng theo từng model (test_suite/test_case/model)."""
    db = get_db()
    group_by_model = request.args.get("by_model") in ("1", "true", "True")
    data = compute_script_cycle_matrix(db, group_by_model=group_by_model)
    tier_order = {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "Verify": 4, "Done": 5, "": 6}
    trend_order = {"regressed": 0, "improved": 1, "unchanged": 2, "insufficient_data": 3}
    data["scripts"].sort(key=lambda s: (
        trend_order.get(s["overall_trend"], 9),
        tier_order.get(s["priority_tier"], 9),
        s["test_suite"], s["test_case"], s.get("model") or "",
    ))
    return jsonify(data)


@app.route("/api/suite-model-matrix")
def api_suite_model_matrix():
    """Pass rate của từng test suite (item) trên từng model theo từng cycle — cho Dashboard."""
    db = get_db()
    return jsonify(compute_suite_model_matrix(db))


# ------------------------------------------------------------------
# Assignments: "who is currently responsible for this script"
# Kept fully separate from `fixes` (the historical fix log) so that
# reassigning a script to a new owner NEVER rewrites past history.
# ------------------------------------------------------------------
@app.route("/api/script-fail-details/<path:suite>/<path:case>")
def api_script_fail_details(suite, case):
    """Lấy danh sách chi tiết tất cả các lần Fail của script trên tất cả model/cycle,
    kèm Test ID để owner có thể kiểm tra trên hệ thống test farm.
    """
    db = get_db()
    rows = db.execute(
        """
        SELECT id, cycle, cycle_date, test_id, model, serial, state, description, result
        FROM results
        WHERE test_suite=? AND test_case=? AND result='Fail'
        ORDER BY cycle DESC, id DESC
        """,
        (suite, case),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/root-cause/breakdown")
def api_root_cause_breakdown():
    """Drill-down cho Pareto 'dang fail vi gi': voi 1 nhan root-cause (query param `label`,
    nhan VN goc tu summarize_root_cause), tra ve danh sach item x model DANG FAIL HIEN TAI
    (latest status = Fail, khong phai toan bo lich su Fail) khop nhan do, kem owner/team,
    de gom batch giao nguoi fix. Khong truyen `label` -> tra ve tong hop theo nhan (groups)."""
    db = get_db()
    label = (request.args.get("label") or "").strip()

    latest_fails = db.execute(
        """
        SELECT test_suite, test_case, model, description, test_id, cycle, cycle_date
        FROM (
            SELECT *, ROW_NUMBER() OVER (
                PARTITION BY test_suite, test_case, model
                ORDER BY cycle DESC, id DESC
            ) as rn
            FROM results WHERE result <> 'Excluded'
        )
        WHERE rn = 1 AND result='Fail'
        """
    ).fetchall()

    owner_map = {(r["test_suite"], r["test_case"]): r["owner"]
                 for r in db.execute("SELECT test_suite, test_case, owner FROM assignments")}
    team_map = {r["name"]: r["team"] for r in db.execute("SELECT name, team FROM owners")}

    if not label:
        groups = {}
        for r in latest_fails:
            key = summarize_root_cause(r["description"], "")
            g = groups.setdefault(key, {"key": key, "current_fail_count": 0, "scripts": set()})
            g["current_fail_count"] += 1
            g["scripts"].add((r["test_suite"], r["test_case"]))
        out = sorted(
            [{"key": g["key"], "label_en": _translate_root_cause_label(g["key"]),
              "current_fail_count": g["current_fail_count"], "affected_scripts": len(g["scripts"])}
             for g in groups.values()],
            key=lambda x: -x["current_fail_count"],
        )
        return jsonify({"groups": out})

    items = []
    scripts = set()
    for r in latest_fails:
        if summarize_root_cause(r["description"], "") != label:
            continue
        owner = owner_map.get((r["test_suite"], r["test_case"]), "")
        items.append({
            "test_suite": r["test_suite"], "test_case": r["test_case"], "test_id": r["test_id"],
            "model": r["model"], "cycle": r["cycle"], "cycle_date": r["cycle_date"],
            "owner": owner or "", "team": team_map.get(owner, "") if owner else "",
            "description": r["description"],
        })
        scripts.add((r["test_suite"], r["test_case"]))
    items.sort(key=lambda x: (x["test_suite"], x["test_case"], x["model"]))

    by_model = {}
    by_suite = {}
    for it in items:
        by_model[it["model"]] = by_model.get(it["model"], 0) + 1
        by_suite[it["test_suite"]] = by_suite.get(it["test_suite"], 0) + 1

    return jsonify({
        "label": label,
        "label_en": _translate_root_cause_label(label),
        "current_fail_count": len(items),
        "affected_scripts": len(scripts),
        "items": items,
        "by_model": sorted([{"model": k, "count": v} for k, v in by_model.items()], key=lambda x: -x["count"]),
        "by_suite": sorted([{"test_suite": k, "count": v} for k, v in by_suite.items()], key=lambda x: -x["count"]),
    })


@app.route("/api/assignments", methods=["GET"])
def api_get_assignments():
    db = get_db()
    rows = db.execute("SELECT test_suite, test_case, owner, assigned_date FROM assignments").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/assignments", methods=["POST"])
@require_perm("priority")
def api_set_assignment():
    """Body: {test_suite, test_case, owner}. Sets/changes who is currently responsible.
    Does NOT touch any existing fixes rows — past fix history stays with whoever actually did it."""
    data = request.get_json(force=True)
    test_suite = (data.get("test_suite") or "").strip()
    test_case = (data.get("test_case") or "").strip()
    owner = (data.get("owner") or "").strip()
    if not test_suite or not test_case:
        return jsonify({"error": "Thieu test_suite/test_case"}), 400
    db = get_db()
    if owner:
        db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (owner,))
    db.execute(
        "INSERT INTO assignments (test_suite, test_case, owner, assigned_date) VALUES (?,?,?,?) "
        "ON CONFLICT(test_suite, test_case) DO UPDATE SET owner=excluded.owner, assigned_date=excluded.assigned_date",
        (test_suite, test_case, owner, date.today().isoformat()),
    )
    db.commit()
    return jsonify({"status": "ok"})


@app.route("/api/handover", methods=["POST"])
@require_perm("settings")
def api_handover():
    """Body: {from_owner, to_owner, only_open (bool, default True)}.
    Reassigns CURRENT script ownership from from_owner to to_owner in bulk —
    used when a team member leaves and someone else takes over their remaining work.
    This never modifies the `fixes` history table, so from_owner's past contributions
    stay correctly attributed to them."""
    data = request.get_json(force=True)
    from_owner = (data.get("from_owner") or "").strip()
    to_owner = (data.get("to_owner") or "").strip()
    only_open = data.get("only_open", True)
    if not from_owner or not to_owner:
        return jsonify({"error": "Thieu from_owner/to_owner"}), 400
    db = get_db()
    db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (to_owner,))

    assigned = db.execute(
        "SELECT test_suite, test_case FROM assignments WHERE owner=?", (from_owner,)
    ).fetchall()

    if only_open:
        priority_map = {(p["test_suite"], p["test_case"]): p for p in get_script_priority(db)}
        targets = [
            (r["test_suite"], r["test_case"]) for r in assigned
            if priority_map.get((r["test_suite"], r["test_case"]), {}).get("priority_tier") != "Done"
        ]
    else:
        targets = [(r["test_suite"], r["test_case"]) for r in assigned]

    for suite, case in targets:
        db.execute(
            "UPDATE assignments SET owner=?, assigned_date=? WHERE test_suite=? AND test_case=?",
            (to_owner, date.today().isoformat(), suite, case),
        )
    log_audit(db, "handover", target=from_owner, detail=f"-> {to_owner}, scripts={len(targets)}")
    db.commit()
    return jsonify({"status": "ok", "reassigned_count": len(targets), "scripts": [{"test_suite": s, "test_case": c} for s, c in targets]})


@app.route("/api/dashboard")
def api_dashboard():
    db = get_db()

    # ---- Trend by cycle (Pass Rate theo cong thuc rieng: NA khong tinh vao tu so & mau so) ----
    trend = compute_cycle_trend(db)

    # ---- Pass rate by model (latest cycle) ----
    models_list = get_models_list(db)
    latest_cycle_row = db.execute("SELECT MAX(cycle) as c FROM results").fetchone()
    latest_cycle = latest_cycle_row["c"] or 0
    model_rows = db.execute(
        """
        SELECT model,
               COUNT(*) as total,
               SUM(CASE WHEN result='Pass' THEN 1 ELSE 0 END) as pass_count
        FROM results WHERE cycle=? AND result <> 'Excluded' GROUP BY model
        """,
        (latest_cycle,),
    ).fetchall()
    model_counts = {m: {"total": 0, "pass": 0} for m in models_list}
    for r in model_rows:
        if r["model"] in model_counts:
            model_counts[r["model"]] = {"total": r["total"], "pass": r["pass_count"]}
        else:
            # Model appears in data but no longer in the managed list (e.g. retired) — still show it.
            model_counts[r["model"]] = {"total": r["total"], "pass": r["pass_count"]}
    model_pass_rate = {}
    for m, c in model_counts.items():
        model_pass_rate[m] = (c["pass"] / c["total"]) if c["total"] else None

    # ---- Priority tier distribution ----
    priority = get_script_priority(db)
    tier_counts = {"P0": 0, "P1": 0, "P2": 0, "P3": 0, "Verify": 0, "Done": 0}
    for p in priority:
        tier_counts[p["priority_tier"]] += 1

    # ---- Root cause pareto (đã GOM NHOM theo nguyên nhân, không group text thô) ----
    # english=True: Dashboard hiển thị tiếng Anh (khác với export Excel toàn bộ dữ liệu ở tab Ưu tiên, vẫn tiếng Việt)
    root_causes = compute_root_cause_pareto(db, limit=15, english=True)
    # ---- Pareto theo nhóm nguyên nhân ĐÃ XÁC NHẬN từ fix log (A3) ----
    fix_root_causes = compute_fix_root_cause_pareto(db)

    # ---- Owner stats (KPI leaderboard) ----
    owner_stats = get_owner_stats(db, priority)

    # ---- Test suite (item) stats ----
    # "still_failing" = script DANG co loi (P0-P3). Verify (het loi nhung chua du chuoi
    # pass) khong tinh vao "con loi" nhung cung chua tinh Done -> theo doi rieng.
    suite_rows = {}
    for p in priority:
        s = suite_rows.setdefault(p["test_suite"], {"total": 0, "done": 0, "verify": 0, "fail_scripts": 0})
        s["total"] += 1
        if p["priority_tier"] == "Done":
            s["done"] += 1
        elif p["priority_tier"] == "Verify":
            s["verify"] += 1
        else:
            s["fail_scripts"] += 1
    suite_stats = [
        {"test_suite": k, "total_scripts": v["total"], "done": v["done"],
         "verify": v["verify"], "still_failing": v["fail_scripts"],
         "done_pct": (v["done"] / v["total"]) if v["total"] else 0}
        for k, v in suite_rows.items()
    ]
    suite_stats.sort(key=lambda x: x["done_pct"])

    # ---- Settings-based KPIs ----
    settings_rows = db.execute("SELECT key, value FROM settings").fetchall()
    settings = {r["key"]: r["value"] for r in settings_rows}
    target_rate = float(settings.get("target_pass_rate") or 0.88)
    deadline = settings.get("deadline_date") or ""
    total_scripts = len(priority)
    # "Con loi" = dang co fail that su (P0-P3). Verify khong tinh (khong co gi de fix,
    # chi cho du chuoi pass) -> khong lam phong fails_to_fix / required_rate_per_day.
    still_failing = sum(1 for p in priority if p["priority_tier"] in ("P0", "P1", "P2", "P3"))
    verify_count = tier_counts["Verify"]
    target_fail_allowed = int(total_scripts * (1 - target_rate))
    fails_to_fix = max(0, still_failing - target_fail_allowed)
    days_remaining = None
    required_rate = None
    if deadline:
        try:
            d = datetime.strptime(deadline, "%Y-%m-%d").date()
            days_remaining = (d - date.today()).days
            if days_remaining and days_remaining > 0:
                required_rate = fails_to_fix / days_remaining
        except Exception:
            pass

    current_pass_rate = trend[-1]["pass_rate"] if trend else None

    # ---- Flaky KPI (A2) ----
    flaky_count = sum(1 for p in priority if p.get("is_flaky"))
    flaky_rate = (flaky_count / total_scripts) if total_scripts else None

    # ---- Pass rate điều chỉnh: loại script mới trong N cycle đầu (B2, tuỳ chọn) ----
    exclude_n = get_setting_int(db, "exclude_new_scripts_cycles", 0)
    current_pass_rate_adjusted = None
    if exclude_n > 0 and trend:
        adjusted = compute_adjusted_trend(db, exclude_n)
        for t in trend:
            t["pass_rate_adjusted"] = adjusted.get(t["cycle"])
        current_pass_rate_adjusted = trend[-1].get("pass_rate_adjusted")

    # ---- Coverage: tiến độ viết script so với tổng động từ hệ thống công ty (B2) ----
    coverage = compute_coverage(db)

    # ---- Automated insights (rule-based), English ----
    insights = []
    if trend and len(trend) >= 2:
        last = trend[-1]
        if last["delta_fail"] is not None:
            if last["delta_fail"] < 0:
                insights.append(f"Cycle {last['cycle']}: failures down {-last['delta_fail']} vs previous cycle — improving.")
            elif last["delta_fail"] > 0:
                insights.append(f"Cycle {last['cycle']}: failures UP {last['delta_fail']} vs previous cycle — check for regressions or new failing scripts.")
    worst_model = None
    worst_rate = 2
    for m, r in model_pass_rate.items():
        if r is not None and r < worst_rate:
            worst_rate = r
            worst_model = m
    if worst_model:
        insights.append(f"Weakest model currently: {worst_model} (pass rate {worst_rate*100:.1f}%) — prioritize checking infra/device or model-specific issues.")
    if root_causes:
        top = root_causes[0]
        insights.append(f"Most common failure cause: '{top['description']}' accounts for {top['pct']*100:.1f}% of all failures — fix the root cause and batch-verify related scripts.")
    if tier_counts["P0"] > 0:
        insights.append(f"{tier_counts['P0']} scripts at P0 (failing on 4-5 models) — prioritize these first, widest impact.")
    if flaky_count:
        insights.append(f"{flaky_count} FLAKY scripts (pass/fail alternating) — handle separately (add wait/sync, check locators) instead of repeatedly re-fixing.")
    if tier_counts["Verify"] > 0:
        insights.append(f"{tier_counts['Verify']} scripts VERIFYING (no longer failing but not yet {get_setting_int(db, 'exit_criteria_cycles', 2)} consecutive passing cycles) — keep monitoring before counting as Done.")
    if required_rate is not None:
        # crude actual rate estimate from last 2 cycles
        actual_rate = None
        if len(trend) >= 2 and trend[-2]["total"]:
            prev_still_fail_est = trend[-2]["fail_count"]
            cur_still_fail_est = trend[-1]["fail_count"]
            actual_rate = max(0, prev_still_fail_est - cur_still_fail_est)
        if actual_rate is not None:
            if actual_rate < required_rate:
                insights.append(f"Estimated fix velocity ({actual_rate:.1f} scripts/day) is BELOW the required velocity ({required_rate:.1f} scripts/day) to hit the deadline — need more resources or batch-fix by root cause.")
            else:
                insights.append(f"Estimated fix velocity ({actual_rate:.1f} scripts/day) is at or above the required velocity ({required_rate:.1f} scripts/day) — keep it up.")

    return jsonify({
        "trend": trend,
        "model_pass_rate": model_pass_rate,
        "tier_counts": tier_counts,
        "root_causes": root_causes,
        "fix_root_causes": fix_root_causes,
        "owner_stats": owner_stats,
        "suite_stats": suite_stats,
        "coverage": coverage,
        "kpi": {
            "total_scripts": total_scripts,
            "still_failing": still_failing,
            "current_pass_rate": current_pass_rate,
            "current_pass_rate_adjusted": current_pass_rate_adjusted,
            "target_pass_rate": target_rate,
            "fails_to_fix": fails_to_fix,
            "days_remaining": days_remaining,
            "required_rate_per_day": required_rate,
            "latest_cycle": latest_cycle,
            "flaky_count": flaky_count,
            "flaky_rate": flaky_rate,
            "verify_count": verify_count,
        },
        "insights": insights,
    })


# ------------------------------------------------------------------
# Export
# ------------------------------------------------------------------
@app.route("/api/export/csv/<table>")
def export_csv(table):
    if table not in ("results", "fixes"):
        return jsonify({"error": "invalid table"}), 400
    db = get_db()
    rows = db.execute(f"SELECT * FROM {table} ORDER BY id").fetchall()
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=rows[0].keys())
        writer.writeheader()
        for r in rows:
            writer.writerow(dict(r))
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    return send_file(mem, mimetype="text/csv", as_attachment=True,
                      download_name=f"{table}_export.csv")


def _heat_hex(rate):
    """Mau heatmap theo pass rate 0..1: do (0) -> vang -> xanh la (1), giong heatColor() ben JS
    (HSL hue 0..120, S=62%, L=86%) nhung tra ve hex RGB de dung lam PatternFill trong Excel."""
    h = (max(0.0, min(1.0, rate)) * 120) / 360
    s, l = 0.62, 0.86

    def hue2rgb(p, q, t):
        if t < 0:
            t += 1
        if t > 1:
            t -= 1
        if t < 1 / 6:
            return p + (q - p) * 6 * t
        if t < 1 / 2:
            return q
        if t < 2 / 3:
            return p + (q - p) * (2 / 3 - t) * 6
        return p

    q = l * (1 + s) if l < 0.5 else l + s - l * s
    p = 2 * l - q
    r = hue2rgb(p, q, h + 1 / 3)
    g = hue2rgb(p, q, h)
    b = hue2rgb(p, q, h - 1 / 3)
    return f"{round(r * 255):02X}{round(g * 255):02X}{round(b * 255):02X}"


@app.route("/api/export/excel/suite-model-matrix")
def export_excel_suite_model_matrix():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    matrix = compute_suite_model_matrix(db)
    all_cycles = matrix["cycles"]

    cycles_param = request.args.get("cycles", "").strip()
    if cycles_param:
        wanted = {int(c) for c in cycles_param.split(",") if c.strip().isdigit()}
        sel_cycles = [c for c in all_cycles if c["cycle"] in wanted]
    else:
        sel_cycles = all_cycles
    if not sel_cycles:
        sel_cycles = all_cycles

    NAVY = "1F4E78"
    HEADER_FILL = PatternFill("solid", fgColor=NAVY)
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    TITLE_FONT = Font(bold=True, name="Calibri", size=14, color=NAVY)
    HINT_FONT = Font(name="Calibri", size=9, italic=True, color="6B7280")
    OVERALL_FILL = PatternFill("solid", fgColor="EEF3FB")
    MODEL_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
    MODEL_FILL = PatternFill("solid", fgColor="3498DB")
    NONE_FILL = PatternFill("solid", fgColor="F5F5F5")
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pass_Rate_Matrix"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "PASS RATE THEO ITEM x MODEL x CYCLE"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Xuat luc: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Cycle: " + \
        ", ".join(f"C{c['cycle']}" for c in sel_cycles)
    ws["A2"].font = HINT_FONT
    ws["A3"] = "Cong thuc Pass Rate = (Pass + Check + Manual Check) / (Tong so - Skip - NA)"
    ws["A3"].font = HINT_FONT

    header_row = 5
    ws.cell(row=header_row, column=1, value="Item (Test suite)")
    ws.cell(row=header_row, column=2, value="Model")
    for i, c in enumerate(sel_cycles):
        col = 3 + i
        label = f"Cycle {c['cycle']}" + (f"\n{c['cycle_date']}" if c.get("cycle_date") else "")
        ws.cell(row=header_row, column=col, value=label)
    style_row_range = range(1, 3 + len(sel_cycles))
    for col in style_row_range:
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
        cell.border = BORDER

    def write_cell(row, col, cell_data):
        c = ws.cell(row=row, column=col)
        if not cell_data or cell_data.get("pass_rate") is None:
            c.value = "—"
            c.fill = NONE_FILL
        else:
            rate = cell_data["pass_rate"]
            c.value = f"{rate * 100:.0f}%\n{cell_data['fail_count']}F / {cell_data['total']}T" + \
                (f" / {cell_data['na_count']}NA" if cell_data.get("na_count") else "")
            c.fill = PatternFill("solid", fgColor=_heat_hex(rate))
        c.alignment = WRAP_CENTER
        c.border = BORDER
        c.font = Font(name="Calibri", size=9, bold=True)

    row = header_row + 1
    ws.cell(row=row, column=1, value="OVERALL - tat ca script")
    ws.cell(row=row, column=1).font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=row, column=1).fill = OVERALL_FILL
    ws.cell(row=row, column=1).border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=2).fill = OVERALL_FILL
    ws.cell(row=row, column=2).border = BORDER
    overall_by_cycle = matrix["overall_by_cycle"]
    for i, c in enumerate(sel_cycles):
        write_cell(row, 3 + i, overall_by_cycle.get(c["cycle"]))
    row += 1

    # Gom nhom theo Item giong tren web: merge cot Item cho tat ca model cua item do.
    groups = {}
    order = []
    for r in matrix["rows"]:
        if r["test_suite"] not in groups:
            groups[r["test_suite"]] = []
            order.append(r["test_suite"])
        groups[r["test_suite"]].append(r)

    for suite in order:
        items = groups[suite]
        start_row = row
        for r in items:
            ws.cell(row=row, column=2, value=r["model"])
            ws.cell(row=row, column=2).font = MODEL_FONT
            ws.cell(row=row, column=2).fill = MODEL_FILL
            ws.cell(row=row, column=2).alignment = WRAP_CENTER
            ws.cell(row=row, column=2).border = BORDER
            for i, c in enumerate(sel_cycles):
                write_cell(row, 3 + i, r["by_cycle"].get(c["cycle"]))
            row += 1
        end_row = row - 1
        ws.cell(row=start_row, column=1, value=suite)
        ws.cell(row=start_row, column=1).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=start_row, column=1).alignment = WRAP_CENTER
        for r_ in range(start_row, end_row + 1):
            ws.cell(row=r_, column=1).border = BORDER
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    for i in range(len(sel_cycles)):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(3 + i)].width = 16
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"PassRate_ItemModelCycle_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )


def build_new_scripts_workbook(db):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    models = get_models_list(db)
    rows = db.execute(
        "SELECT * FROM new_scripts ORDER BY completed_date DESC, id DESC"
    ).fetchall()

    TEAM_FILL = PatternFill("solid", fgColor="4472C4")
    TEAM_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="FCD5B5")
    HEADER_FONT = Font(bold=True, color="000000", name="Calibri", size=10)
    THIN = Side(style="thin", color="000000")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
    CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center")
    LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "(new) UI90 script"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3.3
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[3].height = 31.5

    def style_header_cell(cell, fill, font):
        cell.fill = fill
        cell.font = font
        cell.border = BORDER

    fixed_cols = [
        ("B", "Team", 15),
        ("C", "Item", 19.3),
        ("D", "TC ID", 21.4),
        ("E", "Member", 14.6),
        ("F", "Assign Week", 19.3),
        ("G", "Completed date", 16),
    ]
    for col, label, width in fixed_cols:
        ws.column_dimensions[col].width = width
        ws.merge_cells(f"{col}2:{col}3")
        top_cell = ws[f"{col}2"]
        top_cell.value = label
        top_cell.alignment = CENTER
        if col == "B":
            style_header_cell(top_cell, TEAM_FILL, TEAM_FONT)
        else:
            style_header_cell(top_cell, HEADER_FILL, HEADER_FONT)
        style_header_cell(ws[f"{col}3"], HEADER_FILL, HEADER_FONT)

    # Nhom "Local": 1 cot Status + 1 cot cho moi model (dong theo bang models)
    local_start_col = 8  # H
    local_span = 1 + len(models)
    local_end_col = local_start_col + local_span - 1
    ws.merge_cells(start_row=2, start_column=local_start_col, end_row=2, end_column=local_end_col)
    for c in range(local_start_col, local_end_col + 1):
        style_header_cell(ws.cell(row=2, column=c), HEADER_FILL, HEADER_FONT)
    ws.cell(row=2, column=local_start_col, value="Local").alignment = CENTER

    status_col = local_start_col
    status_cell = ws.cell(row=3, column=status_col, value="Status\n(DONE/SKIP)")
    style_header_cell(status_cell, HEADER_FILL, HEADER_FONT)
    status_cell.alignment = CENTER_WRAP
    ws.column_dimensions[get_column_letter(status_col)].width = 14

    model_col_map = {}
    for i, model in enumerate(models):
        col = local_start_col + 1 + i
        model_col_map[model] = col
        mcell = ws.cell(row=3, column=col, value=model)
        style_header_cell(mcell, HEADER_FILL, HEADER_FONT)
        mcell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Nhom "SDF": chi 1 cot Status (khong track theo tung model), dung de hien thi sdf_id
    sdf_col = local_end_col + 1
    ws.merge_cells(start_row=2, start_column=sdf_col, end_row=2, end_column=sdf_col)
    sdf_header = ws.cell(row=2, column=sdf_col, value="SDF")
    style_header_cell(sdf_header, HEADER_FILL, HEADER_FONT)
    sdf_header.alignment = CENTER
    sdf_status_cell = ws.cell(row=3, column=sdf_col, value="Status")
    style_header_cell(sdf_status_cell, HEADER_FILL, HEADER_FONT)
    sdf_status_cell.alignment = CENTER
    ws.column_dimensions[get_column_letter(sdf_col)].width = 12

    # Remark: cot cuoi, merge doc nhu cac cot co dinh
    remark_col = sdf_col + 1
    remark_letter = get_column_letter(remark_col)
    ws.column_dimensions[remark_letter].width = 91
    ws.merge_cells(start_row=2, start_column=remark_col, end_row=3, end_column=remark_col)
    remark_header = ws.cell(row=2, column=remark_col, value="Remark")
    style_header_cell(remark_header, HEADER_FILL, HEADER_FONT)
    remark_header.alignment = CENTER

    r = 4
    for row in rows:
        ws.cell(row=r, column=2, value=row["team"] or "").alignment = LEFT
        ws.cell(row=r, column=3, value=row["item"] or "").alignment = LEFT
        ws.cell(row=r, column=4, value=row["tc_id"] or "").alignment = LEFT
        ws.cell(row=r, column=5, value=row["member"] or "").alignment = CENTER
        ws.cell(row=r, column=6, value=row["assign_week"]).alignment = CENTER

        dcell = ws.cell(row=r, column=7)
        raw_date = (row["completed_date"] or "")[:10]
        try:
            dcell.value = date.fromisoformat(raw_date)
            dcell.number_format = "d-mmm"
        except ValueError:
            dcell.value = raw_date
        dcell.alignment = CENTER

        ws.cell(row=r, column=status_col, value=row["status"] or "").alignment = CENTER

        written = {m.strip() for m in (row["models_written"] or "").split(",") if m.strip()}
        for model, col in model_col_map.items():
            ws.cell(row=r, column=col, value="O" if model in written else "").alignment = CENTER

        ws.cell(row=r, column=sdf_col, value=row["sdf_id"] or "").alignment = CENTER

        remark_cell = ws.cell(row=r, column=remark_col, value=row["remark"] or "")
        remark_cell.alignment = LEFT_WRAP

        for c in range(2, remark_col + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    ws.freeze_panes = "B4"
    return wb


@app.route("/api/export/excel/new-scripts")
def export_excel_new_scripts():
    db = get_db()
    wb = build_new_scripts_workbook(db)
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"NewScripts_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )


@app.route("/api/export/excel")
def export_excel():
    db = get_db()
    wb = build_export_workbook(db)
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Test_Stabilization_Tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )


# ------------------------------------------------------------------
# Comprehensive Excel workbook builder
# Mirrors (and improves on) the original Test_Stabilization_Tracker.xlsx layout,
# but every cell is a computed VALUE from the live database — no fragile formulas,
# so it stays readable/correct even opened stand-alone, offline, as a backup artifact.
# ------------------------------------------------------------------
def build_export_workbook(db):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.worksheet.table import Table, TableStyleInfo

    NAVY = "1F4E78"
    HEADER_FILL = PatternFill("solid", fgColor=NAVY)
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    TITLE_FONT = Font(bold=True, name="Calibri", size=16, color=NAVY)
    SUB_FONT = Font(bold=True, name="Calibri", size=11, color=NAVY)
    NORMAL = Font(name="Calibri", size=10)
    ITALIC_HINT = Font(name="Calibri", size=9, italic=True, color="6B7280")
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TIER_FILL = {
        "P0": PatternFill("solid", fgColor="FF6B6B"),
        "P1": PatternFill("solid", fgColor="FFB86B"),
        "P2": PatternFill("solid", fgColor="FFE56B"),
        "P3": PatternFill("solid", fgColor="BFE3FF"),
        "Verify": PatternFill("solid", fgColor="D6C9F0"),
        "Done": PatternFill("solid", fgColor="B6E7A0"),
    }
    RESULT_FILL = {
        "Pass": PatternFill("solid", fgColor="D9F2D9"),
        "Fail": PatternFill("solid", fgColor="FADBD8"),
    }

    def style_header_row(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

    def autosize(ws, min_w=9, max_w=42, start_row=1):
        for col in ws.columns:
            letter = None
            max_len = 0
            for cell in col:
                if cell.row < start_row:
                    continue
                if letter is None:
                    letter = cell.column_letter
                v = cell.value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            if letter:
                ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)

    # ---- Pull all data up front ----
    models_list = get_models_list(db)
    priority = get_script_priority(db)
    priority.sort(key=lambda x: (
        {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "Done": 4}.get(x["priority_tier"], 5),
        x["test_suite"], x["test_case"],
    ))
    owner_stats = get_owner_stats(db, priority)

    trend = compute_cycle_trend(db)

    latest_cycle = trend[-1]["cycle"] if trend else 0
    # 1 GROUP BY thay cho 2 COUNT/model (chong N+1). Ket qua Y HET: model khong co du lieu
    # o latest_cycle -> khong co trong map -> (0,0) -> None, dung nhu vong lap cu.
    _mc = {r["model"]: (r["total"], r["pass_count"]) for r in db.execute(
        "SELECT model, COUNT(*) total, SUM(CASE WHEN result='Pass' THEN 1 ELSE 0 END) pass_count "
        "FROM results WHERE cycle=? AND result <> 'Excluded' GROUP BY model", (latest_cycle,)).fetchall()}
    model_pass_rate = {}
    for m in models_list:
        tot, pas = _mc.get(m, (0, 0))
        model_pass_rate[m] = (pas / tot) if tot else None

    tier_counts = {"P0": 0, "P1": 0, "P2": 0, "P3": 0, "Verify": 0, "Done": 0}
    for p in priority:
        tier_counts[p["priority_tier"]] += 1

    root_causes = compute_root_cause_pareto(db, limit=15)

    suite_agg = {}
    for p in priority:
        s = suite_agg.setdefault(p["test_suite"], {"total": 0, "done": 0})
        s["total"] += 1
        if p["priority_tier"] == "Done":
            s["done"] += 1
    suite_stats = sorted(
        [{"test_suite": k, "total": v["total"], "done": v["done"],
          "still_failing": v["total"] - v["done"],
          "done_pct": (v["done"] / v["total"]) if v["total"] else 0} for k, v in suite_agg.items()],
        key=lambda x: x["done_pct"],
    )

    settings_rows = db.execute("SELECT key, value FROM settings").fetchall()
    settings = {r["key"]: r["value"] for r in settings_rows}
    target_rate = float(settings.get("target_pass_rate") or 0.88)
    deadline = settings.get("deadline_date") or ""
    total_scripts = len(priority)
    still_failing = sum(1 for p in priority if p["priority_tier"] in ("P0", "P1", "P2", "P3"))
    current_pass_rate = trend[-1]["pass_rate"] if trend else None

    wb = Workbook()

    # ================================================================
    # SHEET: Instructions
    # ================================================================
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 105
    ws["B2"] = "TEST STABILIZATION TRACKER — Bao cao xuat tu he thong web"
    ws["B2"].font = TITLE_FONT
    lines = [
        ("", False),
        (f"Xuat luc: {datetime.now().strftime('%Y-%m-%d %H:%M')}", False),
        (f"Cycle gan nhat: {latest_cycle}  |  Tong so script: {total_scripts}  |  Con loi: {still_failing}", False),
        ("", False),
        ("VE FILE NAY", True),
        ("Day la ban SNAPSHOT (chup nhanh so lieu tai thoi diem xuat) tu he thong web Test Stabilization Tracker — dung de bao cao, luu tru, hoac lam phuong an du phong khi he thong web tam thoi khong dung duoc.", False),
        ("Khac voi ban Excel dung cong thuc truoc day, moi gia tri trong file nay la SO THUC TE da duoc tinh san (khong phai cong thuc) — mo bang Excel/LibreOffice nao cung doc dung, khong lo loi cong thuc.", False),
        ("Neu he thong web tam ngung hoat dong, co the tiep tuc ghi nhan ket qua/fix thu cong vao 2 sheet 'History_Log' va 'Daily_Fix_Log' o cuoi file (dung dinh dang cot y het), roi nhap lai vao he thong khi hoat dong tro lai.", False),
        ("", False),
        ("CAC SHEET TRONG FILE", True),
        ("- Dashboard: KPI tong quan, cac bieu do xu huong, phan bo priority, root cause pareto.", False),
        ("- Script_Priority_Tracker: danh sach script sap theo do uu tien P0->P3->Done, co chi tiet Pass/Fail tren TUNG MODEL va nguoi dang phu trach.", False),
        ("- Owner_Leaderboard: bang xep hang KPI ca nhan — so script da fix, so script da het loi that su, ty le xac minh, ty le hoan thanh.", False),
        ("- RootCause_Pareto: nhom nguyen nhan loi pho bien nhat (80/20).", False),
        ("- Test_Suite_Summary: tien do hoan thanh theo tung test suite.", False),
        ("- Daily_Fix_Log / History_Log: du lieu tho day du, dung de doi chieu hoac nhap lai neu can.", False),
    ]
    r = 4
    for text, is_head in lines:
        cell = ws.cell(row=r, column=2, value=text)
        cell.font = SUB_FONT if is_head else NORMAL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # ================================================================
    # SHEET: Dashboard
    # ================================================================
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws["B2"] = "DASHBOARD"
    ws["B2"].font = TITLE_FONT

    kpis = [
        ("Pass Rate hien tai", current_pass_rate, "0.0%"),
        ("Target Pass Rate", target_rate, "0.0%"),
        ("Tong so script", total_scripts, "0"),
        ("Script con loi", still_failing, "0"),
        ("Cycle hien tai", latest_cycle, "0"),
        ("Deadline", deadline or "(chua dat)", None),
    ]
    for i, (label, val, fmt) in enumerate(kpis):
        row = 4 + i
        ws.cell(row=row, column=2, value=label).font = SUB_FONT
        c = ws.cell(row=row, column=4, value=val)
        c.font = Font(bold=True, size=12, name="Calibri", color=NAVY)
        if fmt:
            c.number_format = fmt

    trend_hr = 12
    ws.cell(row=trend_hr, column=2, value="TREND THEO CYCLE").font = SUB_FONT
    ws.cell(row=trend_hr, column=6, value="Pass Rate = (Pass+Check+Manual Check) / (Tong - NA)").font = ITALIC_HINT
    thead = ["Cycle", "Ngay", "Tong luot chay", "Pass-like", "NA", "Fail", "Pass Rate", "Delta Fail vs truoc"]
    for i, h in enumerate(thead):
        ws.cell(row=trend_hr + 1, column=2 + i, value=h)
    style_header_row(ws, trend_hr + 1, 0)
    for c in range(2, 2 + len(thead)):
        cell = ws.cell(row=trend_hr + 1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    trend_first = trend_hr + 2
    for i, t in enumerate(trend):
        row = trend_first + i
        ws.cell(row=row, column=2, value=t["cycle"])
        ws.cell(row=row, column=3, value=t["cycle_date"])
        ws.cell(row=row, column=4, value=t["total"])
        ws.cell(row=row, column=5, value=t["pass_count"])
        ws.cell(row=row, column=6, value=t["na_count"])
        ws.cell(row=row, column=7, value=t["fail_count"])
        cc = ws.cell(row=row, column=8, value=t["pass_rate"])
        cc.number_format = "0.0%"
        if t["delta_fail"] is not None:
            dc = ws.cell(row=row, column=9, value=t["delta_fail"])
            dc.font = Font(color="1E8449" if t["delta_fail"] <= 0 else "C0392B", name="Calibri", size=10)
    trend_last = trend_first + len(trend) - 1 if trend else trend_first

    # Model pass rate mini-table (for chart)
    model_hr = trend_last + 3
    ws.cell(row=model_hr, column=2, value="PASS RATE THEO MODEL (cycle gan nhat)").font = SUB_FONT
    ws.cell(row=model_hr + 1, column=2, value="Model")
    ws.cell(row=model_hr + 1, column=3, value="Pass Rate")
    style_header_row(ws, model_hr + 1, 0)
    for c in (2, 3):
        cell = ws.cell(row=model_hr + 1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, m in enumerate(models_list):
        row = model_hr + 2 + i
        ws.cell(row=row, column=2, value=m)
        v = model_pass_rate.get(m)
        cc = ws.cell(row=row, column=3, value=v if v is not None else 0)
        cc.number_format = "0.0%"
    model_last = model_hr + 1 + len(models_list)

    # Tier distribution mini-table (for chart)
    tier_hr = model_last + 3
    ws.cell(row=tier_hr, column=2, value="PHAN BO PRIORITY TIER").font = SUB_FONT
    ws.cell(row=tier_hr + 1, column=2, value="Tier")
    ws.cell(row=tier_hr + 1, column=3, value="So script")
    for c in (2, 3):
        cell = ws.cell(row=tier_hr + 1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    tier_order = ["P0", "P1", "P2", "P3", "Verify", "Done"]
    for i, t in enumerate(tier_order):
        row = tier_hr + 2 + i
        ws.cell(row=row, column=2, value=t)
        ws.cell(row=row, column=3, value=tier_counts[t])
    tier_last = tier_hr + 1 + len(tier_order)

    # ---- Charts ----
    if trend:
        line = LineChart()
        line.title = "Pass Rate theo Cycle"
        line.y_axis.title = "Pass Rate"
        line.y_axis.numFmt = "0%"
        data = Reference(ws, min_col=7, min_row=trend_hr + 1, max_row=trend_last)
        cats = Reference(ws, min_col=2, min_row=trend_first, max_row=trend_last)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height, line.width = 8, 16
        ws.add_chart(line, "F4")

        fail_chart = BarChart()
        fail_chart.title = "Fail Count theo Cycle"
        data2 = Reference(ws, min_col=6, min_row=trend_hr + 1, max_row=trend_last)
        fail_chart.add_data(data2, titles_from_data=True)
        fail_chart.set_categories(cats)
        fail_chart.height, fail_chart.width = 8, 16
        ws.add_chart(fail_chart, "N4")

    model_bar = BarChart()
    model_bar.title = "Pass Rate theo Model"
    model_bar.y_axis.numFmt = "0%"
    mdata = Reference(ws, min_col=3, min_row=model_hr + 1, max_row=model_last)
    mcats = Reference(ws, min_col=2, min_row=model_hr + 2, max_row=model_last)
    model_bar.add_data(mdata, titles_from_data=True)
    model_bar.set_categories(mcats)
    model_bar.height, model_bar.width = 8, 16
    ws.add_chart(model_bar, "F19")

    tier_pie = PieChart()
    tier_pie.title = "Phan bo Priority Tier"
    tdata = Reference(ws, min_col=3, min_row=tier_hr + 1, max_row=tier_last)
    tcats = Reference(ws, min_col=2, min_row=tier_hr + 2, max_row=tier_last)
    tier_pie.add_data(tdata, titles_from_data=True)
    tier_pie.set_categories(tcats)
    tier_pie.height, tier_pie.width = 8, 16
    ws.add_chart(tier_pie, "N19")

    for col, w in zip("BCDEFGH", [26, 14, 15, 10, 10, 12, 16]):
        ws.column_dimensions[col].width = w

    # ================================================================
    # SHEET: Script_Priority_Tracker
    # ================================================================
    ws = wb.create_sheet("Script_Priority_Tracker")
    base_headers = ["Test suite", "Test Case", "Priority_Tier", "Current_Owner", "Team",
                    "Fail_Count", "Pass_Count", "NotRun_Count", "Last_Updated_Cycle"]
    headers = base_headers + models_list
    n_base = len(base_headers)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for i, p in enumerate(priority, start=2):
        ws.cell(row=i, column=1, value=p["test_suite"])
        ws.cell(row=i, column=2, value=p["test_case"])
        tier_cell = ws.cell(row=i, column=3, value=p["priority_tier"])
        tier_cell.fill = TIER_FILL.get(p["priority_tier"], PatternFill())
        tier_cell.font = Font(bold=True, name="Calibri", size=10)
        tier_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=4, value=p["current_owner"])
        ws.cell(row=i, column=5, value=p.get("team", ""))
        ws.cell(row=i, column=6, value=p["fail_count"])
        ws.cell(row=i, column=7, value=p["pass_count"])
        ws.cell(row=i, column=8, value=p["not_run_count"])
        ws.cell(row=i, column=9, value=p["last_updated_cycle"])
        for j, m in enumerate(models_list):
            res = p["model_detail"].get(m)
            label = res if res else "Not Run"
            cc = ws.cell(row=i, column=n_base + 1 + j, value=label)
            cc.alignment = Alignment(horizontal="center")
            cc.fill = RESULT_FILL.get(res, PatternFill("solid", fgColor="F0F0F0"))
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = BORDER
            if ws.cell(row=i, column=c).font is None:
                ws.cell(row=i, column=c).font = NORMAL
    last_row = len(priority) + 1
    if last_row > 1:
        tab = Table(displayName="tblPriorityExport", ref=f"A1:{get_column_letter(len(headers))}{last_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(tab)
    ws.freeze_panes = "C2"
    autosize(ws, min_w=10, max_w=22)

    # ================================================================
    # SHEET: Owner_Leaderboard
    # ================================================================
    ws = wb.create_sheet("Owner_Leaderboard")
    ws["A1"] = "OWNER LEADERBOARD — KPI ca nhan (xep hang theo % script da het loi that su)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 22
    ws["A2"] = ("Resolution_Rate = so script da HET LOI HOAN TOAN / so script khac nhau da tung fix (KPI chinh). "
                "Verification_Rate = ty le cac lan fix duoc xac nhan dung ngay lan dau, khong bi reopen. "
                "Open_Workload = so script dang duoc gan cho nguoi nay ma van con loi (chua Done).")
    ws["A2"].font = ITALIC_HINT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 28

    lb_headers = ["Rank", "Owner", "Scripts_Fixed (khac nhau)", "Scripts_Fully_Resolved",
                  "Resolution_Rate", "Verified", "Reopened", "Pending", "Verification_Rate", "Open_Workload"]
    hr = 4
    for i, h in enumerate(lb_headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(lb_headers))
    for i, o in enumerate(owner_stats, start=hr + 1):
        ws.cell(row=i, column=1, value=o["rank"])
        ws.cell(row=i, column=2, value=o["owner"])
        ws.cell(row=i, column=3, value=o["distinct_scripts_fixed"])
        ws.cell(row=i, column=4, value=o["distinct_scripts_fully_resolved"])
        rc = ws.cell(row=i, column=5, value=o["resolution_rate"])
        rc.number_format = "0.0%"
        ws.cell(row=i, column=6, value=o["verified"])
        ws.cell(row=i, column=7, value=o["reopened"])
        ws.cell(row=i, column=8, value=o["pending"])
        vc = ws.cell(row=i, column=9, value=o["verification_rate"])
        vc.number_format = "0.0%"
        ws.cell(row=i, column=10, value=o["open_workload"])
    lb_last = hr + len(owner_stats)
    if owner_stats:
        tab = Table(displayName="tblOwnerLeaderboard", ref=f"A{hr}:J{lb_last}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(tab)

        chart = BarChart()
        chart.title = "Resolution Rate theo Owner"
        chart.y_axis.numFmt = "0%"
        data = Reference(ws, min_col=5, min_row=hr, max_row=lb_last)
        cats = Reference(ws, min_col=2, min_row=hr + 1, max_row=lb_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height, chart.width = 9, 20
        ws.add_chart(chart, f"A{lb_last + 3}")
    autosize(ws, start_row=4)

    # ================================================================
    # SHEET: RootCause_Pareto
    # ================================================================
    ws = wb.create_sheet("RootCause_Pareto")
    ws["A1"] = "ROOT CAUSE PARETO (Top 15, tinh tren toan bo lich su)"
    ws["A1"].font = TITLE_FONT
    rc_headers = ["Description (nguyen nhan)", "Fail_Count", "% of Total Fails", "Cumulative %"]
    hr = 3
    for i, h in enumerate(rc_headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(rc_headers))
    for i, r in enumerate(root_causes, start=hr + 1):
        ws.cell(row=i, column=1, value=r["description"])
        ws.cell(row=i, column=2, value=r["count"])
        c3 = ws.cell(row=i, column=3, value=r["pct"]); c3.number_format = "0.0%"
        c4 = ws.cell(row=i, column=4, value=r["cum_pct"]); c4.number_format = "0.0%"
    rc_last = hr + len(root_causes)
    if root_causes:
        chart = BarChart()
        chart.title = "Root Cause Pareto"
        data = Reference(ws, min_col=2, min_row=hr, max_row=rc_last)
        cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=rc_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height, chart.width = 9, 22
        ws.add_chart(chart, f"A{rc_last + 3}")
    autosize(ws, start_row=3, max_w=50)

    # ================================================================
    # SHEET: Test_Suite_Summary
    # ================================================================
    ws = wb.create_sheet("Test_Suite_Summary")
    ws["A1"] = "TIEN DO THEO TEST SUITE"
    ws["A1"].font = TITLE_FONT
    ts_headers = ["Test suite", "Tong so script", "Da xong (Done)", "Con loi", "% Hoan thanh"]
    hr = 3
    for i, h in enumerate(ts_headers, start=1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, len(ts_headers))
    for i, s in enumerate(suite_stats, start=hr + 1):
        ws.cell(row=i, column=1, value=s["test_suite"])
        ws.cell(row=i, column=2, value=s["total"])
        ws.cell(row=i, column=3, value=s["done"])
        ws.cell(row=i, column=4, value=s["still_failing"])
        c5 = ws.cell(row=i, column=5, value=s["done_pct"]); c5.number_format = "0.0%"
    ts_last = hr + len(suite_stats)
    if suite_stats:
        chart = BarChart()
        chart.title = "% Hoan thanh theo Test Suite"
        chart.y_axis.numFmt = "0%"
        data = Reference(ws, min_col=5, min_row=hr, max_row=ts_last)
        cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=ts_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height, chart.width = 9, 20
        ws.add_chart(chart, f"A{ts_last + 3}")
    autosize(ws, start_row=3)

    # ================================================================
    # SHEET: Daily_Fix_Log (raw)
    # ================================================================
    ws = wb.create_sheet("Daily_Fix_Log")
    fheaders = ["Fix_Date", "Owner", "Test suite", "Test Case", "Model_Fixed", "Fixed_After_Cycle", "Root_Cause", "Note"]
    ws.append(fheaders)
    style_header_row(ws, 1, len(fheaders))
    fixes_all = db.execute("SELECT * FROM fixes ORDER BY id").fetchall()
    for r in fixes_all:
        rc = r["root_cause"] if "root_cause" in r.keys() else ""
        ws.append([r["fix_date"], r["owner"], r["test_suite"], r["test_case"],
                   r["model_fixed"], r["fixed_after_cycle"], rc, r["note"]])
    if fixes_all:
        tab = Table(displayName="tblFixLogExport", ref=f"A1:H{len(fixes_all) + 1}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(tab)
    ws.freeze_panes = "A2"
    autosize(ws)

    # ================================================================
    # SHEET: History_Log (raw)
    # ================================================================
    ws = wb.create_sheet("History_Log")
    hheaders = ["Cycle", "Cycle_Date", "Test ID", "Model", "Test suite", "Test Case", "State", "Description", "Result"]
    ws.append(hheaders)
    style_header_row(ws, 1, len(hheaders))
    results_all = db.execute("SELECT * FROM results ORDER BY id").fetchall()
    for r in results_all:
        row_vals = [r["cycle"], r["cycle_date"], r["test_id"], r["model"], r["test_suite"],
                    r["test_case"], r["state"], r["description"], r["result"]]
        ws.append(row_vals)
    if results_all:
        tab = Table(displayName="tblHistoryExport", ref=f"A1:I{len(results_all) + 1}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(tab)
    ws.freeze_panes = "A2"
    autosize(ws)

    wb._sheets = [wb[n] for n in [
        "Instructions", "Dashboard", "Script_Priority_Tracker", "Owner_Leaderboard",
        "RootCause_Pareto", "Test_Suite_Summary", "Daily_Fix_Log", "History_Log",
    ]]
    wb.active = 1
    return wb


# ==================================================================
# INTEGRATIONS: farm API / he thong cong ty / GitHub script repo
# Cac adapter goi API ngoai deu la STUB cho toi khi co tai lieu API that
# (bao loi ro rang khi chua cau hinh); fallback nhap tay hoat dong day du.
# ==================================================================
def normalize_farm_rows(test_id, payload):
    """Chuyen JSON tra ve tu farm API (GET .../result?...&requestId={test_id}) thanh list row
    cho insert_result_rows(). Payload dang {"message","data":{"content":[...]},"success"}, moi
    phan tu content[i] = 1 lan chay script (deviceList[0].model = model that su chay).
    Bo qua record thieu model/scriptPath (khong du du lieu de chen mot cach an toan)."""
    data = payload.get("data") if isinstance(payload, dict) else None
    content = data.get("content") if isinstance(data, dict) else None
    if content is None:
        raise ValueError(
            "Response farm không đúng định dạng — thiếu data.content, kiểm tra lại URL API."
        )
    if not isinstance(content, list):
        raise ValueError("Response farm không đúng định dạng — data.content không phải danh sách.")

    rows = []
    for rec in content:
        if not isinstance(rec, dict):
            continue
        device_list = rec.get("deviceList") or []
        model = device_list[0].get("model") if device_list and isinstance(device_list[0], dict) else None
        script_path = rec.get("scriptPath") or ""
        if not model or not script_path:
            continue
        rows.append({
            "test_id": rec.get("requestId") or test_id,
            "model": model,
            "test_suite": rec.get("testFilePath") or "",
            "test_case": script_path,
            "state": rec.get("state") or "",
            "description": rec.get("description") or "",
        })
    return rows


def farm_fetch_results(db, test_ids):
    """Goi farm API lay ket qua theo tung Test ID. Tra ve (rows, errors_per_test_id).
    URL cau hinh o setting 'farm_api_url' — co the chua placeholder {test_id},
    neu khong co thi Test ID duoc noi vao cuoi URL."""
    url_tpl = get_setting(db, "farm_api_url").strip()
    token = get_setting(db, "farm_api_token").strip()
    if not url_tpl:
        raise RuntimeError(
            "Farm API chưa được cấu hình — điền 'URL API farm' trong tab Cài đặt (mục Tích hợp & API) trước."
        )
    all_rows, errors = [], []
    for tid in test_ids:
        try:
            if "{test_id}" in url_tpl:
                url = url_tpl.replace("{test_id}", urllib.parse.quote(str(tid)))
            else:
                url = url_tpl.rstrip("/") + "/" + urllib.parse.quote(str(tid))
            payload = _http_json(url, token=token)
            all_rows.extend(normalize_farm_rows(tid, payload))
        except (RuntimeError, ValueError) as e:
            errors.append({"test_id": tid, "error": str(e)})
    return all_rows, errors


@app.route("/api/results/import", methods=["POST"])
def api_import_results():
    """Import ket qua khong can browser session — de script phia farm push len tu dong.
    Auth chap nhan 1 trong 3: (1) session dang nhap co quyen input-results,
    (2) header X-Import-Token khop setting 'import_token' (phai dat truoc, khong rong),
    (3) header X-Admin-Key. Body giong POST /api/results: {rows: [...], created_by}."""
    db = get_db()
    authed = False
    u = current_user()
    if u and "input-results" in u["permissions"]:
        authed = True
    if not authed:
        tok = request.headers.get("X-Import-Token", "")
        conf = get_setting(db, "import_token").strip()
        if tok and conf and tok == conf:
            authed = True
    if not authed and request.headers.get("X-Admin-Key", "") == ADMIN_SECRET_KEY:
        authed = True
    if not authed:
        return jsonify({"error": "Không có quyền import — cần đăng nhập (quyền input-results) hoặc header X-Import-Token hợp lệ."}), 401

    payload = request.get_json(force=True)
    rows = payload.get("rows", [])
    if not rows:
        return jsonify({"error": "No rows provided"}), 400
    created_by = str(payload.get("created_by") or "").strip() or (u["username"] if u else "farm-import")
    summary = insert_result_rows(db, rows, created_by)
    recompute_cycles(db)
    log_audit(db, "results.import", detail=f"inserted={summary['inserted']}, dup={summary['skipped_duplicate']}, errors={len(summary['errors'])}")
    db.commit()
    return jsonify(summary)


@app.route("/api/integrations/farm/fetch", methods=["POST"])
@require_perm("integrations")
def api_farm_fetch():
    """Fetch ket qua tu farm API theo danh sach Test ID roi chen vao results.
    Body: {test_ids: [...]} hoac {test_ids: "chuoi cach nhau boi xuong dong/dau phay"}."""
    data = request.get_json(force=True)
    test_ids = data.get("test_ids") or []
    if isinstance(test_ids, str):
        test_ids = re.split(r"[\s,;]+", test_ids)
    test_ids = [str(t).strip() for t in test_ids if str(t).strip()]
    if not test_ids:
        return jsonify({"error": "Chưa nhập Test ID nào."}), 400

    db = get_db()
    try:
        rows, fetch_errors = farm_fetch_results(db, test_ids)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 400

    if rows:
        u = current_user()
        summary = insert_result_rows(db, rows, (u["username"] if u else "farm-fetch"))
        recompute_cycles(db)
    else:
        summary = {"inserted": 0, "skipped_running": 0, "skipped_duplicate": 0,
                   "duplicates": [], "errors": [], "warnings": []}
    summary["fetch_errors"] = fetch_errors
    summary["fetched_ok"] = len(test_ids) - len(fetch_errors)
    log_audit(db, "integrations.farm_fetch",
              detail=f"test_ids={len(test_ids)}, ok={summary['fetched_ok']}, inserted={summary['inserted']}")
    db.commit()
    return jsonify(summary)


# ------------------------------------------------------------------
# He thong cong ty: tong so test case can script (dong), status Performed/SKIP...
# ------------------------------------------------------------------
def company_fetch_testcases(db):
    """Goi API TC Hub (GET .../api/tc/{item}, tra ve list TC cua 1 item/request) lay danh sach
    test case + automationTarget (performed/excluded/target) cho tung item. Danh sach item lay
    tu setting 'company_items' (CSV), rong -> fallback toan bo test_suites. Auth bang cookie
    phien dang nhap (setting 'company_cookie'), KHONG dung Bearer token.
    Tra ve (rows, errors) - loi 1 item khong chan cac item con lai. Row: {tc_id, item, status, raw}."""
    url_tpl = get_setting(db, "company_api_url").strip()
    if not url_tpl:
        raise RuntimeError(
            "API TC Hub chưa được cấu hình — điền 'URL API TC Hub' trong tab Cài đặt trước."
        )
    cookie = get_setting(db, "company_cookie").strip()
    items_raw = get_setting(db, "company_items").strip()
    if items_raw:
        items = [s.strip() for s in items_raw.split(",") if s.strip()]
    else:
        items = [r["name"] for r in db.execute("SELECT name FROM test_suites ORDER BY name").fetchall()]
    if not items:
        raise RuntimeError(
            "Không có Item nào để đồng bộ — điền 'Danh sách Item' trong Cài đặt hoặc thêm Test Suite trước."
        )

    rows, errors = [], []
    for item in items:
        try:
            if "{item}" in url_tpl:
                url = url_tpl.replace("{item}", urllib.parse.quote(str(item)))
            else:
                url = url_tpl.rstrip("/") + "/" + urllib.parse.quote(str(item))
            data = _http_json(url, cookie=cookie)
            if not isinstance(data, list):
                raise ValueError("response không phải danh sách TC")
            for rec in data:
                if not isinstance(rec, dict):
                    continue
                tc_id = str(rec.get("id") or "").strip()
                if not tc_id:
                    continue
                rows.append({
                    "tc_id": tc_id,
                    "item": str(rec.get("app") or item).strip() or item,
                    "status": str(rec.get("automationTarget") or "").strip(),
                    "raw": "",
                })
        except (RuntimeError, ValueError) as e:
            errors.append({"item": item, "error": str(e)})
    return rows, errors


def _replace_company_testcases(db, rows, source):
    """Full refresh cache company_testcases (snapshot moi thay toan bo snapshot cu)."""
    db.execute("DELETE FROM company_testcases")
    n = 0
    for r in rows:
        tc_id = str(r.get("tc_id") or "").strip()
        if not tc_id:
            continue
        status = str(r.get("status") or "").strip() or "Not Performed"
        item = str(r.get("item") or "").strip() or item_from_tc_id(tc_id) or "Unknown"
        raw = str(r.get("raw") or "")
        db.execute(
            "INSERT OR REPLACE INTO company_testcases (tc_id, item, status, raw, source, synced_at) "
            "VALUES (?,?,?,?,?, datetime('now'))",
            (tc_id, item, status, raw, source),
        )
        n += 1
    return n


@app.route("/api/integrations/company/sync", methods=["POST"])
@require_perm("integrations")
def api_company_sync():
    db = get_db()
    try:
        rows, errors = company_fetch_testcases(db)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 400
    if not rows and errors:
        return jsonify({"error": "Đồng bộ thất bại toàn bộ item.", "errors": errors}), 400
    n = _replace_company_testcases(db, rows, "api")
    log_audit(db, "integrations.company_sync", detail=f"synced={n}, errors={len(errors)}")
    db.commit()
    return jsonify({"status": "ok", "synced": n, "errors": errors})


@app.route("/api/integrations/company/manual", methods=["POST"])
@require_perm("integrations")
def api_company_manual():
    """Fallback nhap tay: {rows: [{tc_id, status?, item?}], mode: "replace"|"merge"}.
    replace (mac dinh) = thay toan bo snapshot cache; merge = upsert tung dong."""
    data = request.get_json(force=True)
    rows = data.get("rows") or []
    mode = (data.get("mode") or "replace").strip().lower()
    if not rows:
        return jsonify({"error": "Không có dòng nào."}), 400
    db = get_db()
    if mode == "merge":
        n = 0
        for r in rows:
            tc_id = str(r.get("tc_id") or "").strip()
            if not tc_id:
                continue
            status = str(r.get("status") or "").strip() or "Not Performed"
            item = str(r.get("item") or "").strip() or item_from_tc_id(tc_id) or "Unknown"
            db.execute(
                "INSERT OR REPLACE INTO company_testcases (tc_id, item, status, raw, source, synced_at) "
                "VALUES (?,?,?,?, 'manual', datetime('now'))",
                (tc_id, item, status, ""),
            )
            n += 1
    else:
        n = _replace_company_testcases(db, rows, "manual")
    log_audit(db, "integrations.company_manual", detail=f"mode={mode}, imported={n}")
    db.commit()
    return jsonify({"status": "ok", "imported": n, "mode": mode})


@app.route("/api/integrations/company/testcases")
@require_login
def api_company_testcases():
    db = get_db()
    rows = db.execute(
        "SELECT tc_id, item, status, source, synced_at FROM company_testcases ORDER BY item, tc_id"
    ).fetchall()
    total = len(rows)
    skip = sum(1 for r in rows if (r["status"] or "").strip().lower() in COMPANY_SKIP_STATES)
    performed = sum(1 for r in rows if (r["status"] or "").strip().lower() in COMPANY_PERFORMED_STATES)
    # "target" (automationTarget) = TC can hoan thanh script nhung chua xong.
    target = sum(1 for r in rows if (r["status"] or "").strip().lower() in COMPANY_TARGET_STATES)
    synced_at = max((r["synced_at"] or "" for r in rows), default="")
    return jsonify({
        "summary": {"total": total, "total_needed": total - skip, "skip": skip,
                    "performed": performed, "target": target, "synced_at": synced_at},
        "rows": [dict(r) for r in rows],
    })


# ------------------------------------------------------------------
# GitHub script repo: danh sach file script tren nhanh main (doi chieu 3 chieu)
# ------------------------------------------------------------------
@app.route("/api/integrations/github/sync", methods=["POST"])
@require_perm("integrations")
def api_github_sync():
    db = get_db()
    repo = get_setting(db, "github_repo").strip()
    branch = get_setting(db, "github_branch").strip() or "main"
    token = get_setting(db, "github_token").strip()
    base = get_setting(db, "github_api_base").strip() or "https://api.github.com"
    if not repo or "/" not in repo:
        return jsonify({"error": "GitHub repo chưa được cấu hình — điền 'owner/repo' trong tab Cài đặt (mục Tích hợp & API)."}), 400
    url = f"{base.rstrip('/')}/repos/{repo}/git/trees/{urllib.parse.quote(branch)}?recursive=1"
    try:
        data = _http_json(url, token=token)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 400
    tree = data.get("tree") or []
    paths = [t["path"] for t in tree
             if t.get("type") == "blob" and str(t.get("path", "")).lower().endswith(".py")]
    db.execute("DELETE FROM repo_files")
    for p in paths:
        db.execute(
            "INSERT OR REPLACE INTO repo_files (path, source, synced_at) VALUES (?, 'api', datetime('now'))",
            (p,),
        )
    truncated = bool(data.get("truncated"))
    log_audit(db, "integrations.github_sync", detail=f"files={len(paths)}, truncated={truncated}")
    db.commit()
    out = {"status": "ok", "files": len(paths), "truncated": truncated}
    if truncated:
        out["warning"] = "GitHub trả về danh sách bị cắt (repo quá lớn) — kết quả đối chiếu có thể thiếu file."
    return jsonify(out)


@app.route("/api/integrations/github/manual", methods=["POST"])
@require_perm("integrations")
def api_github_manual():
    """Fallback nhap tay: {paths: [...]} — paste output cua
    `git ls-tree -r main --name-only` tu repo script."""
    data = request.get_json(force=True)
    paths = data.get("paths") or []
    if isinstance(paths, str):
        paths = paths.splitlines()
    paths = [str(p).strip().replace("\\", "/").lstrip("/") for p in paths if str(p).strip()]
    if not paths:
        return jsonify({"error": "Không có đường dẫn file nào."}), 400
    db = get_db()
    db.execute("DELETE FROM repo_files")
    for p in paths:
        db.execute(
            "INSERT OR REPLACE INTO repo_files (path, source, synced_at) VALUES (?, 'manual', datetime('now'))",
            (p,),
        )
    log_audit(db, "integrations.github_manual", detail=f"files={len(paths)}")
    db.commit()
    return jsonify({"status": "ok", "imported": len(paths)})


@app.route("/api/integrations/status")
@require_login
def api_integrations_status():
    """Trang thai cau hinh + cache cua cac tich hop, cho tab Dong bo hien thi."""
    db = get_db()
    comp = db.execute(
        "SELECT COUNT(*) c, MAX(synced_at) s, MAX(source) src FROM company_testcases"
    ).fetchone()
    gh = db.execute("SELECT COUNT(*) c, MAX(synced_at) s, MAX(source) src FROM repo_files").fetchone()
    return jsonify({
        "farm_configured": bool(get_setting(db, "farm_api_url").strip()),
        "company_configured": bool(get_setting(db, "company_api_url").strip()),
        "github_configured": bool(get_setting(db, "github_repo").strip()),
        "import_token_set": bool(get_setting(db, "import_token").strip()),
        "company_cache": {"rows": comp["c"], "synced_at": comp["s"] or "", "source": comp["src"] or ""},
        "github_cache": {"files": gh["c"], "synced_at": gh["s"] or "", "source": gh["src"] or ""},
    })


@app.route("/api/integrations/reconcile")
@require_login
def api_reconcile():
    """Doi chieu 3 chieu, ca DONE lan SKIP:
    - DONE: ky vong status Performed ben TC Hub + CO file script that tren nhanh main.
    - SKIP: ky vong status Excluded ben TC Hub + KHONG CON file script (da xoa).
    Member/Team tu dong dien (khi new_scripts chua co san) bang cach tra assignment hien tai
    cua TC do + team cua owner do (chi tinh luc doi chieu, khong ghi de DB)."""
    db = get_db()
    comp_status = {}
    comp_synced = ""
    for r in db.execute("SELECT tc_id, status, synced_at FROM company_testcases").fetchall():
        comp_status[str(r["tc_id"]).strip().lower()] = (r["status"] or "").strip()
        comp_synced = max(comp_synced, r["synced_at"] or "")
    has_company = bool(comp_status)

    stem_index = {}  # ten file (bo .py, lowercase) -> [duong dan day du]
    gh_synced = ""
    n_files = 0
    for r in db.execute("SELECT path, synced_at FROM repo_files").fetchall():
        n_files += 1
        gh_synced = max(gh_synced, r["synced_at"] or "")
        base = r["path"].rsplit("/", 1)[-1]
        stem = base[:-3] if base.lower().endswith(".py") else base
        stem_index.setdefault(stem.lower(), []).append(r["path"])
    has_github = n_files > 0

    suite_paths = {
        r["name"]: (r["script_path"] or "").strip().strip("/")
        for r in db.execute("SELECT name, script_path FROM test_suites").fetchall()
    }
    assign_owner = {
        (r["test_suite"], str(r["test_case"]).strip().lower()): r["owner"]
        for r in db.execute("SELECT test_suite, test_case, owner FROM assignments").fetchall()
    }
    owner_team = {
        r["name"]: (r["team"] or "")
        for r in db.execute("SELECT name, team FROM owners").fetchall()
    }

    rows_out = []
    n_ok = n_missing_company = n_wrong_company = n_missing_github = 0
    n_skip_wrong_company = n_skip_has_github = 0
    total_done = total_skip = 0
    done_ids = set()
    for r in db.execute(
        "SELECT tc_id, item, member, team, status, completed_date FROM new_scripts "
        "WHERE status IN ('DONE','SKIP') ORDER BY item, tc_id"
    ).fetchall():
        kind = r["status"]
        tc_norm = str(r["tc_id"]).strip().lower()
        if kind == "DONE":
            done_ids.add(tc_norm)
            total_done += 1
        else:
            total_skip += 1
        company_status = comp_status.get(tc_norm)

        sp = suite_paths.get(r["item"], "")
        matched_path = None
        for p in stem_index.get(tc_norm, []):
            if not sp or p.lower().startswith(sp.lower() + "/"):
                matched_path = p
                break
        has_file = matched_path is not None

        if kind == "DONE":
            company_ok = bool(company_status) and company_status.lower() in COMPANY_PERFORMED_STATES
            github_ok = has_file
            if has_company:
                if company_status is None:
                    n_missing_company += 1
                elif not company_ok:
                    n_wrong_company += 1
            if has_github and not github_ok:
                n_missing_github += 1
        else:  # SKIP: nguoc lai - khong lam nen khong can Performed, khong can con file
            company_ok = bool(company_status) and company_status.lower() in COMPANY_SKIP_STATES
            github_ok = not has_file
            if has_company and not company_ok:
                n_skip_wrong_company += 1
            if has_github and has_file:
                n_skip_has_github += 1

        ok = (company_ok or not has_company) and (github_ok or not has_github)
        if ok and (has_company or has_github):
            n_ok += 1

        member = r["member"] or assign_owner.get((r["item"], tc_norm), "") or ""
        team = r["team"] or owner_team.get(member, "") or ""

        rows_out.append({
            "kind": kind, "tc_id": r["tc_id"], "item": r["item"], "member": member,
            "team": team, "completed_date": r["completed_date"] or "",
            "company_status": company_status,
            "company_ok": company_ok if has_company else None,
            "github_ok": github_ok if has_github else None,
            "matched_path": matched_path,
            "path_configured": bool(sp),
            "ok": ok,
        })

    # Chieu nguoc: TC Hub bao Performed nhung he thong chua ghi DONE (top 50)
    performed_not_done = []
    if has_company:
        for tc_norm, status in comp_status.items():
            if status.lower() in COMPANY_PERFORMED_STATES and tc_norm not in done_ids:
                performed_not_done.append(tc_norm)
                if len(performed_not_done) >= 50:
                    break

    return jsonify({
        "summary": {
            "total_done": total_done,
            "total_skip": total_skip,
            "ok_all": n_ok,
            "missing_company": n_missing_company,
            "wrong_company_status": n_wrong_company,
            "missing_github": n_missing_github,
            "skip_wrong_company": n_skip_wrong_company,
            "skip_has_github": n_skip_has_github,
            "has_company_data": has_company,
            "has_github_data": has_github,
            "company_synced_at": comp_synced,
            "github_synced_at": gh_synced,
            "github_files": n_files,
        },
        "rows": rows_out,
        "company_performed_not_done": performed_not_done,
    })


# ==================================================================
# REPORTS: sinh markdown bao cao ngay/tuan theo mau docs/maubaocao.md
# ==================================================================
def _pct(x, digits=1):
    return f"{x * 100:.{digits}f}%" if x is not None else "…"


def _num(x):
    return str(x) if x is not None else "…"


def _suite_model_report_table(db, cycle_nums):
    """Bang markdown: pass rate tung Item x tung Model + dong tong tung Item + TONG CHUNG,
    theo tung cycle trong cycle_nums (>=3 cycle gan nhat), kem delta dau-cuoi."""
    data = compute_suite_model_matrix(db)
    cyc_dates = {c["cycle"]: (c["cycle_date"] or "") for c in data["cycles"]}
    sel = [c for c in cycle_nums if c in cyc_dates]
    if not sel:
        return "_No result data._"

    def fmt_cell(cell):
        return _pct(cell["pass_rate"]) if cell else "—"

    header = "| Item | Model | " + " | ".join(
        f"C{c} ({cyc_dates[c][5:] if cyc_dates[c] else ''})" for c in sel
    ) + " | Δ |"
    sep = "|" + "---|" * (len(sel) + 3)
    lines = [header, sep]

    # Gom row theo item de tinh dong tong tung item
    by_item = {}
    for row in data["rows"]:
        by_item.setdefault(row["test_suite"], []).append(row)

    def delta_str(first_rate, last_rate):
        if first_rate is None or last_rate is None:
            return "…"
        d = (last_rate - first_rate) * 100
        arrow = "▲" if d > 0 else ("▼" if d < 0 else "=")
        return f"{arrow}{abs(d):.1f}%"

    for item in sorted(by_item.keys()):
        rows = by_item[item]
        # tung model
        for row in sorted(rows, key=lambda r: r["model"]):
            cells = [row["by_cycle"].get(c) for c in sel]
            rates = [c["pass_rate"] if c else None for c in cells]
            first = next((x for x in rates if x is not None), None)
            last = next((x for x in reversed(rates) if x is not None), None)
            lines.append(
                f"| {item} | {row['model']} | " + " | ".join(fmt_cell(c) for c in cells)
                + f" | {delta_str(first, last)} |"
            )
        # dong tong cua item (gop moi model)
        agg_cells = []
        for c in sel:
            tot = pas = na = 0
            found = False
            for row in rows:
                cell = row["by_cycle"].get(c)
                if cell:
                    found = True
                    tot += cell["total"]
                    pas += cell["pass_count"]
                    na += cell["na_count"]
            denom = tot - na
            agg_cells.append({"pass_rate": (pas / denom) if (found and denom > 0) else None} if found else None)
        rates = [c["pass_rate"] if c else None for c in agg_cells]
        first = next((x for x in rates if x is not None), None)
        last = next((x for x in reversed(rates) if x is not None), None)
        lines.append(
            f"| **{item}** | **Total** | " + " | ".join(f"**{fmt_cell(c)}**" for c in agg_cells)
            + f" | **{delta_str(first, last)}** |"
        )

    # GRAND TOTAL tren tat ca ket qua
    overall_cells = [data["overall_by_cycle"].get(c) for c in sel]
    rates = [c["pass_rate"] if c else None for c in overall_cells]
    first = next((x for x in rates if x is not None), None)
    last = next((x for x in reversed(rates) if x is not None), None)
    lines.append(
        "| **GRAND TOTAL** | **All** | " + " | ".join(f"**{fmt_cell(c)}**" for c in overall_cells)
        + f" | **{delta_str(first, last)}** |"
    )
    return "\n".join(lines)


def _env_fail_by_cycle(db, cycle_nums):
    """{cycle: {fail, env, rate}} — so loi lien quan MOI TRUONG (nhom 'Hạ tầng:' tu
    summarize_root_cause) tren tong fail cua tung cycle."""
    out = {c: {"fail": 0, "env": 0} for c in cycle_nums}
    if not cycle_nums:
        return out
    qmarks = ",".join("?" * len(cycle_nums))
    for r in db.execute(
        f"SELECT cycle, description, state FROM results WHERE result='Fail' AND cycle IN ({qmarks})",
        list(cycle_nums),
    ).fetchall():
        d = out[r["cycle"]]
        d["fail"] += 1
        if summarize_root_cause(r["description"], r["state"]).startswith("Hạ tầng"):
            d["env"] += 1
    for c, d in out.items():
        d["rate"] = (d["env"] / d["fail"]) if d["fail"] else None
    return out


def _person_count_table(db, table, person_col, date_col, date_from, date_to, extra_where=""):
    """Bang markdown dem so dong theo nguoi trong khoang ngay [date_from, date_to]."""
    rows = db.execute(
        f"SELECT COALESCE(NULLIF(TRIM({person_col}), ''), '(unknown)') person, COUNT(*) c "
        f"FROM {table} WHERE {date_col} >= ? AND {date_col} <= ? {extra_where} "
        f"GROUP BY person ORDER BY c DESC, person",
        (date_from, date_to),
    ).fetchall()
    total = sum(r["c"] for r in rows)
    if not rows:
        return 0, 0, "_None._"
    lines = ["| Person | Count |", "|---|---|"]
    for r in rows:
        lines.append(f"| {r['person']} | {r['c']} |")
    lines.append(f"| **Total** | **{total}** |")
    return total, len(rows), "\n".join(lines)


def _velocity_last7(db, table, date_col, extra_where=""):
    """Toc do trung binh (dong/ngay) trong 7 ngay gan nhat theo date_col."""
    since = (date.today() - timedelta(days=6)).isoformat()
    n = db.execute(
        f"SELECT COUNT(*) c FROM {table} WHERE {date_col} >= ? AND {date_col} <= ? {extra_where}",
        (since, date.today().isoformat()),
    ).fetchone()["c"]
    return n / 7.0


def _eta_section(db, kpi, coverage):
    """Muc 'Uoc luong hoan thanh du an' — 2 duong: stabilization (fix) va viet moi."""
    lines = []
    today = date.today()
    deadline = get_setting(db, "deadline_date").strip()
    fix_speed = _velocity_last7(db, "fixes", "fix_date")
    write_speed = _velocity_last7(db, "new_scripts", "completed_date", "AND status='DONE'")

    eta_dates = []
    # (a) stabilization: scripts left to fix to reach target / fix velocity
    fails_to_fix = kpi.get("fails_to_fix")
    if fails_to_fix is not None:
        if fails_to_fix <= 0:
            lines.append("- **Stabilization**: current pass-rate target already met ✅")
        elif fix_speed > 0:
            days = int(fails_to_fix / fix_speed + 0.999)
            eta = today + timedelta(days=days)
            eta_dates.append(eta)
            lines.append(
                f"- **Stabilization**: {fails_to_fix} scripts left to fix ÷ {fix_speed:.1f} fixes/day "
                f"(7-day avg) → target reached ~**{eta.strftime('%Y-%m-%d')}** (~{days} days)"
            )
        else:
            lines.append(f"- **Stabilization**: {fails_to_fix} scripts left to fix — cannot estimate (no fixes in the last 7 days)")
    # (b) new scripts: remaining TCs / write velocity
    if coverage.get("configured"):
        remaining = coverage["total_needed"] - coverage["done"]
        if remaining <= 0:
            lines.append("- **New scripts**: all required TCs covered ✅")
        elif write_speed > 0:
            days = int(remaining / write_speed + 0.999)
            eta = today + timedelta(days=days)
            eta_dates.append(eta)
            lines.append(
                f"- **New scripts**: {remaining} TCs left ÷ {write_speed:.1f} scripts/day (7-day avg) "
                f"→ coverage complete ~**{eta.strftime('%Y-%m-%d')}** (~{days} days)"
            )
        else:
            lines.append(f"- **New scripts**: {remaining} TCs left — cannot estimate (no DONE scripts in the last 7 days)")
    else:
        lines.append("- **New scripts**: company system not synced — total required TCs unknown")

    if eta_dates:
        eta_all = max(eta_dates)
        line = f"- **Project ETA at current pace: {eta_all.strftime('%Y-%m-%d')}**"
        if deadline:
            try:
                dl = date.fromisoformat(deadline)
                diff = (dl - eta_all).days
                line += f" — deadline {dl.strftime('%Y-%m-%d')}: " + (
                    f"**on track, {diff} days spare** ✅" if diff >= 0 else f"**~{-diff} days late** ⚠️"
                )
            except ValueError:
                pass
        lines.append(line)
    return "\n".join(lines)


def _report_kpis(db, priority=None):
    """Cac con so dieu hanh dung chung cho ca 2 bao cao (giong logic /api/dashboard)."""
    if priority is None:
        priority = get_script_priority(db)
    settings = {r["key"]: r["value"] for r in db.execute("SELECT key, value FROM settings")}
    target_rate = float(settings.get("target_pass_rate") or 0.88)
    total_scripts = len(priority)
    still_failing = sum(1 for p in priority if p["priority_tier"] in ("P0", "P1", "P2", "P3"))
    fails_to_fix = max(0, still_failing - int(total_scripts * (1 - target_rate)))
    days_remaining = required_rate = None
    deadline = settings.get("deadline_date") or ""
    if deadline:
        try:
            days_remaining = (date.fromisoformat(deadline) - date.today()).days
            if days_remaining and days_remaining > 0:
                required_rate = fails_to_fix / days_remaining
        except ValueError:
            pass
    tier_counts = {"P0": 0, "P1": 0, "P2": 0, "P3": 0, "Verify": 0, "Done": 0}
    for p in priority:
        tier_counts[p["priority_tier"]] += 1
    return {
        "target_rate": target_rate, "total_scripts": total_scripts,
        "still_failing": still_failing, "fails_to_fix": fails_to_fix,
        "days_remaining": days_remaining, "required_rate_per_day": required_rate,
        "tier_counts": tier_counts, "deadline": deadline,
        "flaky_count": sum(1 for p in priority if p.get("is_flaky")),
    }


def _completion_table_md(coverage):
    """Markdown 'Script completion by Test Suite' — nguon CHI TU he thong cong ty
    (compute_coverage). Cot: Item / Total TC / Completed script / Completed Rate + Total row.
    Cache rong -> ghi ro 'not synced' (khong bia so)."""
    if not coverage.get("configured"):
        return "_Company system not synced — total required TCs unknown (sync/paste in the 🔗 Sync tab)._"
    lines = ["| Item | Total TC | Completed script | Completed Rate |", "|---|---|---|---|"]
    for it in coverage["by_item"]:
        lines.append(f"| {it['item']} | {it['needed']} | {it['done']} | {_pct(it['pct'])} |")
    lines.append(f"| **Total** | **{coverage['total_needed']}** | **{coverage['done']}** | **{_pct(coverage['pct'])}** |")
    return "\n".join(lines)


def build_daily_report(db, date_str=None):
    """English markdown DAILY report. Cycle KHONG bat buoc: neu ngay do khong co cycle
    nhung co fix hoac script viet moi thi van sinh bao cao (muc dua-tren-cycle ghi
    'No cycle ran'). Chi loi khi ngay do khong co du lieu gi."""
    trend = compute_cycle_trend(db)
    # Xac dinh ngay bao cao: uu tien date_str; else cycle moi nhat; else hom nay.
    if not date_str:
        date_str = trend[-1]["cycle_date"] if trend else date.today().isoformat()
    entry = next((t for t in trend if t["cycle_date"] == date_str), None)
    idx = trend.index(entry) if entry is not None else None
    prev = trend[idx - 1] if (idx is not None and idx > 0) else None

    # Co bat ky du lieu tien do nao trong ngay khong?
    n_fix_today = db.execute("SELECT COUNT(*) c FROM fixes WHERE fix_date=?", (date_str,)).fetchone()["c"]
    n_new_today = db.execute(
        "SELECT COUNT(*) c FROM new_scripts WHERE status='DONE' AND completed_date=?", (date_str,)
    ).fetchone()["c"]
    if entry is None and not n_fix_today and not n_new_today:
        return None, f"Không có dữ liệu nào (cycle / fix / script viết mới) cho ngày {date_str}."

    priority = get_script_priority(db)
    kpi = _report_kpis(db, priority)
    coverage = compute_coverage(db)

    title_cycle = f"Cycle {entry['cycle']}" if entry is not None else "No cycle"
    md = [f"# [Stabilization Daily] {title_cycle} — {date_str}", ""]

    # ---- 1. Cycle result today ----
    md += ["## 1. Cycle result today", ""]
    if entry is not None:
        md += ["| Metric | Today | Prev cycle | Δ |", "|---|---|---|---|"]
        def _delta(cur, prv, pct=False, invert=False):
            if cur is None or prv is None:
                return "…"
            d = cur - prv
            good = (d <= 0) if invert else (d >= 0)
            arrow = "▲" if d > 0 else ("▼" if d < 0 else "=")
            mark = "🙂" if good and d != 0 else ("⚠️" if d != 0 else "")
            val = f"{abs(d)*100:.1f}%" if pct else f"{abs(d)}"
            return f"{arrow}{val} {mark}".strip()
        md.append(f"| Pass rate | {_pct(entry['pass_rate'])} | {_pct(prev['pass_rate']) if prev else '…'} | {_delta(entry['pass_rate'], prev['pass_rate'] if prev else None, pct=True)} |")
        md.append(f"| Total runs | {entry['total']} | {prev['total'] if prev else '…'} | |")
        md.append(f"| Failures in cycle | {entry['fail_count']} | {prev['fail_count'] if prev else '…'} | {_delta(entry['fail_count'], prev['fail_count'] if prev else None, invert=True)} |")
    else:
        md += [f"_No cycle ran on {date_str} — showing current standings and today's progress only._", ""]
        md += ["| Metric | Value |", "|---|---|"]
    md.append(f"| Still failing (current) | {kpi['still_failing']} / {kpi['total_scripts']} |" + (" |" if entry is None else " | |"))
    md.append(f"| P0 (wide-impact) | {kpi['tier_counts']['P0']} |" + (" |" if entry is None else " | |"))
    md.append(f"| Verifying (Verify) | {kpi['tier_counts']['Verify']} |" + (" |" if entry is None else " | |"))
    md.append(f"| Flaky scripts | {kpi['flaky_count']} |" + (" |" if entry is None else " | |"))
    md.append("")

    # ---- 2. Pass rate Item x Model (last >=3 cycles) ----
    if trend:
        end = (idx + 1) if idx is not None else len(trend)
        recent_cycles = [t["cycle"] for t in trend[:end]][-3:]
        md += [f"## 2. Pass rate Item × Model (last {len(recent_cycles)} cycles)", "",
               _suite_model_report_table(db, recent_cycles), ""]
    else:
        recent_cycles = []
        md += ["## 2. Pass rate Item × Model", "", "_No result data yet._", ""]

    # ---- 3. Environment-related failures ----
    md += ["## 3. Environment-related failures (Infra/Device)", ""]
    if recent_cycles:
        env = _env_fail_by_cycle(db, recent_cycles)
        md += ["| Cycle | Total fail | Env fail | Rate |", "|---|---|---|---|"]
        for c in recent_cycles:
            d = env[c]
            md.append(f"| C{c} | {d['fail']} | {d['env']} | {_pct(d['rate'])} |")
    else:
        md.append("_No cycle data._")
    md.append("")

    # ---- 4. Fixes today + verify previous cycle ----
    n_fix, n_fixers, fix_table = _person_count_table(db, "fixes", "owner", "fix_date", date_str, date_str)
    md += [f"## 4. Fixes today: **{n_fix}** fixes (by {n_fixers} people)", "", fix_table, ""]
    if prev:
        tracking = compute_fix_tracking(db)
        counts = {"verified": 0, "regressed": 0, "still_failing": 0, "pending": 0}
        attention = []
        for f in tracking:
            if f["fixed_after_cycle"] == prev["cycle"]:
                counts[f["status"]] += 1
                if f["status"] in ("still_failing", "regressed"):
                    attention.append(f"`{f['test_case']}` — {f['owner']} ({f['status_label']})")
        md += [f"**Verify fixes from previous cycle (C{prev['cycle']})**: "
               f"✅ resolved: {counts['verified']} · ⚠️ regressed: {counts['regressed']} · "
               f"❌ still failing: {counts['still_failing']} · ⏳ awaiting data: {counts['pending']}", ""]
        if attention:
            md += ["Needs attention:", *[f"- {a}" for a in attention], ""]

    # ---- 5. New scripts today ----
    n_new, n_writers, new_table = _person_count_table(
        db, "new_scripts", "member", "completed_date", date_str, date_str, "AND status='DONE'")
    md += [f"## 5. New scripts today: **{n_new}** DONE (by {n_writers} people)", "", new_table, ""]

    # ---- 6. Script completion by Test Suite (company system) ----
    md += ["## 6. Script completion by Test Suite", "", _completion_table_md(coverage), ""]

    # ---- 7. Notes & velocity ----
    md += ["## 7. Notes & velocity", ""]
    if entry is not None:
        model_rows = db.execute(
            "SELECT model, COUNT(*) t, SUM(CASE WHEN result='Pass' THEN 1 ELSE 0 END) p "
            "FROM results WHERE cycle=? AND result <> 'Excluded' GROUP BY model", (entry["cycle"],),
        ).fetchall()
        worst = min(((r["p"] / r["t"], r["model"]) for r in model_rows if r["t"]), default=None)
        if worst:
            md.append(f"- Weakest model this cycle: **{worst[1]}** ({_pct(worst[0])}) — device issue or script issue?")
    pareto = compute_root_cause_pareto(db, limit=3, english=True)
    if pareto:
        md.append(f"- Most common failure cause (all cycles): **{pareto[0]['description']}** ({pareto[0]['count']} failures, {_pct(pareto[0]['pct'])})")
    fix_speed = _velocity_last7(db, "fixes", "fix_date")
    req = kpi["required_rate_per_day"]
    if req is not None:
        verdict = "ENOUGH ✅" if fix_speed >= req else "BELOW ⚠️"
        md.append(f"- Actual fix velocity (7-day avg): **{fix_speed:.1f} fixes/day** vs required **{req:.1f}/day** → {verdict}")
    else:
        md.append(f"- Actual fix velocity (7-day avg): **{fix_speed:.1f} fixes/day** (no deadline set to compare against)")
    md.append("")

    # ---- 8. Project completion estimate ----
    md += ["## 8. Project completion estimate", "", _eta_section(db, kpi, coverage), ""]

    # ---- 9. Test farm (fill manually) ----
    md += ["## 9. Test farm (fill manually)", "",
           "| Metric | Value | Note |", "|---|---|---|",
           "| Active devices / total | …/… | |",
           "| Runs completed / planned | …/… | |",
           "| Leftover RUNNING results | … | |",
           "| Infra incidents today | … | |", ""]

    # ---- 10. Plan for tomorrow ----
    top_assigned = [p for p in priority if p["priority_tier"] in ("P0", "P1")]
    top_assigned.sort(key=lambda p: (-p["priority_score"], p["test_case"]))
    md += ["## 10. Plan for tomorrow", ""]
    if top_assigned:
        for p in top_assigned[:10]:
            owner = p["current_owner"] or "(unassigned)"
            md.append(f"- [{p['priority_tier']}] `{p['test_case']}` ({p['test_suite']}) — {owner}")
        if len(top_assigned) > 10:
            md.append(f"- … and {len(top_assigned) - 10} more P0/P1 scripts (see Priority table)")
    else:
        md.append("- No P0/P1 scripts left 🎉")
    md += ["", "## 11. Blockers needing support", "", "- …", ""]

    return "\n".join(md), None


def build_weekly_report(db, week_str=None):
    """English markdown WEEKLY report (ISO week 'YYYY-Wnn'). Cycle KHONG bat buoc:
    tuan khong co cycle nhung co fix / script viet moi thi van sinh bao cao (muc
    dua-tren-cycle ghi 'No cycle ran'). Chi loi khi tuan do khong co du lieu gi."""
    if week_str:
        m = re.match(r"^(\d{4})-W(\d{1,2})$", week_str.strip())
        if not m:
            return None, "Định dạng tuần không hợp lệ — dùng YYYY-Wnn (VD 2026-W28)."
        year, wk = int(m.group(1)), int(m.group(2))
    else:
        iso = date.today().isocalendar()
        year, wk = iso[0], iso[1]
        week_str = f"{year}-W{wk:02d}"
    try:
        monday = date.fromisocalendar(year, wk, 1)
    except ValueError:
        return None, f"Tuần {week_str} không hợp lệ."
    sunday = monday + timedelta(days=6)
    prev_monday = monday - timedelta(days=7)
    prev_sunday = monday - timedelta(days=1)
    mon_s, sun_s = monday.isoformat(), sunday.isoformat()

    trend = compute_cycle_trend(db)
    in_week = [t for t in trend if t["cycle_date"] and mon_s <= t["cycle_date"] <= sun_s]
    in_prev = [t for t in trend if t["cycle_date"] and prev_monday.isoformat() <= t["cycle_date"] <= prev_sunday.isoformat()]

    n_fix_week = db.execute("SELECT COUNT(*) c FROM fixes WHERE fix_date>=? AND fix_date<=?", (mon_s, sun_s)).fetchone()["c"]
    n_new_week = db.execute("SELECT COUNT(*) c FROM new_scripts WHERE status='DONE' AND completed_date>=? AND completed_date<=?", (mon_s, sun_s)).fetchone()["c"]
    if not in_week and not n_fix_week and not n_new_week:
        return None, f"Không có dữ liệu nào (cycle / fix / script viết mới) trong tuần {week_str} ({monday.strftime('%d/%m')}–{sunday.strftime('%d/%m')})."

    priority = get_script_priority(db)
    kpi = _report_kpis(db, priority)
    coverage = compute_coverage(db)
    first = in_week[0] if in_week else None
    last = in_week[-1] if in_week else None
    prev_last = in_prev[-1] if in_prev else None

    cyc_txt = f" — Cycle {first['cycle']}–{last['cycle']}" if in_week else " — no cycle"
    md = [f"# [Stabilization Weekly] Week {wk} ({monday.strftime('%Y-%m-%d')}–{sunday.strftime('%Y-%m-%d')}){cyc_txt}", ""]

    # ---- 1. Executive summary ----
    dl_txt = ""
    if kpi["deadline"]:
        dl_txt = f", deadline {kpi['deadline']}" + (f" ({kpi['days_remaining']} days left)" if kpi["days_remaining"] is not None else "")
    pass_line = (f"- Pass rate: **{_pct(first['pass_rate'])} → {_pct(last['pass_rate'])}** this week (target ≥{_pct(kpi['target_rate'], 0)}{dl_txt})"
                 if in_week else f"- Pass rate: _no cycle ran this week_ (target ≥{_pct(kpi['target_rate'], 0)}{dl_txt})")
    md += ["## 1. Executive summary", "",
           pass_line,
           f"- Logged **{n_fix_week} fixes** and **{n_new_week} new scripts (DONE)** this week",
           f"- Still failing: **{kpi['still_failing']}/{kpi['total_scripts']}** · P0: {kpi['tier_counts']['P0']} · Flaky: {kpi['flaky_count']}",
           "- _(One-line assessment: on track / behind & why — fill manually)_", ""]

    # ---- 2. Key metrics vs target ----
    n_fix_prev = db.execute("SELECT COUNT(*) c FROM fixes WHERE fix_date>=? AND fix_date<=?",
                            (prev_monday.isoformat(), prev_sunday.isoformat())).fetchone()["c"]
    days_week = max(1, len(in_week))
    md += ["## 2. Key metrics vs target", "",
           "| Metric | Prev week | This week | Target |", "|---|---|---|---|",
           f"| Pass rate (last cycle of week) | {_pct(prev_last['pass_rate']) if prev_last else '…'} | {_pct(last['pass_rate']) if last else '…'} | ≥{_pct(kpi['target_rate'], 0)} |",
           f"| Failures (last cycle of week) | {prev_last['fail_count'] if prev_last else '…'} | {last['fail_count'] if last else '…'} | ↓ |",
           f"| Still failing (current) | | {kpi['still_failing']} | ↓ |",
           f"| Scripts left to fix for target | | {kpi['fails_to_fix']} | 0 |",
           (f"| Required fix velocity | | {kpi['required_rate_per_day']:.1f}/day | |" if kpi["required_rate_per_day"] is not None else "| Required fix velocity | | … (no deadline set) | |"),
           f"| Actual fix velocity | {n_fix_prev / 7:.1f}/day | {n_fix_week / days_week:.1f}/day | ≥ required |",
           f"| P0 / P1 | | {kpi['tier_counts']['P0']} / {kpi['tier_counts']['P1']} | P0 = 0 |",
           (f"| New-script progress | | {coverage['done']}/{coverage['total_needed']} ({_pct(coverage['pct'])}) | on plan |" if coverage.get("configured") else "| New-script progress | | company system not synced | |"),
           ""]

    # ---- 3. Pass rate Item x Model (>=3 cycles) ----
    if trend:
        week_cycles = [t["cycle"] for t in in_week]
        if len(week_cycles) < 3:
            ref = last["cycle"] if last else trend[-1]["cycle"]
            week_cycles = [t["cycle"] for t in trend if t["cycle"] <= ref][-3:]
        md += [f"## 3. Pass rate Item × Model ({len(week_cycles)} cycles)", "",
               _suite_model_report_table(db, week_cycles), ""]
    else:
        week_cycles = []
        md += ["## 3. Pass rate Item × Model", "", "_No result data yet._", ""]

    # ---- 4. Environment-related failures ----
    md += ["## 4. Environment-related failures (Infra/Device)", ""]
    tot_fail = tot_env = 0
    if week_cycles:
        env = _env_fail_by_cycle(db, week_cycles)
        tot_fail = sum(d["fail"] for d in env.values())
        tot_env = sum(d["env"] for d in env.values())
        md += ["| Cycle | Total fail | Env fail | Rate |", "|---|---|---|---|"]
        for c in week_cycles:
            d = env[c]
            md.append(f"| C{c} | {d['fail']} | {d['env']} | {_pct(d['rate'])} |")
        md.append(f"| **Whole week** | **{tot_fail}** | **{tot_env}** | **{_pct((tot_env / tot_fail) if tot_fail else None)}** |")
    else:
        md.append("_No cycle data._")
    md.append("")

    # ---- 5. By app / test suite ----
    suite_rows = {}
    for p in priority:
        s = suite_rows.setdefault(p["test_suite"], {"total": 0, "done": 0, "verify": 0, "fail": 0})
        s["total"] += 1
        if p["priority_tier"] == "Done":
            s["done"] += 1
        elif p["priority_tier"] == "Verify":
            s["verify"] += 1
        else:
            s["fail"] += 1
    md += ["## 5. By app / test suite", "",
           "| App/Suite | Total scripts | Done | Verifying | Still failing | Done % |", "|---|---|---|---|---|---|"]
    t_tot = t_done = t_ver = t_fail = 0
    for name in sorted(suite_rows, key=lambda k: suite_rows[k]["done"] / suite_rows[k]["total"] if suite_rows[k]["total"] else 0):
        s = suite_rows[name]
        t_tot += s["total"]; t_done += s["done"]; t_ver += s["verify"]; t_fail += s["fail"]
        md.append(f"| {name} | {s['total']} | {s['done']} | {s['verify']} | {s['fail']} | {_pct(s['done'] / s['total'] if s['total'] else None)} |")
    md.append(f"| **Total** | **{t_tot}** | **{t_done}** | **{t_ver}** | **{t_fail}** | **{_pct(t_done / t_tot if t_tot else None)}** |")
    md.append("")

    # ---- 6. Script completion by Test Suite (company system) ----
    md += ["## 6. Script completion by Test Suite", "", _completion_table_md(coverage), ""]

    # ---- 7. Root cause Pareto (from fix log, top 5) ----
    fix_pareto = compute_fix_root_cause_pareto(db, mon_s, sun_s)
    md += ["## 7. Root cause Pareto this week (from fix log, top 5)", ""]
    if fix_pareto:
        md += ["| Root cause group | Fixes | % | Next-week action |", "|---|---|---|---|"]
        for g in fix_pareto[:5]:
            md.append(f"| {g['group']} | {g['count']} | {_pct(g['pct'])} | _(fill manually)_ |")
    else:
        md.append("_No fix recorded a root-cause group this week._")
    md.append("")

    # ---- 8. Output per person this week ----
    n_fix, n_fixers, fix_table = _person_count_table(db, "fixes", "owner", "fix_date", mon_s, sun_s)
    n_new, n_writers, new_table = _person_count_table(db, "new_scripts", "member", "completed_date", mon_s, sun_s, "AND status='DONE'")
    md += ["## 8. Output this week", "",
           f"**Fixes: {n_fix}** (by {n_fixers} people)", "", fix_table, "",
           f"**New scripts DONE: {n_new}** (by {n_writers} people)", "", new_table, ""]

    # ---- 9. Fix quality & team ----
    tracking = compute_fix_tracking(db)
    n_verified = sum(1 for f in tracking if f["status"] == "verified")
    n_regressed = sum(1 for f in tracking if f["status"] == "regressed")
    reopen_rate = (n_regressed / (n_verified + n_regressed)) if (n_verified + n_regressed) else None
    owner_stats = get_owner_stats(db, priority)
    overloaded = [o for o in owner_stats if o["open_workload"] > 10]
    md += ["## 9. Fix quality & team", "",
           f"- Reopen rate (fixed then failed again): **{_pct(reopen_rate)}** (target ≤10%) — ✅ {n_verified} verified / ⚠️ {n_regressed} regressed",
           ]
    if owner_stats:
        best = owner_stats[0]
        md.append(f"- Top resolution owner: **{best['owner']}** ({_pct(best['resolution_rate'])} resolution, {best['distinct_scripts_fully_resolved']} scripts fully resolved)")
    if overloaded:
        md.append("- Overloaded (Open Workload > 10): " + ", ".join(f"{o['owner']} ({o['open_workload']})" for o in overloaded))
    md.append("")

    # ---- 10. Project completion estimate ----
    md += ["## 10. Project completion estimate", "", _eta_section(db, kpi, coverage), ""]

    # ---- 11. Farm (fill manually) + risks ----
    md += ["## 11. Test farm this week (fill manually)", "",
           "| Metric | Value | Prev week |", "|---|---|---|",
           "| Avg device availability | …% | |",
           "| Full cycles run / planned | …/… | |",
           f"| % failures from Infra/Device | {_pct((tot_env / tot_fail) if tot_fail else None)} | |",
           "| Major incidents + recovery time | | |", "",
           "## 12. Risks & recommendations", "", "- …", ""]

    return "\n".join(md), None


@app.route("/api/report/daily")
@require_login
def api_report_daily():
    db = get_db()
    date_str = (request.args.get("date") or "").strip() or None
    md, err = build_daily_report(db, date_str)
    if err:
        return jsonify({"error": err}), 404
    return jsonify({"markdown": md, "meta": {"type": "daily", "date": date_str}})


@app.route("/api/report/weekly")
@require_login
def api_report_weekly():
    db = get_db()
    week_str = (request.args.get("week") or "").strip() or None
    md, err = build_weekly_report(db, week_str)
    if err:
        return jsonify({"error": err}), 404
    return jsonify({"markdown": md, "meta": {"type": "weekly", "week": week_str}})


def _excel_style_header(sheet, row, values, color="4472C4"):
    """Helper: style header row (bold, white text, colored background)."""
    for col_idx, val in enumerate(values, 1):
        cell = sheet.cell(row=row, column=col_idx, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _excel_style_data_row(sheet, row, values):
    """Helper: style data row with light background. Float values (VD ty le 0..1)
    duoc format tu dong thanh "0.0%" (numeric that, khong phai chuoi) de conditional
    formatting / color-scale hoat dong duoc tren cac cot ty le."""
    for col_idx, val in enumerate(values, 1):
        cell = sheet.cell(row=row, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if isinstance(val, float):
            cell.number_format = "0.0%"
        if row % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


_TIER_COLORS = [("P0", "E74C3C"), ("P1", "E67E22"), ("P2", "F1C40F"),
                 ("P3", "3498DB"), ("Verify", "9B59B6"), ("Done", "2ECC71")]


def _excel_overview_sheet(wb, title_text, db, trend, kpi, coverage):
    """Sheet 'Overview' — trang tong quan DUY NHAT dat dau file: health status mau,
    vai KPI lon, 1-2 nhan dinh ngan, + 2 chart (pass rate trend toan bo lich su,
    phan bo tier). Muc dich: doc trong ~10 giay nam duoc tinh hinh du an, cac sheet
    chi tiet phia sau danh cho ai can dao sau. Dung chung cho ca daily & weekly."""
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value=title_text).font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    current_rate = trend[-1]["pass_rate"] if trend else None
    target_rate = kpi["target_rate"]

    # ---- Health status (traffic light) ----
    if current_rate is None:
        status_text, status_color = "NO DATA", "BDC3C7"
    else:
        diff = current_rate - target_rate
        if diff >= 0:
            status_text, status_color = "ON TRACK ✅", "2ECC71"
        elif diff >= -0.05:
            status_text, status_color = "AT RISK ⚠️", "F1C40F"
        else:
            status_text, status_color = "BEHIND ❌", "E74C3C"
    ws.merge_cells("A3:F4")
    cell = ws["A3"]
    cell.value = status_text
    cell.font = Font(bold=True, size=20, color="FFFFFF")
    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Big KPI cells ----
    fix_speed = _velocity_last7(db, "fixes", "fix_date")
    if kpi["days_remaining"] is not None:
        days_label, days_val = "Days to Deadline", kpi["days_remaining"]
    elif kpi["fails_to_fix"] is not None and kpi["fails_to_fix"] > 0 and fix_speed > 0:
        days_label = "Est. Days to Target"
        days_val = int(kpi["fails_to_fix"] / fix_speed + 0.999)
    else:
        days_label, days_val = "Days to Deadline", "—"

    kpis = [
        ("Current Pass Rate", current_rate if current_rate is not None else "—"),
        ("Target Pass Rate", target_rate),
        ("Still Failing", f"{kpi['still_failing']} / {kpi['total_scripts']}"),
        ("P0 (wide-impact)", kpi["tier_counts"]["P0"]),
        (days_label, days_val),
    ]
    r = 6
    for i, (label, val) in enumerate(kpis):
        col = 1 + i
        c1 = ws.cell(row=r, column=col, value=label)
        c1.font = Font(size=10, color="666666")
        c1.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=r + 1, column=col, value=val)
        c2.font = Font(size=18, bold=True)
        c2.alignment = Alignment(horizontal="center")
        if isinstance(val, float):
            c2.number_format = "0.0%"
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ---- Short insights (1-2 lines) ----
    r = 9
    insight_lines = []
    if trend and len(trend) >= 2:
        last, prev = trend[-1], trend[-2]
        if last["delta_fail"] is not None and last["delta_fail"] != 0:
            direction = "down" if last["delta_fail"] < 0 else "UP"
            insight_lines.append(f"• Cycle {last['cycle']}: failures {direction} {abs(last['delta_fail'])} vs previous cycle.")
    if kpi["required_rate_per_day"] is not None:
        verdict = "meeting" if fix_speed >= kpi["required_rate_per_day"] else "BELOW"
        insight_lines.append(f"• Fix velocity (7-day avg {fix_speed:.1f}/day) is {verdict} the {kpi['required_rate_per_day']:.1f}/day required to hit the deadline.")
    if coverage.get("configured"):
        insight_lines.append(f"• Script coverage: {coverage['done']}/{coverage['total_needed']} ({coverage['pct']*100:.1f}%) of required test cases done.")
    if not insight_lines:
        insight_lines = ["• Not enough data yet for a trend assessment."]
    for line in insight_lines:
        ws.cell(row=r, column=1, value=line).font = Font(size=11)
        ws.merge_cells(f"A{r}:F{r}")
        r += 1

    # ---- Chart 1: Pass rate trend (full history) ----
    chart_top_row = r + 2
    ws.cell(row=chart_top_row - 1, column=1, value="Chart data (do not edit)").font = Font(size=9, italic=True, color="999999")
    ws.cell(row=chart_top_row, column=1, value="Cycle")
    ws.cell(row=chart_top_row, column=2, value="Pass Rate")
    trow = chart_top_row
    for t in trend:
        trow += 1
        ws.cell(row=trow, column=1, value=f"C{t['cycle']}")
        pc = ws.cell(row=trow, column=2, value=t["pass_rate"])
        if t["pass_rate"] is not None:
            pc.number_format = "0.0%"
    if trend:
        line = LineChart()
        line.title = "Pass Rate Trend (all cycles)"
        line.y_axis.numFmt = "0%"
        line.y_axis.title = "Pass Rate"
        line.x_axis.title = "Cycle"
        line.height, line.width = 8, 16
        data_ref = Reference(ws, min_col=2, min_row=chart_top_row, max_row=trow)
        cats_ref = Reference(ws, min_col=1, min_row=chart_top_row + 1, max_row=trow)
        line.add_data(data_ref, titles_from_data=True)
        line.set_categories(cats_ref)
        ws.add_chart(line, f"H{r}")

    # ---- Chart 2: Tier distribution ----
    tier_start_row = trow + 3
    ws.cell(row=tier_start_row, column=1, value="Tier")
    ws.cell(row=tier_start_row, column=2, value="Count")
    for i, (tier, _color) in enumerate(_TIER_COLORS):
        ws.cell(row=tier_start_row + 1 + i, column=1, value=tier)
        ws.cell(row=tier_start_row + 1 + i, column=2, value=kpi["tier_counts"].get(tier, 0))
    pie = PieChart()
    pie.title = "Priority Tier Distribution"
    pie.height, pie.width = 8, 12
    data_ref = Reference(ws, min_col=2, min_row=tier_start_row, max_row=tier_start_row + len(_TIER_COLORS))
    cats_ref = Reference(ws, min_col=1, min_row=tier_start_row + 1, max_row=tier_start_row + len(_TIER_COLORS))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    try:
        from openpyxl.chart.marker import DataPoint
        pie.series[0].data_points = [
            DataPoint(idx=i, spPr=None) for i in range(len(_TIER_COLORS))
        ]
        for dp, (_tier, color) in zip(pie.series[0].data_points, _TIER_COLORS):
            dp.graphicalProperties.solidFill = color
    except Exception:
        pass  # mau slice mac dinh neu API chart noi bo openpyxl thay doi giua cac phien ban
    ws.add_chart(pie, f"H{r + 18}")

    return ws


def _excel_pass_rate_matrix_sheet(wb, db, cycles):
    """Sheet 'Pass Rate Matrix': Item x Model x cycle, kem dong Total tung Item + GRAND TOTAL.
    Aggregation logic khop 1:1 voi _suite_model_report_table() (markdown)."""
    ws = wb.create_sheet("Pass Rate Matrix")
    data = compute_suite_model_matrix(db)
    header = ["Item", "Model"] + [f"C{c}" for c in cycles] + ["Δ"]
    _excel_style_header(ws, 1, header, "70AD47")
    r = 2
    first_data_row = r
    if not cycles:
        _excel_style_data_row(ws, r, ["No result data yet."] + [""] * (len(header) - 1))
        r += 1
    else:
        def fmt(cell):
            # Gia tri numeric that (float 0..1) de ColorScaleRule hoat dong; "—" khi khong co du lieu.
            return cell["pass_rate"] if cell and cell["pass_rate"] is not None else "—"

        def delta(first, last):
            if first is None or last is None:
                return "…"
            d = (last - first) * 100
            arrow = "▲" if d > 0 else ("▼" if d < 0 else "=")
            return f"{arrow}{abs(d):.1f}%"

        by_item = {}
        for row in data["rows"]:
            by_item.setdefault(row["test_suite"], []).append(row)

        for item in sorted(by_item):
            rows = by_item[item]
            for row in sorted(rows, key=lambda x: x["model"]):
                cells = [row["by_cycle"].get(c) for c in cycles]
                rates = [c["pass_rate"] if c else None for c in cells]
                first = next((x for x in rates if x is not None), None)
                last = next((x for x in reversed(rates) if x is not None), None)
                _excel_style_data_row(ws, r, [item, row["model"]] + [fmt(c) for c in cells] + [delta(first, last)])
                r += 1
            agg_cells = []
            for c in cycles:
                tot = pas = na = 0
                found = False
                for row in rows:
                    cell = row["by_cycle"].get(c)
                    if cell:
                        found = True
                        tot += cell["total"]; pas += cell["pass_count"]; na += cell["na_count"]
                denom = tot - na
                agg_cells.append({"pass_rate": (pas / denom) if (found and denom > 0) else None} if found else None)
            rates = [c["pass_rate"] if c else None for c in agg_cells]
            first = next((x for x in rates if x is not None), None)
            last = next((x for x in reversed(rates) if x is not None), None)
            _excel_style_data_row(ws, r, [item, "Total"] + [fmt(c) for c in agg_cells] + [delta(first, last)])
            r += 1

        overall_cells = [data["overall_by_cycle"].get(c) for c in cycles]
        rates = [c["pass_rate"] if c else None for c in overall_cells]
        first = next((x for x in rates if x is not None), None)
        last = next((x for x in reversed(rates) if x is not None), None)
        _excel_style_data_row(ws, r, ["GRAND TOTAL", "All"] + [fmt(c) for c in overall_cells] + [delta(first, last)])

        # Conditional formatting: to do-vang-xanh theo pass rate tren toan bo cot cycle (khong tinh Item/Model/Delta).
        last_row = r
        rate_range = f"C{first_data_row}:{get_column_letter(2 + len(cycles))}{last_row}"
        ws.conditional_formatting.add(rate_range, ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ))

    ws.column_dimensions["A"].width = 22
    for col in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    return ws


def _excel_env_failures_sheet(wb, db, cycles):
    """Sheet 'Env Failures': loi lien quan moi truong (Infra/Device) tung cycle + Total row."""
    ws = wb.create_sheet("Env Failures")
    _excel_style_header(ws, 1, ["Cycle", "Total Fail", "Env Fail", "Rate"])
    r = 2
    if not cycles:
        _excel_style_data_row(ws, r, ["No cycle data.", "", "", ""])
    else:
        env = _env_fail_by_cycle(db, cycles)
        tot_fail = tot_env = 0
        for c in cycles:
            d = env[c]
            tot_fail += d["fail"]; tot_env += d["env"]
            _excel_style_data_row(ws, r, [f"C{c}", d["fail"], d["env"], f"{d['rate']*100:.1f}%" if d["rate"] is not None else "…"])
            r += 1
        _excel_style_data_row(ws, r, ["Total", tot_fail, tot_env, f"{(tot_env/tot_fail*100):.1f}%" if tot_fail else "…"])
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 16
    return ws


def _excel_productivity_sheet(wb, db, date_from, date_to):
    """Sheet 'Productivity': fix & script viet moi tung nguoi trong khoang [date_from, date_to] + Total row moi bang."""
    ws = wb.create_sheet("Productivity")
    r = 1
    ws.cell(row=r, column=1, value="Fixes per person").font = Font(bold=True, size=12)
    r += 1
    _excel_style_header(ws, r, ["Person", "Fixes"])
    r += 1
    fx_rows = db.execute(
        "SELECT COALESCE(NULLIF(TRIM(owner), ''), '(unknown)') person, COUNT(*) c FROM fixes "
        "WHERE fix_date>=? AND fix_date<=? GROUP BY person ORDER BY c DESC, person",
        (date_from, date_to),
    ).fetchall()
    for x in fx_rows:
        _excel_style_data_row(ws, r, [x["person"], x["c"]])
        r += 1
    _excel_style_data_row(ws, r, ["Total", sum(x["c"] for x in fx_rows)])
    r += 2

    ws.cell(row=r, column=1, value="New scripts (DONE) per person").font = Font(bold=True, size=12)
    r += 1
    _excel_style_header(ws, r, ["Person", "Scripts Written"])
    r += 1
    ws_rows = db.execute(
        "SELECT COALESCE(NULLIF(TRIM(member), ''), '(unknown)') person, COUNT(*) c FROM new_scripts "
        "WHERE status='DONE' AND completed_date>=? AND completed_date<=? GROUP BY person ORDER BY c DESC, person",
        (date_from, date_to),
    ).fetchall()
    for x in ws_rows:
        _excel_style_data_row(ws, r, [x["person"], x["c"]])
        r += 1
    _excel_style_data_row(ws, r, ["Total", sum(x["c"] for x in ws_rows)])

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    return ws


def _excel_completion_sheet(wb, coverage):
    """Sheet 'Completion by Suite': nguon CHI TU he thong cong ty (compute_coverage). Chua sync -> ghi ro."""
    ws = wb.create_sheet("Completion by Suite")
    if not coverage.get("configured"):
        ws.cell(row=1, column=1, value="Company system not synced — total required TCs unknown.").font = Font(italic=True)
        ws.column_dimensions["A"].width = 60
        return ws
    _excel_style_header(ws, 1, ["Item", "Total TC", "Completed Script", "Completed Rate"], "FFC000")
    r = 2
    for it in coverage["by_item"]:
        _excel_style_data_row(ws, r, [it["item"], it["needed"], it["done"], it["pct"]])
        r += 1
    _excel_style_data_row(ws, r, ["Total", coverage["total_needed"], coverage["done"], coverage["pct"]])
    ws.conditional_formatting.add(f"D2:D{r}", ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    ))
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20
    return ws


def _excel_eta_plan_sheet(wb, db, kpi, coverage, priority):
    """Sheet 'ETA & Plan': uoc luong hoan thanh (tai su dung _eta_section) + top P0/P1 can lam truoc."""
    ws = wb.create_sheet("ETA & Plan")
    r = 1
    ws.cell(row=r, column=1, value="Project completion estimate").font = Font(bold=True, size=12)
    r += 1
    for line in _eta_section(db, kpi, coverage).split("\n"):
        ws.cell(row=r, column=1, value=line.lstrip("- ").replace("**", ""))
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Top P0/P1 scripts").font = Font(bold=True, size=12)
    r += 1
    _excel_style_header(ws, r, ["Tier", "Test Suite", "Test Case", "Owner"])
    r += 1
    top = sorted(
        (p for p in priority if p["priority_tier"] in ("P0", "P1")),
        key=lambda p: (-p["priority_score"], p["test_case"]),
    )
    for p in top[:30]:
        _excel_style_data_row(ws, r, [p["priority_tier"], p["test_suite"], p["test_case"], p["current_owner"] or "(unassigned)"])
        r += 1
    ws.column_dimensions["A"].width = 55
    for col in "BCD":
        ws.column_dimensions[col].width = 22
    return ws


def _excel_root_cause_pareto_sheet(wb, fix_pareto):
    """Sheet 'Root Cause Pareto' (weekly only): tu fix log, group theo root_cause_group."""
    ws = wb.create_sheet("Root Cause Pareto")
    _excel_style_header(ws, 1, ["Root Cause Group", "Fixes", "%"])
    r = 2
    if not fix_pareto:
        ws.cell(row=r, column=1, value="No fix recorded a root-cause group this week.").font = Font(italic=True)
    else:
        for g in fix_pareto:
            _excel_style_data_row(ws, r, [g["group"], g["count"], f"{g['pct']*100:.1f}%"])
            r += 1
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 26
    return ws


def _excel_team_quality_sheet(wb, owner_stats, reopen_rate, n_verified, n_regressed):
    """Sheet 'Team Quality' (weekly only): mirror Owner Leaderboard (all-time rates) + Total row."""
    ws = wb.create_sheet("Team Quality")
    r = 1
    ws.cell(row=r, column=1,
            value=f"Reopen rate: {reopen_rate*100:.1f}%  (Verified: {n_verified}  Regressed: {n_regressed})"
            if reopen_rate is not None else f"Reopen rate: …  (Verified: {n_verified}  Regressed: {n_regressed})"
            ).font = Font(bold=True)
    r += 2
    header = ["#", "Owner", "Scripts Written", "Fixes", "Distinct Fixed", "Fully Resolved",
              "Resolution Rate", "Verified", "Reopen", "Verification Rate", "Open Workload"]
    _excel_style_header(ws, r, header)
    r += 1
    first_data_row = r
    for o in owner_stats:
        _excel_style_data_row(ws, r, [
            o["rank"], o["owner"], o.get("scripts_written", 0), o["fixes_logged"],
            o["distinct_scripts_fixed"], o["distinct_scripts_fully_resolved"],
            o["resolution_rate"] if o["resolution_rate"] is not None else "…",
            o["verified"], o["reopened"],
            o["verification_rate"] if o["verification_rate"] is not None else "…",
            o["open_workload"],
        ])
        r += 1
    def sm(k):
        return sum(o.get(k, 0) for o in owner_stats)
    _excel_style_data_row(ws, r, ["", "Total", sm("scripts_written"), sm("fixes_logged"), sm("distinct_scripts_fixed"),
                                   sm("distinct_scripts_fully_resolved"), "—", sm("verified"), sm("reopened"), "—",
                                   sm("open_workload")])
    if owner_stats:
        for col_letter in ("G", "J"):  # Resolution Rate, Verification Rate
            ws.conditional_formatting.add(f"{col_letter}{first_data_row}:{col_letter}{r - 1}", ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ))
    for col in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    return ws


def _build_daily_report_excel(db, date_str):
    """Excel DAILY report: nhieu sheet, English, dong Total, dat parity voi build_daily_report()
    (markdown). Cycle KHONG bat buoc (giong markdown) - dung chung logic loi/relax."""
    trend = compute_cycle_trend(db)
    if not date_str:
        date_str = trend[-1]["cycle_date"] if trend else date.today().isoformat()
    md, err = build_daily_report(db, date_str)
    if err:
        raise ValueError(err)

    entry = next((t for t in trend if t["cycle_date"] == date_str), None)
    idx = trend.index(entry) if entry is not None else None
    prev = trend[idx - 1] if (idx is not None and idx > 0) else None
    priority = get_script_priority(db)
    kpi = _report_kpis(db, priority)
    coverage = compute_coverage(db)

    wb = Workbook()
    title_cycle = f"Cycle {entry['cycle']}" if entry is not None else "No cycle"
    _excel_overview_sheet(wb, f"Stabilization Daily Report — {title_cycle} ({date_str})", db, trend, kpi, coverage)

    ws = wb.create_sheet("Summary")
    r = 1
    ws.cell(row=r, column=1, value=f"Stabilization Daily Report — {title_cycle} ({date_str})").font = Font(bold=True, size=14)
    r += 2
    _excel_style_header(ws, r, ["Metric", "Today", "Prev Cycle"])
    r += 1
    if entry is not None:
        rows = [
            ("Pass rate", f"{entry['pass_rate']*100:.1f}%", f"{prev['pass_rate']*100:.1f}%" if prev else "…"),
            ("Total runs", entry["total"], prev["total"] if prev else "…"),
            ("Failures in cycle", entry["fail_count"], prev["fail_count"] if prev else "…"),
        ]
    else:
        rows = [(f"No cycle ran on {date_str}", "", "")]
    for lbl, cur, prv in rows:
        _excel_style_data_row(ws, r, [lbl, cur, prv])
        r += 1
    for lbl, val in [
        ("Still failing", f"{kpi['still_failing']}/{kpi['total_scripts']}"),
        ("P0", kpi["tier_counts"]["P0"]),
        ("Verifying (Verify)", kpi["tier_counts"]["Verify"]),
        ("Flaky scripts", kpi["flaky_count"]),
    ]:
        _excel_style_data_row(ws, r, [lbl, val, ""])
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    if trend:
        end = (idx + 1) if idx is not None else len(trend)
        recent_cycles = [t["cycle"] for t in trend[:end]][-3:]
    else:
        recent_cycles = []
    _excel_pass_rate_matrix_sheet(wb, db, recent_cycles)
    _excel_env_failures_sheet(wb, db, recent_cycles)
    _excel_productivity_sheet(wb, db, date_str, date_str)
    _excel_completion_sheet(wb, coverage)
    _excel_eta_plan_sheet(wb, db, kpi, coverage, priority)
    return wb


def _build_weekly_report_excel(db, week_str):
    """Excel WEEKLY report: nhieu sheet, English, dong Total, dat parity voi build_weekly_report()
    (markdown) + 2 sheet rieng cho tuan (Root Cause Pareto, Team Quality)."""
    if not week_str:
        iso = date.today().isocalendar()
        week_str = f"{iso[0]}-W{iso[1]:02d}"
    md, err = build_weekly_report(db, week_str)
    if err:
        raise ValueError(err)

    m = re.match(r"^(\d{4})-W(\d{1,2})$", week_str.strip())
    year, wk = int(m.group(1)), int(m.group(2))
    monday = date.fromisocalendar(year, wk, 1)
    sunday = monday + timedelta(days=6)
    mon_s, sun_s = monday.isoformat(), sunday.isoformat()

    trend = compute_cycle_trend(db)
    in_week = [t for t in trend if t["cycle_date"] and mon_s <= t["cycle_date"] <= sun_s]
    priority = get_script_priority(db)
    kpi = _report_kpis(db, priority)
    coverage = compute_coverage(db)

    wb = Workbook()
    cyc_txt = f"Cycle {in_week[0]['cycle']}–{in_week[-1]['cycle']}" if in_week else "No cycle"
    title_text = f"Stabilization Weekly Report — Week {wk} ({monday.strftime('%Y-%m-%d')}–{sunday.strftime('%Y-%m-%d')}) — {cyc_txt}"
    _excel_overview_sheet(wb, title_text, db, trend, kpi, coverage)

    ws = wb.create_sheet("Summary")
    r = 1
    ws.cell(row=r, column=1, value=title_text).font = Font(bold=True, size=14)
    r += 2
    n_fix = db.execute("SELECT COUNT(*) c FROM fixes WHERE fix_date>=? AND fix_date<=?", (mon_s, sun_s)).fetchone()["c"]
    n_new = db.execute("SELECT COUNT(*) c FROM new_scripts WHERE status='DONE' AND completed_date>=? AND completed_date<=?",
                      (mon_s, sun_s)).fetchone()["c"]
    _excel_style_header(ws, r, ["Metric", "Value"])
    r += 1
    data = [
        ("Pass rate start→end of week",
         f"{in_week[0]['pass_rate']*100:.1f}% → {in_week[-1]['pass_rate']*100:.1f}%" if in_week else "No cycle this week"),
        ("Fixes this week", n_fix),
        ("New scripts (DONE)", n_new),
        ("Still failing", f"{kpi['still_failing']}/{kpi['total_scripts']}"),
        ("Completion (company system)",
         f"{coverage['done']}/{coverage['total_needed']} ({coverage['pct']*100:.1f}%)" if coverage.get("configured") else "Not synced"),
        ("P0", kpi["tier_counts"]["P0"]),
        ("Verifying (Verify)", kpi["tier_counts"]["Verify"]),
        ("Flaky scripts", kpi["flaky_count"]),
    ]
    for lbl, val in data:
        _excel_style_data_row(ws, r, [lbl, val])
        r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28

    if trend:
        week_cycles = [t["cycle"] for t in in_week]
        if len(week_cycles) < 3:
            ref = in_week[-1]["cycle"] if in_week else trend[-1]["cycle"]
            week_cycles = [t["cycle"] for t in trend if t["cycle"] <= ref][-3:]
    else:
        week_cycles = []
    _excel_pass_rate_matrix_sheet(wb, db, week_cycles)
    _excel_env_failures_sheet(wb, db, week_cycles)
    _excel_productivity_sheet(wb, db, mon_s, sun_s)
    _excel_completion_sheet(wb, coverage)

    fix_pareto = compute_fix_root_cause_pareto(db, mon_s, sun_s)
    _excel_root_cause_pareto_sheet(wb, fix_pareto)

    tracking = compute_fix_tracking(db)
    n_verified = sum(1 for f in tracking if f["status"] == "verified")
    n_regressed = sum(1 for f in tracking if f["status"] == "regressed")
    reopen_rate = (n_regressed / (n_verified + n_regressed)) if (n_verified + n_regressed) else None
    owner_stats = get_owner_stats(db, priority)
    _excel_team_quality_sheet(wb, owner_stats, reopen_rate, n_verified, n_regressed)

    _excel_eta_plan_sheet(wb, db, kpi, coverage, priority)
    return wb


@app.route("/api/report/daily/export")
@require_login
def api_report_daily_export():
    """Export daily report as Excel file."""
    db = get_db()
    date_str = (request.args.get("date") or "").strip() or None
    try:
        wb = _build_daily_report_excel(db, date_str)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"daily_cycle_{date_str or 'latest'}.xlsx")


@app.route("/api/report/weekly/export")
@require_login
def api_report_weekly_export():
    """Export weekly report as Excel file."""
    db = get_db()
    week_str = (request.args.get("week") or "").strip() or None
    try:
        wb = _build_weekly_report_excel(db, week_str)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"weekly_report_{week_str or 'current'}.xlsx")


# ==================================================================
# BACKUP: snapshot hang ngay tracker.db + users.db vao backups/YYYY-MM-DD/
# (sqlite3 backup API - an toan voi WAL; daemon thread, khong can cron ngoai)
# ==================================================================
_backup_thread_started = False


def perform_backup(reason="auto"):
    """Snapshot 2 file DB vao backups/<today>/. Dung sqlite3 backup API (WAL-safe).
    Chay duoc ca ngoai request context (daemon thread) - tu mo connection rieng.
    Tra ve dict tom tat."""
    today = date.today().isoformat()
    target_dir = os.path.join(BACKUP_DIR, today)
    os.makedirs(target_dir, exist_ok=True)
    files = []
    for src_path in (DB_PATH, USERS_DB_PATH):
        if not os.path.exists(src_path):
            continue
        dst_path = os.path.join(target_dir, os.path.basename(src_path))
        src = sqlite3.connect(src_path)
        dst = sqlite3.connect(dst_path)
        try:
            with dst:
                src.backup(dst)
        finally:
            dst.close()
            src.close()
        files.append({"file": os.path.basename(src_path), "size": os.path.getsize(dst_path)})

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute("SELECT value FROM settings WHERE key='backup_retention'").fetchone()
        try:
            retention = max(1, int(row["value"])) if row and row["value"] else 30
        except (ValueError, TypeError):
            retention = 30
        deleted = []
        if os.path.isdir(BACKUP_DIR):
            dirs = sorted(d for d in os.listdir(BACKUP_DIR)
                          if re.match(r"^\d{4}-\d{2}-\d{2}$", d) and os.path.isdir(os.path.join(BACKUP_DIR, d)))
            while len(dirs) > retention:
                oldest = dirs.pop(0)
                shutil.rmtree(os.path.join(BACKUP_DIR, oldest), ignore_errors=True)
                deleted.append(oldest)
        conn.execute(
            "INSERT INTO settings (key, value) VALUES ('last_backup_date', ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value", (today,))
        conn.execute(
            "INSERT INTO audit_log (username, action, target, detail) VALUES (?,?,?,?)",
            ("(system)", "backup.run", today, f"reason={reason}, files={len(files)}, deleted_old={len(deleted)}"))
        conn.commit()
    finally:
        conn.close()
    return {"backup_dir": target_dir, "files": files, "deleted_old": deleted, "kept": retention}


def backup_daemon():
    """Vong lap nen: moi 30 phut kiem tra hom nay da backup chua (theo setting
    last_backup_date); chua thi backup. Khong bao gio de exception giet thread."""
    while True:
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            try:
                vals = {r["key"]: r["value"] for r in conn.execute(
                    "SELECT key, value FROM settings WHERE key IN ('backup_enabled','last_backup_date')")}
            finally:
                conn.close()
            enabled = (vals.get("backup_enabled") or "1") != "0"
            last = vals.get("last_backup_date") or ""
            if enabled and last < date.today().isoformat():
                perform_backup("auto")
                print(f"[backup] Da backup tu dong vao {BACKUP_DIR}/{date.today().isoformat()}")
        except Exception as e:
            print(f"[backup] Loi backup tu dong: {e}")
        time.sleep(1800)


def start_backup_daemon():
    global _backup_thread_started
    if _backup_thread_started:
        return
    _backup_thread_started = True
    threading.Thread(target=backup_daemon, daemon=True).start()


@app.route("/api/backup/run", methods=["POST"])
@require_perm("settings")
def api_backup_run():
    try:
        out = perform_backup("manual")
    except Exception as e:
        return jsonify({"error": f"Backup thất bại: {e}"}), 500
    return jsonify({"status": "ok", **out})


@app.route("/api/backup/status")
@require_login
def api_backup_status():
    db = get_db()
    backups = []
    if os.path.isdir(BACKUP_DIR):
        for d in sorted(os.listdir(BACKUP_DIR), reverse=True):
            full = os.path.join(BACKUP_DIR, d)
            if not (re.match(r"^\d{4}-\d{2}-\d{2}$", d) and os.path.isdir(full)):
                continue
            fs = [{"file": f, "size": os.path.getsize(os.path.join(full, f))} for f in sorted(os.listdir(full))]
            backups.append({"dir": d, "files": fs})
    return jsonify({
        "enabled": get_setting(db, "backup_enabled", "1") != "0",
        "retention": get_setting_int(db, "backup_retention", 30),
        "last_backup_date": get_setting(db, "last_backup_date", ""),
        "backups": backups,
    })


# ------------------------------------------------------------------
# Admin Dashboard - Management Panel
# ------------------------------------------------------------------
@app.route("/admin/<secret_key>")
def admin_dashboard(secret_key):
    if secret_key != ADMIN_SECRET_KEY:
        return "Unauthorized", 403
    return render_template("admin.html")


@app.route("/api/admin/results", methods=["GET"])
def admin_get_results():
    secret_key = request.args.get("key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    db = get_db()
    limit = int(request.args.get("limit", 1000))
    search = request.args.get("search", "").strip()

    if search:
        rows = db.execute(
            """SELECT * FROM results
               WHERE test_suite LIKE ? OR test_case LIKE ? OR model LIKE ? OR created_by LIKE ?
               ORDER BY id DESC LIMIT ?""",
            (f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", limit)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM results ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/results/<int:result_id>", methods=["PUT"])
def admin_update_result(result_id):
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json(force=True)
    db = get_db()

    # Build UPDATE query dynamically. Cycle (số) không cho sửa trực tiếp — nó được suy
    # tự động từ cycle_date; nếu sửa Test ID thì cycle_date sẽ tự cập nhật theo.
    allowed_fields = ["cycle_date", "test_id", "model", "test_suite", "test_case", "state", "description", "author", "team"]
    updates = []
    values = []

    for field in allowed_fields:
        if field in data:
            val = normalize_model_name(str(data[field]).strip()) if field == "model" else data[field]
            updates.append(f"{field}=?")
            values.append(val)

    # Nếu sửa Test ID và Test ID mã hoá ngày -> tự cập nhật cycle_date theo ngày đó.
    if "test_id" in data:
        d = extract_date_from_test_id(data["test_id"])
        if d:
            updates.append("cycle_date=?")
            values.append(d)

    # Recalculate result (Pass/Fail) if state changed
    if "state" in data:
        state = str(data["state"]).strip()
        result_class = classify_result(state)
        updates.append("result=?")
        values.append(result_class)

    if not updates:
        return jsonify({"error": "No fields to update"}), 400

    values.append(result_id)
    query = f"UPDATE results SET {', '.join(updates)} WHERE id=?"
    db.execute(query, values)
    recompute_cycles(db)  # danh so lai cycle theo ngay sau khi sua
    log_audit(db, "admin.result.update", target=f"results#{result_id}", detail=", ".join(sorted(k for k in data)))
    db.commit()

    return jsonify({"status": "updated"})


@app.route("/api/admin/results/<int:result_id>", methods=["DELETE"])
def admin_delete_result(result_id):
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    db = get_db()
    db.execute("DELETE FROM results WHERE id=?", (result_id,))
    recompute_cycles(db)  # danh so lai cycle theo ngay sau khi xoa (co the mat het 1 ngay)
    log_audit(db, "admin.result.delete", target=f"results#{result_id}")
    db.commit()

    return jsonify({"status": "deleted"})


# ------------------------------------------------------------------
# Admin: quan ly du lieu new_scripts (Script viet moi) - xem/sua/xoa tung dong +
# nhap hang loat (bulk). Cung mo hinh bao ve bang ADMIN_SECRET_KEY nhu Cycle Results.
# ------------------------------------------------------------------
@app.route("/api/admin/new-scripts", methods=["GET"])
def admin_get_new_scripts():
    secret_key = request.args.get("key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    db = get_db()
    limit = int(request.args.get("limit", 2000))
    search = request.args.get("search", "").strip()

    if search:
        rows = db.execute(
            """SELECT * FROM new_scripts
               WHERE tc_id LIKE ? OR item LIKE ? OR member LIKE ? OR team LIKE ?
               ORDER BY id DESC LIMIT ?""",
            (f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", limit)
        ).fetchall()
    else:
        rows = db.execute("SELECT * FROM new_scripts ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/new-scripts/<int:sid>", methods=["PUT"])
def admin_update_new_script(sid):
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json(force=True)
    db = get_db()

    # Sua TC ID thi tinh lai Item theo tien to moi (giu dung nguyen tac TC ID -> Item).
    updates = []
    values = []
    if "tc_id" in data:
        tc_id = extract_test_case_name(str(data["tc_id"] or "").strip())
        item = item_from_tc_id(tc_id)
        if not item:
            examples = ", ".join(ITEM_TC_EXAMPLES.values())
            return jsonify({"error": f"TC ID '{tc_id}' không đúng định dạng. Ví dụ hợp lệ: {examples}."}), 400
        dup = db.execute("SELECT 1 FROM new_scripts WHERE tc_id=? AND id!=? LIMIT 1", (tc_id, sid)).fetchone()
        if dup:
            return jsonify({"error": f"TC ID '{tc_id}' đã tồn tại ở dòng khác."}), 400
        updates += ["tc_id=?", "item=?"]
        values += [tc_id, item]

    for field in ("member", "team", "completed_date", "status", "models_written", "sdf_id", "remark"):
        if field in data:
            val = data[field]
            if field == "status":
                val = str(val or "").strip().upper()
                if val not in ("DONE", "SKIP", "ASSIGNED"):
                    return jsonify({"error": "Status phải là DONE, SKIP hoặc ASSIGNED."}), 400
            elif field == "models_written":
                parts = re.split(r"[,/]", val) if isinstance(val, str) else (val or [])
                val = ", ".join(normalize_model_name(str(m).strip()) for m in parts if str(m).strip())
            updates.append(f"{field}=?")
            values.append(val)

    if "assign_week" in data:
        try:
            wk = int(data["assign_week"]) if data["assign_week"] not in (None, "") else None
        except (ValueError, TypeError):
            wk = None
        updates.append("assign_week=?")
        values.append(wk)

    if not updates:
        return jsonify({"error": "No fields to update"}), 400

    values.append(sid)
    db.execute(f"UPDATE new_scripts SET {', '.join(updates)} WHERE id=?", values)
    log_audit(db, "admin.new_script.update", target=f"new_scripts#{sid}", detail=", ".join(sorted(k for k in data)))
    db.commit()
    return jsonify({"status": "updated"})


@app.route("/api/admin/new-scripts/<int:sid>", methods=["DELETE"])
def admin_delete_new_script(sid):
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403
    db = get_db()
    db.execute("DELETE FROM new_scripts WHERE id=?", (sid,))
    log_audit(db, "admin.new_script.delete", target=f"new_scripts#{sid}")
    db.commit()
    return jsonify({"status": "deleted"})


@app.route("/api/admin/new-scripts/bulk", methods=["POST"])
def admin_bulk_new_scripts():
    """Nhap hang loat: nhan { rows: [ {tc_id, member, status, completed_date,
    models_written, sdf_id, remark}, ... ] }. Moi dong duoc validate/chuan hoa qua
    dung ham validate_new_script_row() nhu form nhap don le -> khong lech quy tac."""
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    payload = request.get_json(force=True)
    rows_in = payload.get("rows", [])
    if not rows_in:
        return jsonify({"error": "Không có dòng dữ liệu nào."}), 400

    db = get_db()
    inserted = 0
    errors = []
    for i, raw in enumerate(rows_in):
        row, error = validate_new_script_row(db, raw)
        if error:
            errors.append({"row_index": i, "tc_id": raw.get("tc_id", ""), "error": error})
            continue
        insert_new_script_row(db, row)
        inserted += 1
    log_audit(db, "admin.new_script.bulk", detail=f"inserted={inserted}, errors={len(errors)}")
    db.commit()
    return jsonify({"inserted": inserted, "errors": errors})


# ------------------------------------------------------------------
# Admin: quan ly bang fixes (Daily_Fix_Log). GET/POST/PUT/DELETE tung dong giu
# NGUYEN theo dung ban da chay tren du lieu that (origin/main) - khong doi hanh vi.
# ------------------------------------------------------------------
@app.route("/api/admin/fixes", methods=["GET"])
def admin_get_fixes():
    secret_key = request.args.get("key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    db = get_db()
    limit = int(request.args.get("limit", 1000))
    search = request.args.get("search", "").strip()

    if search:
        rows = db.execute(
            """SELECT * FROM fixes
               WHERE owner LIKE ? OR test_suite LIKE ? OR test_case LIKE ?
                  OR model_fixed LIKE ? OR root_cause LIKE ?
               ORDER BY id DESC LIMIT ?""",
            (f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", limit)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM fixes ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/fixes", methods=["POST"])
def admin_create_fix():
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json(force=True)
    required_fields = ["fix_date", "owner", "test_suite", "test_case", "model_fixed", "fixed_after_cycle"]
    for f in required_fields:
        if not str(data.get(f, "")).strip():
            return jsonify({"error": f"Thiếu trường bắt buộc: {f}"}), 400

    db = get_db()
    root_cause = data.get("root_cause", "").strip()
    group = str(data.get("root_cause_group") or "").strip() or classify_root_cause_group(root_cause)
    db.execute(
        """INSERT INTO fixes (fix_date, owner, test_suite, test_case, model_fixed, fixed_after_cycle, note, root_cause, root_cause_group, sdf_id)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            data["fix_date"], data["owner"].strip(), data["test_suite"].strip(), data["test_case"].strip(),
            normalize_model_name(data["model_fixed"].strip()), int(data["fixed_after_cycle"]),
            data.get("note", "").strip(), root_cause, group, str(data.get("sdf_id") or "").strip(),
        )
    )
    log_audit(db, "admin.fix.create", target=data["test_case"].strip())
    db.commit()
    return jsonify({"status": "created"})


@app.route("/api/admin/fixes/<int:fix_id>", methods=["PUT"])
def admin_update_fix(fix_id):
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json(force=True)
    db = get_db()

    allowed_fields = ["fix_date", "owner", "test_suite", "test_case", "model_fixed", "fixed_after_cycle", "note", "root_cause", "root_cause_group", "sdf_id"]
    updates = []
    values = []
    for field in allowed_fields:
        if field in data:
            val = data[field]
            if field == "fixed_after_cycle":
                val = int(val)
            elif field == "model_fixed":
                val = normalize_model_name(str(val or "").strip())
            updates.append(f"{field}=?")
            values.append(val)

    # Sua root_cause text ma khong gui kem group -> tinh lai group theo text moi.
    if "root_cause" in data and "root_cause_group" not in data:
        updates.append("root_cause_group=?")
        values.append(classify_root_cause_group(str(data["root_cause"] or "")))

    if not updates:
        return jsonify({"error": "No fields to update"}), 400

    values.append(fix_id)
    query = f"UPDATE fixes SET {', '.join(updates)} WHERE id=?"
    db.execute(query, values)
    log_audit(db, "admin.fix.update", target=f"fixes#{fix_id}", detail=", ".join(sorted(k for k in data)))
    db.commit()

    return jsonify({"status": "updated"})


@app.route("/api/admin/fixes/<int:fix_id>", methods=["DELETE"])
def admin_delete_fix(fix_id):
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    db = get_db()
    db.execute("DELETE FROM fixes WHERE id=?", (fix_id,))
    log_audit(db, "admin.fix.delete", target=f"fixes#{fix_id}")
    db.commit()

    return jsonify({"status": "deleted"})


# ------------------------------------------------------------------
# Admin: nhap hang loat Fix Log (tinh nang rieng, khong co ban tuong duong tren
# origin/main - khong dung chung route voi 4 route CRUD don le o tren).
# ------------------------------------------------------------------
def validate_fix_row(db, data):
    """Validate + chuan hoa 1 dong fixes tu dict tho. Tra ve (row_dict, None) hoac
    (None, error_message). CHI dung cho bulk import - khong anh huong CRUD don le."""
    owner = str(data.get("owner") or "").strip()
    test_suite = str(data.get("test_suite") or "").strip()
    test_case = str(data.get("test_case") or "").strip()
    model_fixed = str(data.get("model_fixed") or "").strip()
    root_cause = str(data.get("root_cause") or "").strip()
    fix_date = str(data.get("fix_date") or "").strip()
    fixed_after_cycle = data.get("fixed_after_cycle")

    if not owner:
        return None, "Thiếu Owner."
    if not test_suite or not test_case:
        return None, "Thiếu Test suite/Test case."
    if not model_fixed:
        return None, "Thiếu Model fixed."
    if not root_cause:
        return None, "Root cause là bắt buộc, không được để trống."
    if not fix_date:
        return None, "Thiếu Fix date."
    try:
        fixed_after_cycle = int(fixed_after_cycle)
    except (ValueError, TypeError):
        return None, "Fixed after cycle phải là số nguyên."

    group = str(data.get("root_cause_group") or "").strip() or classify_root_cause_group(root_cause)
    return {
        "fix_date": fix_date, "owner": owner, "test_suite": test_suite, "test_case": test_case,
        "model_fixed": model_fixed, "fixed_after_cycle": fixed_after_cycle,
        "note": str(data.get("note") or "").strip(), "root_cause": root_cause,
        "root_cause_group": group,
    }, None


@app.route("/api/admin/fixes/bulk", methods=["POST"])
def admin_bulk_fixes():
    """Nhap hang loat Fix Log. Cung logic chong trung nhu /api/fixes (POST don le):
    (owner, test_suite, test_case, model_fixed, fixed_after_cycle) trung -> cap nhat
    thay vi tao dong moi, tranh nhan doi trong Theo doi Fix."""
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    payload = request.get_json(force=True)
    rows_in = payload.get("rows", [])
    if not rows_in:
        return jsonify({"error": "Không có dòng dữ liệu nào."}), 400

    db = get_db()
    inserted = 0
    updated = 0
    errors = []
    for i, raw in enumerate(rows_in):
        row, error = validate_fix_row(db, raw)
        if error:
            errors.append({"row_index": i, "test_case": raw.get("test_case", ""), "error": error})
            continue
        db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (row["owner"],))
        existing = db.execute(
            "SELECT id FROM fixes WHERE owner=? AND test_suite=? AND test_case=? AND model_fixed=? AND fixed_after_cycle=?",
            (row["owner"], row["test_suite"], row["test_case"], row["model_fixed"], row["fixed_after_cycle"]),
        ).fetchone()
        if existing:
            db.execute(
                "UPDATE fixes SET fix_date=?, note=?, root_cause=?, root_cause_group=? WHERE id=?",
                (row["fix_date"], row["note"], row["root_cause"], row["root_cause_group"], existing["id"]),
            )
            updated += 1
        else:
            db.execute(
                "INSERT INTO fixes (fix_date, owner, test_suite, test_case, model_fixed, fixed_after_cycle, note, root_cause, root_cause_group) "
                "VALUES (?,?,?,?,?,?,?,?,?)",
                (row["fix_date"], row["owner"], row["test_suite"], row["test_case"], row["model_fixed"],
                 row["fixed_after_cycle"], row["note"], row["root_cause"], row["root_cause_group"]),
            )
            inserted += 1
    log_audit(db, "admin.fix.bulk", detail=f"inserted={inserted}, updated={updated}, errors={len(errors)}")
    db.commit()
    return jsonify({"inserted": inserted, "updated": updated, "errors": errors})


# ------------------------------------------------------------------
# Admin: quan ly bang assignments (dang phu trach) - khoa chinh la (test_suite, test_case),
# khong co cot id rieng.
# ------------------------------------------------------------------
@app.route("/api/admin/assignments", methods=["GET"])
def admin_get_assignments():
    secret_key = request.args.get("key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    db = get_db()
    search = request.args.get("search", "").strip()
    if search:
        rows = db.execute(
            """SELECT rowid, test_suite, test_case, owner, assigned_date FROM assignments
               WHERE test_suite LIKE ? OR test_case LIKE ? OR owner LIKE ?
               ORDER BY assigned_date DESC""",
            (f"%{search}%", f"%{search}%", f"%{search}%")
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT rowid, test_suite, test_case, owner, assigned_date FROM assignments ORDER BY assigned_date DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/assignments/<int:rowid>", methods=["PUT"])
def admin_update_assignment(rowid):
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json(force=True)
    db = get_db()
    row = db.execute("SELECT test_suite, test_case FROM assignments WHERE rowid=?", (rowid,)).fetchone()
    if not row:
        return jsonify({"error": "Không tìm thấy assignment."}), 404

    owner = str(data.get("owner") or "").strip()
    assigned_date = str(data.get("assigned_date") or "").strip()
    if owner:
        db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (owner,))
    db.execute(
        "UPDATE assignments SET owner=?, assigned_date=? WHERE rowid=?",
        (owner, assigned_date, rowid),
    )
    log_audit(db, "admin.assignment.update", target=f"{row['test_suite']}/{row['test_case']}", detail=f"owner={owner}")
    db.commit()
    return jsonify({"status": "updated"})


@app.route("/api/admin/assignments/<int:rowid>", methods=["DELETE"])
def admin_delete_assignment(rowid):
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403
    db = get_db()
    db.execute("DELETE FROM assignments WHERE rowid=?", (rowid,))
    log_audit(db, "admin.assignment.delete", target=f"assignments#{rowid}")
    db.commit()
    return jsonify({"status": "deleted"})


@app.route("/api/admin/assignments/bulk", methods=["POST"])
def admin_bulk_assignments():
    """Nhap hang loat assignment. Khoa (test_suite, test_case) -> trung thi UPDATE
    (upsert), giong dung hanh vi cua POST /api/assignments don le."""
    secret_key = request.headers.get("X-Admin-Key", "")
    if secret_key != ADMIN_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    payload = request.get_json(force=True)
    rows_in = payload.get("rows", [])
    if not rows_in:
        return jsonify({"error": "Không có dòng dữ liệu nào."}), 400

    db = get_db()
    upserted = 0
    errors = []
    for i, raw in enumerate(rows_in):
        test_suite = str(raw.get("test_suite") or "").strip()
        test_case = str(raw.get("test_case") or "").strip()
        owner = str(raw.get("owner") or "").strip()
        assigned_date = str(raw.get("assigned_date") or "").strip() or date.today().isoformat()
        if not test_suite or not test_case:
            errors.append({"row_index": i, "test_case": test_case, "error": "Thiếu Test suite/Test case."})
            continue
        if owner:
            db.execute("INSERT OR IGNORE INTO owners (name, active) VALUES (?, 1)", (owner,))
        db.execute(
            "INSERT INTO assignments (test_suite, test_case, owner, assigned_date) VALUES (?,?,?,?) "
            "ON CONFLICT(test_suite, test_case) DO UPDATE SET owner=excluded.owner, assigned_date=excluded.assigned_date",
            (test_suite, test_case, owner, assigned_date),
        )
        upserted += 1
    log_audit(db, "admin.assignment.bulk", detail=f"upserted={upserted}, errors={len(errors)}")
    db.commit()
    return jsonify({"upserted": upserted, "errors": errors})


# ------------------------------------------------------------------
# Admin: Quan ly tai khoan (users.db) - bao ve bang ADMIN_SECRET_KEY nhu route admin khac.
# ------------------------------------------------------------------
def _admin_key_ok_get():
    return request.args.get("key", "") == ADMIN_SECRET_KEY


def _admin_key_ok_mut():
    return request.headers.get("X-Admin-Key", "") == ADMIN_SECRET_KEY


@app.route("/api/admin/users", methods=["GET"])
def admin_get_users():
    if not _admin_key_ok_get():
        return jsonify({"error": "Unauthorized"}), 403
    udb = get_users_db()
    rows = udb.execute(
        "SELECT username, role, permissions, active, created_at FROM users ORDER BY username"
    ).fetchall()
    out = []
    for r in rows:
        out.append({
            "username": r["username"],
            "role": r["role"],
            "permissions": perms_to_list(r["permissions"]),
            "active": r["active"],
            "created_at": r["created_at"],
        })
    return jsonify({"users": out, "all_tabs": ALL_TABS, "extra_perms": NS_EXTRA_PERMS})


@app.route("/api/admin/users/<path:username>", methods=["PUT"])
def admin_update_user(username):
    if not _admin_key_ok_mut():
        return jsonify({"error": "Unauthorized"}), 403
    data = request.get_json(force=True)
    udb = get_users_db()
    if not udb.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
        return jsonify({"error": "Tài khoản không tồn tại."}), 404
    # Doi role -> dat lai permissions ve mac dinh cua role do (tru khi client gui kem permissions).
    if "role" in data:
        role = data["role"]
        if role not in ROLE_DEFAULT_PERMS:
            return jsonify({"error": "Role không hợp lệ."}), 400
        udb.execute("UPDATE users SET role=? WHERE username=?", (role, username))
        if "permissions" not in data:
            udb.execute("UPDATE users SET permissions=? WHERE username=?",
                        (",".join(ROLE_DEFAULT_PERMS[role]), username))
    if "permissions" in data:
        perms = [p for p in data["permissions"] if p in ALL_PERMS]
        udb.execute("UPDATE users SET permissions=? WHERE username=?",
                    (",".join(perms), username))
    if "active" in data:
        udb.execute("UPDATE users SET active=? WHERE username=?",
                    (1 if data["active"] else 0, username))
    udb.commit()
    db = get_db()
    log_audit(db, "user.update", target=username, detail=", ".join(sorted(k for k in data)))
    db.commit()
    return jsonify({"status": "updated"})


@app.route("/api/admin/users/<path:username>/reset-password", methods=["POST"])
def admin_reset_password(username):
    if not _admin_key_ok_mut():
        return jsonify({"error": "Unauthorized"}), 403
    udb = get_users_db()
    if not udb.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
        return jsonify({"error": "Tài khoản không tồn tại."}), 404
    udb.execute("UPDATE users SET password_hash=? WHERE username=?",
                (generate_password_hash(DEFAULT_RESET_PASSWORD), username))
    udb.commit()
    db = get_db()
    log_audit(db, "user.reset_password", target=username)
    db.commit()
    return jsonify({"status": "reset", "default_password": DEFAULT_RESET_PASSWORD})


@app.route("/api/admin/users/<path:username>", methods=["DELETE"])
def admin_delete_user(username):
    if not _admin_key_ok_mut():
        return jsonify({"error": "Unauthorized"}), 403
    if username == BOOTSTRAP_ADMIN:
        return jsonify({"error": "Không thể xoá tài khoản admin khởi tạo."}), 400
    udb = get_users_db()
    row = udb.execute("SELECT active FROM users WHERE username=?", (username,)).fetchone()
    if not row:
        return jsonify({"error": "Tài khoản không tồn tại."}), 404
    if row["active"] == 1:
        return jsonify({"error": "Chỉ xoá được tài khoản đã ngừng hoạt động. Hãy vô hiệu hoá trước."}), 400
    udb.execute("DELETE FROM users WHERE username=?", (username,))
    udb.commit()
    db = get_db()
    log_audit(db, "user.delete", target=username)
    db.commit()
    return jsonify({"status": "deleted"})


# ------------------------------------------------------------------
# Admin: Audit log viewer (B5) - xem lich su thao tac nguy hiem.
# ------------------------------------------------------------------
@app.route("/api/admin/audit-log", methods=["GET"])
def admin_get_audit_log():
    if not _admin_key_ok_get():
        return jsonify({"error": "Unauthorized"}), 403
    db = get_db()
    limit = min(int(request.args.get("limit", 200)), 2000)
    q = request.args.get("q", "").strip()
    if q:
        like = f"%{q}%"
        rows = db.execute(
            "SELECT id, ts, username, action, target, detail FROM audit_log "
            "WHERE username LIKE ? OR action LIKE ? OR target LIKE ? OR detail LIKE ? "
            "ORDER BY id DESC LIMIT ?",
            (like, like, like, like, limit),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT id, ts, username, action, target, detail FROM audit_log ORDER BY id DESC LIMIT ?",
            (limit,),
        ).fetchall()
    return jsonify([dict(r) for r in rows])


# ------------------------------------------------------------------
# Admin: Quan ly Owner & Team - ADMIN_KEY-gated, dung chung helper owner_op_*.
# ------------------------------------------------------------------
@app.route("/api/admin/owners", methods=["GET"])
def admin_get_owners():
    if not _admin_key_ok_get():
        return jsonify({"error": "Unauthorized"}), 403
    db = get_db()
    rows = db.execute("SELECT name, active, team FROM owners ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/owners", methods=["POST"])
def admin_add_owner():
    if not _admin_key_ok_mut():
        return jsonify({"error": "Unauthorized"}), 403
    data = request.get_json(force=True)
    body, code = owner_op_add(get_db(), (data.get("name") or "").strip(), (data.get("team") or "").strip())
    return jsonify(body), code


@app.route("/api/admin/owners/<path:name>/team", methods=["PUT"])
def admin_set_owner_team(name):
    if not _admin_key_ok_mut():
        return jsonify({"error": "Unauthorized"}), 403
    data = request.get_json(force=True)
    body, code = owner_op_set_team(get_db(), name, (data.get("team") or "").strip())
    return jsonify(body), code


@app.route("/api/admin/owners/<path:old_name>", methods=["PUT", "DELETE"])
def admin_edit_owner(old_name):
    if not _admin_key_ok_mut():
        return jsonify({"error": "Unauthorized"}), 403
    db = get_db()
    if request.method == "DELETE":
        # ?hard=1 -> xoa han (chi khi khong con tham chieu); mac dinh deactivate (giu lich su).
        if request.args.get("hard") == "1":
            body, code = owner_op_hard_delete(db, old_name)
        else:
            body, code = owner_op_deactivate(db, old_name)
        return jsonify(body), code
    data = request.get_json(force=True)
    # PUT: rename hoac reactivate (active=1).
    if data.get("reactivate"):
        db.execute("UPDATE owners SET active=1 WHERE name=?", (old_name,))
        log_audit(db, "owner.reactivate", target=old_name)
        db.commit()
        return jsonify({"status": "reactivated"})
    body, code = owner_op_rename(db, old_name, (data.get("new_name") or "").strip())
    return jsonify(body), code


if __name__ == "__main__":
    init_db()
    init_users_db()
    start_backup_daemon()  # backup tu dong hang ngay (daemon thread, khong can cron)
    print("=" * 60)
    print(" Test Stabilization Tracker dang chay!")
    print(" May nay:        http://localhost:5000")
    print(" May khac trong LAN: http://<IP-may-ban>:5000")
    print("=" * 60)
    # Uu tien waitress (WSGI production-grade, chiu tai tot hon dev server).
    # Neu chua cai waitress -> fallback ve Flask dev server (van chay duoc ngay).
    try:
        from waitress import serve
        print(" Server: waitress (production) | threads=8")
        serve(app, host="0.0.0.0", port=5000, threads=8)
    except ImportError:
        print(" Server: Flask dev (chua cai waitress -> 'pip install -r requirements.txt')")
        app.run(host="0.0.0.0", port=5000, threaded=True, debug=False)
