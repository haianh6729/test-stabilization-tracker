"""
Trich xuat data that tu sheet "(new) UI90 script" trong Script_Tracker_.xlsx
thanh 1 file tab-delimited de dan vao textarea nhap hang loat o trang Admin
(/admin/<key> -> tab "Script viet moi" -> #nsBulkArea).

Khong sua file Excel goc - chi doc.
"""
import openpyxl
from datetime import datetime

SRC = r"D:\Learning\Script_Tracker_.xlsx"
SHEET = "(new) UI90 script"
OUT = r"C:\Users\haian\AppData\Local\Temp\claude\d--Learning-test-stabilization-app-test-stabilization-app\57be0019-0680-4bfa-bff3-e1ed3b39ef43\scratchpad\new_scripts_import.txt"

# Cot "Local": B=team C=item D=tcid E=member F=week G=date H=status I..M=models
MODEL_COLS = {
    9: "SM-F966", 10: "SM-F741", 11: "SM-S721", 12: "SM-X920", 13: "SM-X526",
}


def norm_status(raw):
    s = str(raw or "").strip().upper()
    if s == "DONE":
        return "DONE"
    if s == "SKIP":
        return "SKIP"
    return "ASSIGNED"


def norm_date(raw):
    if not raw:
        return ""
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    return str(raw).strip()[:10]


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[SHEET]

    # Gom theo tc_id, giu ban ghi o row LON HON khi trung (row sau = cap nhat hon).
    by_tcid = {}
    for r in range(4, ws.max_row + 1):
        tc_id = ws.cell(row=r, column=4).value
        if not tc_id:
            continue
        tc_id = str(tc_id).strip()

        member = ws.cell(row=r, column=5).value or ""
        status = norm_status(ws.cell(row=r, column=8).value)
        completed_date = norm_date(ws.cell(row=r, column=7).value)
        remark = ws.cell(row=r, column=20).value or ""

        models = []
        for col, name in MODEL_COLS.items():
            val = ws.cell(row=r, column=col).value
            if val is not None and str(val).strip():
                models.append(name)
        models_str = ", ".join(models)

        by_tcid[tc_id] = {
            "row": r,
            "member": str(member).strip(),
            "status": status,
            "completed_date": completed_date,
            "models": models_str,
            "sdf_id": "",
            "remark": str(remark).strip(),
        }

    lines = []
    for tc_id, d in by_tcid.items():
        cols = [
            tc_id, d["member"], d["status"], d["completed_date"],
            d["models"], d["sdf_id"], d["remark"],
        ]
        lines.append("\t".join(c.replace("\t", " ").replace("\n", " ") for c in cols))

    with open(OUT, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines))

    status_counts = {}
    for d in by_tcid.values():
        status_counts[d["status"]] = status_counts.get(d["status"], 0) + 1

    print(f"Tong so dong (da dedupe theo tc_id): {len(by_tcid)}")
    print(f"Phan bo status: {status_counts}")
    print(f"Da ghi file: {OUT}")


if __name__ == "__main__":
    main()
